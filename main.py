import asyncio
import re
import os
import logging
from datetime import datetime, timedelta

from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.types import BufferedInputFile
import asyncpg
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
    WebAppInfo,
)
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

import json as _json
import aiohttp.web as _web
from database import (
    init_db, get_user, db_run, db_get, db_all, db_insert,
    get_setting, update_setting, add_balance,
)
from texts import t, REGIONS, REGIONS_RU

load_dotenv()
logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────
BOT_TOKEN  = os.getenv("BOT_TOKEN")
CHANNEL_ID = os.getenv("CHANNEL_ID", "@xazdent")
ADMIN_IDS  = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip()]
WEBAPP_URL = os.getenv("WEBAPP_URL", "").rstrip("/")
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))

# E'lon narxlari (ball)
AD_PRICE_TOSHKENT = 200   # Toshkent shahri uchun
AD_PRICE_REGION   = 50    # Boshqa viloyatlar uchun
AD_PRICE_BOTH_AUD = 2     # Ikki auditoriya (clinic+zubtex) ko'paytiruvchi

bot    = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="Markdown"))
dp     = Dispatcher(storage=MemoryStorage())
router = Router()

# ── STATES ────────────────────────────────────────────────────────────────────
class RegState(StatesGroup):
    name   = State()
    phone  = State()
    region = State()
    terms  = State()

class AdminState(StatesGroup):
    setting_input = State()

class AdState(StatesGroup):
    audience   = State()   # Kim: clinic/zubtex/seller
    regions    = State()   # Viloyatlar (checkbox)
    content    = State()   # Matn/rasm/link
    confirm    = State()   # Tasdiqlash

class AdState(StatesGroup):
    """Reklama e'lon berish."""
    audience  = State()   # Kimlar: klinika/zubtex/seller
    regions   = State()   # Viloyatlar
    content   = State()   # Rasm/matn/link
    confirm   = State()   # Tasdiqlash

class NeedState(StatesGroup):
    product  = State()
    qty      = State()
    deadline = State()

class BulkState(StatesGroup):
    items    = State()   # Mahsulotlar ro'yxati (list of dicts)
    deadline = State()

class TopupState(StatesGroup):
    amount  = State()
    receipt = State()

class OfferState(StatesGroup):
    price = State()
    note  = State()

class ShopState(StatesGroup):
    cat  = State()
    name = State()

class MyProductsState(StatesGroup):
    editing = State()   # ro'yxat tahrirlash

# ── KEYBOARDS ─────────────────────────────────────────────────────────────────
def ik(*rows):
    return InlineKeyboardMarkup(inline_keyboard=list(rows))

def ib(text, data=None, url=None, web_app=None):
    if web_app:
        return InlineKeyboardButton(text=text, web_app=web_app)
    if url:
        return InlineKeyboardButton(text=text, url=url)
    return InlineKeyboardButton(text=text, callback_data=data)

def rk(*rows, one_time=False):
    return ReplyKeyboardMarkup(keyboard=list(rows), resize_keyboard=True, one_time_keyboard=one_time)

def kb_lang():
    return ik([ib("🇺🇿 O'zbekcha", "lang_uz"), ib("🇷🇺 Русский", "lang_ru")])

def kb_role(lg):
    return ik(
        [ib("🏥 Vrach / Klinika",  "role_clinic")],
        [ib("🔬 Zubtexnik",         "role_zubtex")],
        [ib("🛒 Sotuvchi",          "role_seller")],
    )

def kb_clinic(lg, uid=0, webapp_url=""):
    """Klinika klaviaturasi — Dental Market WebApp bilan."""
    if webapp_url and uid:
        mkt_url = f"{webapp_url}/catalog?uid={uid}&role=clinic"
        return rk(
            [KeyboardButton(text="🛍 Dental Market",
                           web_app=WebAppInfo(url=mkt_url))],
            [KeyboardButton(text="✏️ Buyurtma yozish"),
             KeyboardButton(text="📩 Takliflar")],
            [KeyboardButton(text="📋 Ehtiyojlarim"),
             KeyboardButton(text="💰 Hisobim")],
            [KeyboardButton(text="📊 Tejash"),
             KeyboardButton(text="⚙️ Profil")],
            [KeyboardButton(text="📖 Yordam")],
        )
    return rk(
        [KeyboardButton(text="🛍 Dental Market")],
        [KeyboardButton(text="✏️ Buyurtma yozish"), KeyboardButton(text="📩 Takliflar")],
        [KeyboardButton(text="📋 Ehtiyojlarim"),    KeyboardButton(text="💰 Hisobim")],
        [KeyboardButton(text="📊 Tejash"),          KeyboardButton(text="⚙️ Profil")],
        [KeyboardButton(text="📖 Yordam")],
    )

def kb_seller(lg, uid=0, webapp_url=""):
    """Sotuvchi klaviaturasi — WebApp tugmalari bilan."""
    if webapp_url and uid:
        mkt_url = f"{webapp_url}/catalog?uid={uid}&role=seller"
        add_url = f"{webapp_url}/catalog?uid={uid}&role=seller&action=add"
        ord_url = f"{webapp_url}/catalog?uid={uid}&role=seller#orders"
        return rk(
            [KeyboardButton(text="🛍 Dental Market",
                           web_app=WebAppInfo(url=mkt_url))],
            [KeyboardButton(text="➕ Mahsulot qo\'shish",
                           web_app=WebAppInfo(url=add_url)),
             KeyboardButton(text="🔔 Buyurtmalar")],
            [KeyboardButton(text="💰 Hisobim"),
             KeyboardButton(text="⚙️ Profil")],
            [KeyboardButton(text="📖 Yordam")],
        )
    # Fallback — WebApp URL yo'q
    return rk(
        [KeyboardButton(text="🛍 Dental Market")],
        [KeyboardButton(text="➕ Mahsulot qo\'shish"),
         KeyboardButton(text="🔔 Buyurtmalar")],
        [KeyboardButton(text="💰 Hisobim"),
         KeyboardButton(text="⚙️ Profil")],
        [KeyboardButton(text="📖 Yordam")],
    )

def kb_regions(lg):
    regs = REGIONS if lg != "ru" else REGIONS_RU
    rows = []
    for i in range(0, len(regs), 2):
        row = [ib(regs[i], f"reg_{i}")]
        if i + 1 < len(regs):
            row.append(ib(regs[i + 1], f"reg_{i+1}"))
        rows.append(row)
    return InlineKeyboardMarkup(inline_keyboard=rows)

def kb_deadline():
    return ik(
        [ib("⚡️ 2 soat", "dl_2"),  ib("🕐 24 soat", "dl_24")],
        [ib("📅 3 kun",  "dl_72"), ib("🗓 1 hafta",  "dl_168")],
    )

def kb_units():
    return ik([ib("📌 Dona", "unit_dona"), ib("⚖️ Kg", "unit_kg"), ib("💧 Litr", "unit_litr")])

def kb_delivery():
    return ik(
        [ib("⚡️ 2 soat", "del_2"),  ib("🕐 24 soat", "del_24")],
        [ib("📅 2 kun",  "del_48"), ib("🗓 1 hafta",  "del_168")],
    )

def kb_confirm():
    return ik([ib("✅ Ha, joylash", "confirm"), ib("❌ Bekor", "cancel")])

def kb_cancel():
    return ik([ib("❌ Bekor", "cancel")])

def kb_shop_cats():
    return ik(
        [ib("🦷 Terapevtik", "cat_1")],
        [ib("⚙️ Jarrohlik & Implant", "cat_2")],
        [ib("🔬 Zubtexnik", "cat_3")],
        [ib("🧪 Dezinfeksiya", "cat_4")],
        [ib("💡 Asbob-uskunalar", "cat_5")],
    )

# ── HELPERS ───────────────────────────────────────────────────────────────────
async def lang(uid):
    u = await get_user(uid)
    return (u["lang"] if u else None) or "uz"

async def has_profile(uid):
    u = await get_user(uid)
    return bool(u and u["clinic_name"] and u["phone"] and u["region"])

async def get_or_create_room(uid):
    """Foydalanuvchining default omborxonasini topadi yoki yaratadi (ko'rinmaydi)."""
    room = await db_get(
        "SELECT * FROM rooms WHERE owner_id=? AND status='active' ORDER BY id LIMIT 1", (uid,)
    )
    if room:
        return room
    rid = await db_insert(
        "INSERT INTO rooms(room_code,room_type,owner_id,max_needs) VALUES(?,?,?,?)",
        (f"AUTO{uid}", "premium", uid, 9999),
    )
    return await db_get("SELECT * FROM rooms WHERE id=?", (rid,))

async def _handle_web_cart(msg: Message, state: FSMContext, u, cart_data: str):
    """Web saytdan savat buyurtmasi. Format: 123x2,456x1"""
    uid = msg.from_user.id

    # Savat ma'lumotlarini parse qilish
    items_raw = []
    try:
        for part in cart_data.split(','):
            if 'x' in part:
                pid, qty = part.split('x', 1)
                items_raw.append({'pid': int(pid), 'qty': int(qty)})
    except Exception:
        items_raw = []

    if not items_raw:
        await msg.answer("⚠️ Savat bo'sh yoki xato. Qayta urinib ko'ring.")
        return

    # Ro'yxatdan o'tmagan bo'lsa
    if not u or not u.get('phone') or not u.get('region'):
        # Savatni state ga saqlaymiz
        await state.update_data(web_cart=cart_data)
        if not u:
            await db_run(
                "INSERT INTO users(id,username,full_name) VALUES(?,?,?) ON CONFLICT(id) DO NOTHING",
                (uid, msg.from_user.username, msg.from_user.full_name)
            )
        lg = 'uz'
        await state.set_state(RegState.phone)
        kb = rk([KeyboardButton(text="📱 Telefon yuborish", request_contact=True)], one_time=True)
        await msg.answer(
            "🛒 *Savatdagi buyurtmangiz tayyor!*\n\n"
            "Buyurtmani rasmiylashtirish uchun\n"
            "telefon raqamingizni yuboring:",
            reply_markup=kb
        )
        return

    # Ro'yxatdan o'tgan — buyurtmani yaratamiz
    await _create_web_cart_order(msg, u, items_raw)

async def _create_web_cart_order(msg, u, items_raw):
    """Web cart dan buyurtma yaratish."""
    uid = msg.from_user.id
    uname = u.get("clinic_name") or u.get("full_name") or str(uid)
    uphone = u.get("phone") or "—"
    uregion = u.get("region") or "—"
    import json as _pj2
    seller_map = {}
    for item in items_raw:
        prod = await db_get(
            "SELECT p.*, s.owner_id as seller_id, s.shop_name "
            "FROM products p JOIN shops s ON p.shop_id=s.id WHERE p.id=?",
            (item["pid"],)
        )
        if not prod: continue
        sid = prod["seller_id"]
        if sid not in seller_map:
            seller_map[sid] = {"shop_name": prod["shop_name"], "items": []}
        seller_map[sid]["items"].append({
            "name": prod["name"], "qty": item["qty"],
            "price": prod["price"], "unit": prod["unit"],
            "product_id": prod["id"]
        })
    if not seller_map:
        await msg.answer("Mahsulotlar topilmadi. Katalog yangilangan bolishi mumkin.")
        return
    sent = 0
    for seller_id, data in seller_map.items():
        items = data["items"]
        total = sum(i["price"] * i["qty"] for i in items)
        lines_list = []
        for i, it in enumerate(items):
            lines_list.append(
                f"{i+1}. *{it['name']}* — {it['qty']} {it['unit']} x "
                f"{fmt_price(it['price'])} = *{fmt_price(it['price']*it['qty'])} som*"
            )
        lines_txt = "\n".join(lines_list)
        order_id = await db_insert(
            "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) VALUES(?,?,?,?)",
            (uid, seller_id, _pj2.dumps(items, ensure_ascii=False), total)
        )
        msg_txt = (
            f"🌐 *Veb-saytdan buyurtma #{order_id}!*\n\n"
            f"📦 *{data['shop_name']}:*\n{lines_txt}\n\n"
            f"💰 *Jami: {fmt_price(total)} som*\n\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🏥 *{uname}*\n"
            f"📞 {uphone}\n"
            f"📍 {uregion}"
        )
        confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Qabul qildim", callback_data=f"co_confirm_{order_id}_{uid}"),
            InlineKeyboardButton(text="❌ Mavjud emas", callback_data=f"co_reject_{order_id}_{uid}")
        ]])
        try:
            await bot.send_message(seller_id, msg_txt, reply_markup=confirm_kb)
            sent += 1
        except Exception as e:
            log.error(f"Web cart seller notify: {e}")
    lg = await lang(uid)
    if not u.get("role") or u["role"] in (None,"none",""):
        await db_run("UPDATE users SET role=? WHERE id=?", ("clinic", uid))
        u2 = await get_user(uid)
    else:
        u2 = u
    kb = kb_clinic(lg) if u2 and u2.get("role") in ("clinic","zubtex") else kb_seller(lg)
    await msg.answer(
        f"✅ *Buyurtmangiz yuborildi!*\n\n"
        f"📦 {sent} ta sotuvchiga xabar ketdi.\n"
        f"Sotuvchilar tez orada boglanadi.",
        reply_markup=kb
    )


def fmt_price(n):
    return f"{int(n):,}".replace(',', ' ')

async def post_to_channel(need_id, need):
    """1 ta ehtiyoj uchun kanal posti (qayta post uchun)."""
    dl_map = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
    dl_txt = dl_map.get(need["deadline_hours"], f"{need['deadline_hours']} soat")
    owner  = await get_user(need["owner_id"])
    words  = need["product_name"].split()
    tags   = " ".join(f"#{w.lower()}" for w in words[:3] if len(w)>2)
    txt = (
        f"📋 *BUYURTMA #{need_id}*\n\n"
        f"🦷 {need['product_name']}\n"
        f"📦 {need['quantity']} {need['unit']}\n"
        f"⏱ {dl_txt} ichida\n\n"
        f"📍 {owner['region'] or ''}\n\n"
        f"{tags}\n💬 @XazdentBot"
    )
    batch_id = need.get("batch_id")
    # Kanalga WebAppInfo ishlamaydi — deep link ishlatamiz
    bot_info = await bot.get_me()
    deep_url = f"https://t.me/{bot_info.username}?start=offer_{need_id}"
    kb = ik([ib("📤 Taklif yuborish", url=deep_url)])
    try:
        m = await bot.send_message(CHANNEL_ID, txt, reply_markup=kb)
        return m.message_id
    except Exception as e:
        log.error(f"❌ Kanal xato: {e}")
        return None

async def post_batch_to_channel(batch_id, needs_list, owner, photo_file_id=None):
    """Ko'p ehtiyoj uchun BITTA paket post."""
    if not needs_list:
        return None
    dl_map = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
    dl_txt = dl_map.get(needs_list[0]["deadline_hours"], "?")
    lines  = "\n".join([
        f"• {n['product_name']} — {n['quantity']} {n['unit']}"
        for n in needs_list[:15]
    ])
    if len(needs_list) > 15:
        lines += f"\n• ...va yana {len(needs_list)-15} ta"
    all_words = " ".join(n["product_name"] for n in needs_list[:5]).split()
    tags = " ".join(f"#{w.lower()}" for w in dict.fromkeys(all_words) if len(w)>2)[:80]
    # To'lov turlari
    pay_icons = {"p2p":"💳 P2P","cash":"💵 Naqd","bank":"🏦 Hisob raqam"}
    pm_raw = needs_list[0].get("payment_methods","") if needs_list else ""
    pm_txt = " · ".join(pay_icons[p] for p in (pm_raw or "").split(",") if p in pay_icons)
    pm_line = f"\n💳 {pm_txt}" if pm_txt else ""
    txt = (
        f"📋 *BUYURTMA #{batch_id}* — {len(needs_list)} ta mahsulot\n\n"
        f"{lines}\n\n"
        f"📍 {owner.get('region') or ''}\n"
        f"⏱ {dl_txt} ichida{pm_line}\n\n"
        f"{tags}\n💬 @XazdentBot"
    )
    # Kanalga WebAppInfo yuborib bo'lmaydi — oddiy URL link ishlatamiz
    bot_info = await bot.get_me()
    if WEBAPP_URL:
        deep_url = f"https://t.me/{bot_info.username}?start=batch_{batch_id}"
    else:
        deep_url = f"https://t.me/{bot_info.username}?start=batch_{batch_id}"
    kb = ik([ib("📤 Taklif yuborish", url=deep_url)])
    try:
        if photo_file_id:
            m = await bot.send_photo(CHANNEL_ID, photo_file_id,
                                     caption=txt, reply_markup=kb)
        else:
            m = await bot.send_message(CHANNEL_ID, txt, reply_markup=kb)
        log.info(f"✅ Batch kanal post: batch={batch_id} msg={m.message_id}")
        return m.message_id
    except Exception as e:
        log.error(f"❌ Batch kanal xato: {e}")
        return None

async def notify_sellers(need_id, need, owner):
    """Barcha sotuvchilarga lichkada xabar."""
    sellers = await db_all(
        "SELECT id FROM users WHERE role='seller' AND id!=? AND is_blocked=0",
        (owner["id"],),
    )
    if WEBAPP_URL and need.get("batch_id"):
        url = f"{WEBAPP_URL}/offer/{need['batch_id']}"
        kb = ik(
            [ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))],
            [ib("⏭ Keyinroq", "skip_notify")],
        )
    else:
        bot_info = await bot.get_me()
        deep_url = f"https://t.me/{bot_info.username}?start=offer_{need_id}"
        kb = ik(
            [ib("📤 Taklif yuborish", url=deep_url)],
            [ib("⏭ Keyinroq", "skip_notify")],
        )

    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    txt = (
        f"📦 *Yangi buyurtma!*\n\n"
        f"🦷 {need['product_name']}\n"
        f"📦 {need['quantity']} {need['unit']}\n"
        f"⏱ {dl_map.get(need['deadline_hours'], '?')} ichida\n"
        f"📍 {owner['region'] or ''}"
    )
    sent = 0
    for s in sellers:
        try:
            await bot.send_message(s["id"], txt, reply_markup=kb)
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass
    log.info(f"Notify: {sent}/{len(sellers)} sotuvchiga yuborildi")

# ── /start ─────────────────────────────────────────────────────────────────────
@router.message(CommandStart())
async def cmd_start(msg: Message, state: FSMContext):
    await state.clear()
    uid = msg.from_user.id
    u = await get_user(uid)
    if not u:
        await db_run(
            "INSERT INTO users(id,username,full_name) VALUES(?,?,?) ON CONFLICT(id) DO NOTHING",
            (uid, msg.from_user.username, msg.from_user.full_name),
        )
        u = await get_user(uid)

    # Deep link: /start offer_42 yoki /start batch_5
    args = msg.text.split(maxsplit=1)[1] if " " in (msg.text or "") else ""
    # Mahsulot deep link: /start p_123 yoki /start xd_XD00005
    if args.startswith("xz_") or args.startswith("xd_"):
        try:
            code = args[3:].upper()
            prod = await db_get(
                "SELECT id FROM products WHERE UPPER(article_code)=? AND is_active=1",
                (code,)
            )
            if prod:
                if u and u["role"] not in (None, "none", ""):
                    await _show_product_start(msg, prod["id"])
                    return
                else:
                    await state.update_data(pending_product_id=prod["id"])
        except Exception as e:
            log.error(f"xd_ deep link xato: {e}")

    if args.startswith("p_"):
        try:
            pid = int(args[2:])
            if u and u["role"] not in (None, "none", ""):
                # Ro'yxatdan o'tgan — mahsulotni ko'rsatamiz
                await _show_product_start(msg, pid)
                return
            else:
                # Ro'yxatdan o'tmagan — eslab qolamiz, keyin ko'rsatamiz
                await state.update_data(pending_product_id=pid)
                # Ro'yxatdan o'tishni boshlaymiz
                await msg.answer(
                    "🦷 *XazDent Dental Market*\n\n"
                    "Mahsulotni ko\'rish uchun avval\n"
                    "ro\'yxatdan o\'ting — 30 soniya!",
                )
                # Standart registration flow
        except Exception as e:
            log.error(f"p_ deep link xato: {e}")

    # Web saytdan savat buyurtmasi
    if args.startswith("web_cart_"):
        cart_data = args[9:]  # "web_cart_" dan keyin
        await _handle_web_cart(msg, state, u, cart_data)
        return

    if args.startswith("offer_") and u and u["role"] in ("seller",):
        try:
            nid = int(args.split("_")[1])
            await _start_offer_bot(msg, state, nid)
            return
        except Exception:
            pass
    if args.startswith("batch_") and u and u["role"] in ("seller",):
        try:
            batch_id = int(args.split("_")[1])
            uid = msg.from_user.id
            if WEBAPP_URL:
                url = f"{WEBAPP_URL}/offer/{batch_id}"
                await msg.answer(
                    "💰 *Narx kiriting:*",
                    reply_markup=ik([ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))])
                )
            else:
                needs = await db_all(
                    "SELECT * FROM needs WHERE batch_id=? AND status='active'", (batch_id,)
                )
                if needs:
                    await _start_offer_bot(msg, state, needs[0]["id"])
            return
        except Exception as e:
            log.error(f"batch deep link xato: {e}")

    if u and u["role"] not in (None, "none", ""):
        # Pending product — ulashilgan mahsulot
        d = await state.get_data()
        pending_pid = d.get("pending_product_id")
        if pending_pid:
            await state.update_data(pending_product_id=None)
            await _show_product_start(msg, pending_pid)
            return

        lg  = u["lang"] or "uz"
        kb   = kb_clinic(lg, uid=uid, webapp_url=WEBAPP_URL) if u["role"] in ("clinic", "zubtex") else kb_seller(lg, uid=uid, webapp_url=WEBAPP_URL)
        role = u["role"]
        if role in ("clinic", "zubtex"):
            txt = "🏥 *Klinika paneli*"
            # Dental Market inline tugmasi
            if WEBAPP_URL:
                mkt_url = f"{WEBAPP_URL}/catalog?uid={uid}&role=clinic"
                inline_kb = ik([ib("🛍 Dental Market — Online do\'kon →", web_app=WebAppInfo(url=mkt_url))])
                await msg.answer(txt, reply_markup=kb)
                await msg.answer(
                    "🛍 *Dental Market*\n_Stomatologik materiallar online do\'koni_",
                    reply_markup=inline_kb
                )
            else:
                await msg.answer(txt, reply_markup=kb)
        else:
            shop = await db_get("SELECT * FROM shops WHERE owner_id=?", (uid,))
            if not shop and u:
                sname0 = u.get("clinic_name") or u.get("full_name") or "Do\'konim"
                new_sid = await db_insert(
                    "INSERT INTO shops(owner_id,shop_name,category,phone,region,status) "
                    "VALUES(?,?,?,?,?,'active')",
                    (uid, sname0, "Stomatologiya", u.get("phone",""), u.get("region",""))
                )
                shop = await db_get("SELECT * FROM shops WHERE id=?", (new_sid,))
            avg = await db_get("SELECT AVG(rating) as a FROM reviews WHERE seller_id=?", (uid,))
            prod_count = 0
            if shop:
                pc = await db_get(
                    "SELECT COUNT(*) as c FROM products WHERE shop_id=? AND is_active=1",
                    (shop["id"],))
                prod_count = pc["c"] if pc else 0
            deal_count = shop["total_deals"] if shop else 0
            rating = float(avg["a"] or 0) if avg and avg["a"] else 0
            stars = ("⭐ %.1f" % rating) if rating > 0 else "⭐ Yangi do\'kon"
            shop_name = shop["shop_name"] if shop else "Do\'konim"
            region = shop["region"] if shop else (u.get("region") if u else "")
            txt = (
                f"🏪 *{shop_name}*\n"
                f"📍 {region} · {stars}\n"
                f"📦 {prod_count} ta mahsulot · 🤝 {deal_count} ta bitim"
            )
            if WEBAPP_URL:
                mkt_url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller"
                add_url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller&action=add"
                inline_kb = ik(
                    [ib("🛍 Dental Market →", web_app=WebAppInfo(url=mkt_url))],
                    [ib("➕ Mahsulot qo\'shish →", web_app=WebAppInfo(url=add_url))],
                )
                await msg.answer(txt, reply_markup=kb)
                await msg.answer("👆 Do\'koningizni boshqaring:", reply_markup=inline_kb)
            else:
                await msg.answer(txt, reply_markup=kb)
        return

    await msg.answer(t("uz", "welcome"), reply_markup=kb_lang())

@router.callback_query(F.data == "terms_accept", RegState.terms)
async def terms_accept(call: CallbackQuery, state: FSMContext):
    d   = await state.get_data()
    uid = call.from_user.id
    # Foydalanuvchi ma'lumotlarini saqlaymiz
    clinic_name = d.get("clinic_name", "")
    phone       = d.get("phone", "")
    region      = d.get("region", "")
    full_name   = d.get("full_name", "") or call.from_user.full_name or ""
    await db_run(
        "UPDATE users SET clinic_name=?, phone=?, region=?, full_name=? WHERE id=?",
        (clinic_name, phone, region, full_name, uid)
    )
    await call.message.edit_reply_markup(reply_markup=None)
    await _finish_reg(call.message, state)
    await call.answer()

@router.callback_query(F.data == "skip_notify")
async def skip_notify(call: CallbackQuery):
    try:
        await call.message.delete()
    except Exception:
        pass
    await call.answer("Keyinroq ko'rasiz")

# ── TIL ────────────────────────────────────────────────────────────────────────
@router.callback_query(F.data.startswith("lang_"))
async def cb_lang(call: CallbackQuery):
    lg = call.data[5:]
    await db_run("UPDATE users SET lang=? WHERE id=?", (lg, call.from_user.id))
    await call.message.edit_text(t(lg, "welcome"), reply_markup=kb_role(lg))
    await call.answer()

@router.callback_query(F.data.startswith("role_"))
async def cb_role(call: CallbackQuery, state: FSMContext):
    role = call.data[5:]
    uid  = call.from_user.id
    # zubtex ham clinic kabi saqlaymiz
    await db_run("UPDATE users SET role=? WHERE id=?", (role, uid))
    u  = await get_user(uid)
    lg = u["lang"] or "uz"
    await state.set_state(RegState.name)
    if role == "clinic":
        ask = "🏥 Klinika nomini kiriting:\n\n_Masalan: Sadaf Dental_"
    elif role == "zubtex":
        ask = "🔬 Zubtexnik ismingiz va ish joyingizni kiriting:"
    else:
        ask = "🏪 Do\'kon / kompaniya nomingizni kiriting:"
    await call.message.answer(ask, reply_markup=ReplyKeyboardRemove())
    await call.answer()

# ── RO'YXATDAN O'TISH ─────────────────────────────────────────────────────────
@router.message(RegState.name)
async def reg_name(msg: Message, state: FSMContext):
    name = (msg.text or "").strip()
    if len(name) < 2:
        await msg.answer("⚠️ Kamida 2 ta harf kiriting.")
        return
    await state.update_data(clinic_name=name)
    lg = await lang(msg.from_user.id)
    kb = rk([KeyboardButton(text=t(lg, "btn_send_phone"), request_contact=True)], one_time=True)
    await state.set_state(RegState.phone)
    await msg.answer(t(lg, "ask_phone"), reply_markup=kb)

@router.message(RegState.phone, F.contact)
async def reg_phone(msg: Message, state: FSMContext):
    await state.update_data(
        phone=msg.contact.phone_number,
        full_name=msg.contact.first_name or msg.from_user.full_name or ""
    )
    lg = await lang(msg.from_user.id)
    await state.set_state(RegState.region)
    await msg.answer(t(lg, "ask_region"), reply_markup=kb_regions(lg))

@router.callback_query(F.data.startswith("reg_"), RegState.region)
async def reg_region(call: CallbackQuery, state: FSMContext):
    lg   = await lang(call.from_user.id)
    idx  = int(call.data[4:])
    regs = REGIONS if lg != "ru" else REGIONS_RU
    reg  = regs[idx].split(" ", 1)[1] if " " in regs[idx] else regs[idx]
    await state.update_data(region=reg)
    await state.set_state(RegState.terms)
    u = await get_user(call.from_user.id)
    role = u["role"] if u else "clinic"
    # Foydalanish shartlari + obuna xabari
    if role == "seller":
        terms_txt = (
            "📋 *Foydalanish shartlari*\n\n"
            "XazDent B2B dental platformasiga xush kelibsiz!\n\n"
            "✅ *Bepul davr:* 1 Avgustgacha to'liq bepul\n"
            "💰 *Oylik obuna:* miqdori haqida keyinroq "
            "xabardor qilinasiz\n\n"
            "📌 *Qoidalar:*\n"
            "• Faqat stomatologik mahsulotlar joylash mumkin\n"
            "• Narxlar so'mda ko'rsatilsin\n"
            "• Soxta mahsulot va ma'lumot joylash taqiqlanadi\n\n"
            "Davom etish uchun shartlarni qabul qiling:"
        )
    else:
        terms_txt = (
            "📋 *XazDent*ga xush kelibsiz!\n\n"
            "✅ Dental materiallar eng qulay narxda\n"
            "✅ 100+ ishonchli sotuvchi\n"
            "✅ Butun O'zbekiston bo'ylab yetkazib berish\n\n"
            "Davom etish uchun qabul qiling:"
        )
    await call.message.answer(
        terms_txt,
        reply_markup=ik([ib("✅ Qabul qilaman — Davom etish", "terms_accept")])
    )
    await call.answer()

# addr handler olib tashlandi — manzil so'ralmaydi

# location handler olib tashlandi

async def _finish_reg(msg: Message, state: FSMContext):
    uid = msg.from_user.id
    await state.clear()
    lg = await lang(uid)
    u  = await get_user(uid)
    if u and u["role"] in ("clinic", "zubtex"):
        kb    = kb_clinic(lg, uid=uid, webapp_url=WEBAPP_URL)
        txt   = (
            f"✅ *Ro\'yxatdan o\'tdingiz!*\n\n"
            f"🏥 {u.get('clinic_name','')}\n"
            f"📍 {u.get('region','')}\n\n"
            f"🛍 Dental Market orqali materiallar buyurtma qiling!"
        )
        await msg.answer(txt, reply_markup=kb)
    else:
        # Do'kon yo'q bo'lsa avtomatik yaratamiz
        shop = await db_get("SELECT id FROM shops WHERE owner_id=?", (uid,))
        if not shop and u:
            sname = u.get("clinic_name") or u.get("full_name") or "Do\'konim"
            await db_insert(
                "INSERT INTO shops(owner_id,shop_name,category,phone,region,status) "
                "VALUES(?,?,?,?,?,'active')",
                (uid, sname, "Stomatologiya", u.get("phone",""), u.get("region",""))
            )
        kb  = kb_seller(lg, uid=uid, webapp_url=WEBAPP_URL)
        txt = (
            f"✅ *Do\'koningiz ochildi!*\n\n"
            f"🏪 {u.get('clinic_name','')}\n"
            f"📍 {u.get('region','')}\n\n"
            f"➕ Mahsulot qo\'shing va xaridorlar sizni topsın!"
        )
        await msg.answer(txt, reply_markup=kb)
        # Adminga xabar
        for aid in ADMIN_IDS:
            try:
                await bot.send_message(
                    aid,
                    f"🆕 *Yangi sotuvchi!*\n\n"
                    f"🏪 {u.get('clinic_name','')}\n"
                    f"👤 {u.get('full_name','')}\n"
                    f"📞 {u.get('phone','')}\n"
                    f"📍 {u.get('region','')}\n"
                    f"🆔 ID: `{uid}`"
                )
            except Exception:
                pass

# ── PROFIL ─────────────────────────────────────────────────────────────────────
@router.message(F.text == "⚙️ Profil")
async def show_profile(msg: Message, state: FSMContext):
    u  = await get_user(msg.from_user.id)
    lg = u["lang"] or "uz"
    txt = (
        f"⚙️ *Profil*\n\n"
        f"👤 {u['clinic_name'] or '—'}\n"
        f"📞 {u['phone'] or '—'}\n"
        f"📍 {u['region'] or '—'}\n"
        f"🏠 {u['address'] or '—'}\n"
        f"💰 Balans: {u['balance'] or 0:.1f} ball"
    )
    role_label = {"clinic":"🏥 Klinika","zubtex":"🔬 Zubtexnik","seller":"🛒 Sotuvchi"}.get(u["role"],"—")
    txt += f"\n👤 Rol: {role_label}"
    await msg.answer(txt, reply_markup=ik(
        [ib("✏️ Tahrirlash", "edit_profile")],
        [ib("🔄 Rolni o'zgartirish", "change_role")],
        [ib("📢 E'lon berish", "ad_start")],
    ))

def _payment_kb(selected: list) -> InlineKeyboardMarkup:
    opts = [("p2p","💳 P2P (karta)"), ("bank","🏦 Hisob raqam"), ("cash","💵 Naqd pul")]
    rows = []
    for key, label in opts:
        chk = "✅ " if key in selected else "☐ "
        rows.append([ib(f"{chk}{label}", f"pm_tog_{key}")])
    rows.append([ib("💾 Saqlash", "pm_save"), ib("◀️ Bekor", "pm_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "set_payment")
async def set_payment(call: CallbackQuery, state: FSMContext):
    u   = await get_user(call.from_user.id)
    cur = (u.get("payment_methods") or "").split(",") if u else []
    cur = [x for x in cur if x]
    await state.update_data(pm_selected=cur)
    await call.message.answer(
        "💳 *To\'lov usullarini tanlang:*\n_(Bir nechta tanlash mumkin)_",
        reply_markup=_payment_kb(cur)
    )
    await call.answer()

@router.callback_query(F.data.startswith("pm_tog_"))
async def pm_toggle(call: CallbackQuery, state: FSMContext):
    key = call.data[7:]
    d   = await state.get_data()
    sel = list(d.get("pm_selected", []))
    if key in sel: sel.remove(key)
    else: sel.append(key)
    await state.update_data(pm_selected=sel)
    await call.message.edit_reply_markup(reply_markup=_payment_kb(sel))
    await call.answer()

@router.callback_query(F.data == "pm_save")
async def pm_save(call: CallbackQuery, state: FSMContext):
    d   = await state.get_data()
    sel = d.get("pm_selected", [])
    val = ",".join(sel)
    await db_run("UPDATE users SET payment_methods=? WHERE id=?", (val or None, call.from_user.id))
    await state.clear()
    pay_icons = {"p2p":"💳 P2P","cash":"💵 Naqd","bank":"🏦 Hisob raqam"}
    pm_txt = " · ".join(pay_icons[p] for p in sel if p in pay_icons) or "Belgilanmagan"
    await call.message.edit_text(f"✅ Saqlandi: *{pm_txt}*")
    await call.answer("✅")

@router.callback_query(F.data == "pm_cancel")
async def pm_cancel(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.delete()
    await call.answer()

@router.callback_query(F.data == "change_role")
async def change_role(call: CallbackQuery):
    lg = await lang(call.from_user.id)
    await call.message.answer(
        "🔄 *Yangi rolingizni tanlang:*",
        reply_markup=ik(
            [ib("🏥 Vrach / Klinika", "role_clinic")],
            [ib("🔬 Zubtexnik",        "role_zubtex")],
            [ib("🛒 Sotuvchi",          "role_seller")],
            [ib("◀️ Orqaga",            "back_profile")],
        )
    )
    await call.answer()

@router.callback_query(F.data == "back_profile")
async def back_profile(call: CallbackQuery):
    await call.message.delete()
    await call.answer()

@router.callback_query(F.data == "edit_profile")
async def edit_profile(call: CallbackQuery, state: FSMContext):
    lg = await lang(call.from_user.id)
    u  = await get_user(call.from_user.id)
    await state.set_state(RegState.name)
    if u and u["role"] == "seller":
        ask = "🏪 Do'kon/kompaniya nomingizni kiriting:"
    elif u and u["role"] == "zubtex":
        ask = "🔬 Ismingiz va ish joyingizni kiriting:"
    else:
        ask = "🏥 Klinika nomingizni kiriting:"
    await call.message.answer(ask, reply_markup=ReplyKeyboardRemove())
    await call.answer()

# ── EHTIYOJ YOZISH (3 savol) ──────────────────────────────────────────────────
@router.message(F.text == "✏️ Ehtiyoj yozish")
async def need_start(msg: Message, state: FSMContext):
    uid = msg.from_user.id
    if not await has_profile(uid):
        await msg.answer(
            "⚠️ Avval profilingizni to'ldiring!",
            reply_markup=ik([ib("✏️ Profilni to'ldirish", "edit_profile")]),
        )
        return
    await msg.answer(
        "📋 *Ehtiyoj yozish*\n\nQancha mahsulot kerak?",
        reply_markup=ik(
            [ib("1️⃣ 1 ta mahsulot", "need_single")],
            [ib("📋 Ko'p mahsulot (Mini App)", "need_bulk")],
        ),
    )

@router.callback_query(F.data == "need_single")
async def need_single(call: CallbackQuery, state: FSMContext):
    await state.set_state(NeedState.product)
    await call.message.answer(
        "🦷 *Nima kerak?*\n\n_Masalan: Xarizma plomba A2_",
        reply_markup=ReplyKeyboardRemove(),
    )
    await call.answer()

@router.callback_query(F.data == "need_bulk")
async def need_bulk(call: CallbackQuery, state: FSMContext):
    webapp_url = WEBAPP_URL
    if not webapp_url:
        # WEBAPP_URL yo'q — matn orqali kiritish
        await state.set_state(BulkState.items)
        await call.message.answer(
            "📋 *Ko'p mahsulot*\n\n"
            "Har bir mahsulotni yangi qatorda yozing:\n\n"
            "```\n5 dona Xarizma A2\n2 kg GC Fuji IX\n1 dona Endomotor\n```\n\n"
            "_Format: miqdor + birlik + nom_",
            reply_markup=ReplyKeyboardRemove(),
        )
    else:
        uid  = call.from_user.id
        u    = await get_user(uid)
        name = (u["clinic_name"] or u["full_name"] or "Klinika").replace(" ", "+")
        url  = f"{webapp_url}/order?clinic={name}&uid={uid}"
        await call.message.answer(
            "📋 *Ko'p mahsulot buyurtmasi*\n\nQuyidagi tugmani bosing:",
            reply_markup=ik([ib("📋 Buyurtma yaratish →", web_app=WebAppInfo(url=url))]),
        )
    await call.answer()

@router.message(NeedState.product)
async def need_product(msg: Message, state: FSMContext):
    await state.update_data(product=msg.text)
    await state.set_state(NeedState.qty)
    await msg.answer(
        "📦 *Miqdori?*\n\n_Masalan: 5 dona, 2 kg, 1 litr_\n\nYoki tezkor:",
        reply_markup=ik(
            [ib("1 dona", "q_1_dona"), ib("2 dona", "q_2_dona"), ib("5 dona", "q_5_dona")],
            [ib("1 kg",   "q_1_kg"),  ib("2 kg",   "q_2_kg"),  ib("10 kg",  "q_10_kg")],
        ),
    )

@router.callback_query(F.data.startswith("q_"), NeedState.qty)
async def need_qty_btn(call: CallbackQuery, state: FSMContext):
    parts = call.data[2:].split("_")
    qty, unit = float(parts[0]), parts[1]
    await state.update_data(qty=qty, unit=unit)
    await state.set_state(NeedState.deadline)
    await call.message.answer("⏱ *Qachongacha kerak?*", reply_markup=kb_deadline())
    await call.answer()

@router.message(NeedState.qty)
async def need_qty_text(msg: Message, state: FSMContext):
    text  = msg.text.strip()
    parts = text.split(None, 1)
    try:
        qty  = float(parts[0].replace(",", "."))
        unit = parts[1].lower() if len(parts) > 1 else None
    except Exception:
        await msg.answer("❌ Noto'g'ri format. _Masalan: 5 dona yoki 2 kg_")
        return
    if unit:
        await state.update_data(qty=qty, unit=unit)
        await state.set_state(NeedState.deadline)
        await msg.answer("⏱ *Qachongacha kerak?*", reply_markup=kb_deadline())
    else:
        await state.update_data(qty=qty)
        await msg.answer("⚖️ *O'lchov birligi:*", reply_markup=kb_units())

@router.callback_query(F.data.startswith("unit_"), NeedState.qty)
async def need_unit(call: CallbackQuery, state: FSMContext):
    await state.update_data(unit=call.data[5:])
    await state.set_state(NeedState.deadline)
    await call.message.answer("⏱ *Qachongacha kerak?*", reply_markup=kb_deadline())
    await call.answer()

@router.callback_query(F.data.startswith("dl_"), NeedState.deadline)
async def need_deadline(call: CallbackQuery, state: FSMContext):
    dl = int(call.data[3:])
    await state.update_data(deadline=dl)
    d = await state.get_data()
    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    preview = (
        f"🦷 *{d['product']}*\n"
        f"📦 {d['qty']} {d.get('unit','dona')}\n"
        f"⏱ {dl_map.get(dl, str(dl)+' soat')} ichida"
    )
    await call.message.answer(
        f"✅ *Tekshiring:*\n\n{preview}\n\nKanalga joylashtirilamizmi?",
        reply_markup=kb_confirm(),
    )
    await call.answer()

@router.callback_query(F.data == "confirm", NeedState.deadline)
async def need_confirm(call: CallbackQuery, state: FSMContext):
    d   = await state.get_data()
    uid = call.from_user.id
    u   = await get_user(uid)

    # Auto room
    room    = await get_or_create_room(uid)
    expires = (datetime.now() + timedelta(hours=d["deadline"])).isoformat()

    # Batch yaratish (jadval uchun)
    batch_id = await db_insert(
        "INSERT INTO batches(owner_id,deadline_hours,expires_at) VALUES(?,?,?)",
        (uid, d["deadline"], expires),
    )

    need_id = await db_insert(
        "INSERT INTO needs(batch_id,room_id,owner_id,product_name,quantity,unit,deadline_hours,expires_at) "
        "VALUES(?,?,?,?,?,?,?,?)",
        (batch_id, room["id"], uid, d["product"], d["qty"], d.get("unit", "dona"), d["deadline"], expires),
    )

    # Batch ga need bog'lash
    await db_run("UPDATE batches SET status='active' WHERE id=?", (batch_id,))

    need = await db_get("SELECT * FROM needs WHERE id=?", (need_id,))
    mid  = await post_to_channel(need_id, dict(need))
    if mid:
        await db_run("UPDATE needs SET channel_message_id=? WHERE id=?", (mid, need_id))

    await state.clear()
    # Kanalga link
    product = d["product"]
    qty     = d["qty"]
    unit    = d.get("unit", "dona")
    if mid and isinstance(CHANNEL_ID, str):
        chan = CHANNEL_ID.lstrip("@")
        link = f"\n[Kanalda ko'rish](https://t.me/{chan}/{mid})"
    else:
        link = ""

    await call.message.edit_text(
        f"✅ *E'lon joylashtirildi!*\n\n"
        f"🦷 {product}\n"
        f"📦 {qty} {unit}"
        f"{link}",
    )
    await call.answer("✅")

    # Sotuvchilarga xabar (fon da)
    asyncio.create_task(notify_sellers(need_id, dict(need), dict(u)))

# ── BULK (matn orqali, WEBAPP_URL yo'q bo'lsa) ──────────────────────────────
@router.message(BulkState.items)
async def bulk_items(msg: Message, state: FSMContext):
    lines  = [l.strip() for l in msg.text.strip().split("\n") if l.strip()]
    parsed = []
    errors = []
    for line in lines:
        parts = line.split(None, 2)
        if len(parts) >= 3:
            try:
                qty  = float(parts[0].replace(",", "."))
                unit = parts[1]
                name = parts[2]
                parsed.append({"qty": qty, "unit": unit, "name": name})
            except Exception:
                errors.append(line)
        elif len(parts) == 2:
            try:
                qty  = float(parts[0].replace(",", "."))
                name = parts[1]
                parsed.append({"qty": qty, "unit": "dona", "name": name})
            except Exception:
                errors.append(line)
        else:
            errors.append(line)
    if not parsed:
        await msg.answer("❌ Format xato. Masalan:\n`5 dona Xarizma A2`")
        return
    await state.update_data(bulk_items=parsed)
    await state.set_state(BulkState.deadline)
    preview = "\n".join([f"• {p['qty']} {p['unit']} — {p['name']}" for p in parsed])
    err_txt = f"\n\n⚠️ Qabul qilinmadi: {', '.join(errors)}" if errors else ""
    await msg.answer(
        f"📋 *{len(parsed)} ta mahsulot:*\n\n{preview}{err_txt}\n\n⏱ Qachongacha kerak?",
        reply_markup=kb_deadline(),
    )

@router.callback_query(F.data.startswith("dl_"), BulkState.deadline)
async def bulk_deadline(call: CallbackQuery, state: FSMContext):
    dl = int(call.data[3:])
    await state.update_data(deadline=dl)
    d     = await state.get_data()
    items = d["bulk_items"]
    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    preview = "\n".join([f"• {p['qty']} {p['unit']} — {p['name']}" for p in items])
    await call.message.answer(
        f"✅ *Tasdiqlang:*\n\n{preview}\n\n⏱ {dl_map.get(dl, str(dl)+' soat')} ichida\n\n"
        f"*{len(items)} ta e'lon joylashtiriladi*",
        reply_markup=kb_confirm(),
    )
    await call.answer()

@router.callback_query(F.data == "confirm", BulkState.deadline)
async def bulk_confirm(call: CallbackQuery, state: FSMContext):
    d       = await state.get_data()
    items   = d["bulk_items"]
    uid     = call.from_user.id
    u       = await get_user(uid)
    room    = await get_or_create_room(uid)
    expires = (datetime.now() + timedelta(hours=d["deadline"])).isoformat()

    batch_id = await db_insert(
        "INSERT INTO batches(owner_id,deadline_hours,expires_at) VALUES(?,?,?)",
        (uid, d["deadline"], expires),
    )

    saved_needs = []
    for item in items:
        nid = await db_insert(
            "INSERT INTO needs(batch_id,room_id,owner_id,product_name,quantity,unit,deadline_hours,expires_at) "
            "VALUES(?,?,?,?,?,?,?,?)",
            (batch_id, room["id"], uid, item["name"], item["qty"], item["unit"], d["deadline"], expires),
        )
        need = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
        saved_needs.append(dict(need))

    # Bitta paket post
    mid = await post_batch_to_channel(batch_id, saved_needs, dict(u))
    if mid:
        for n in saved_needs:
            await db_run("UPDATE needs SET channel_message_id=? WHERE id=?", (mid, n["id"]))

    asyncio.create_task(notify_sellers_batch(batch_id, uid))
    await state.clear()

    chan = CHANNEL_ID.lstrip("@") if isinstance(CHANNEL_ID, str) else str(CHANNEL_ID)
    link = f"\n[Kanalda ko'rish](https://t.me/{chan}/{mid})" if mid else ""
    await call.message.edit_text(
        f"✅ *{len(saved_needs)} ta mahsulot joylashtirildi!*{link}\n\n"
        f"Sotuvchilar ko'rib taklif yuboradi.\n"
        f"📋 *Ehtiyojlarim* bo'limida kuzating."
    )
    await call.answer("✅")

async def notify_sellers_batch(batch_id: int, owner_id: int):
    """Batch dagi barcha ehtiyojlar haqida sotuvchilarga xabar."""
    needs   = await db_all("SELECT * FROM needs WHERE batch_id=?", (batch_id,))
    owner   = await get_user(owner_id)
    sellers = await db_all(
        "SELECT id FROM users WHERE role='seller' AND id!=? AND is_blocked=0", (owner_id,)
    )
    if not needs or not sellers:
        return

    preview = "\n".join([f"• {n['quantity']} {n['unit']} — {n['product_name']}" for n in needs[:5]])
    if len(needs) > 5:
        preview += f"\n• ...va yana {len(needs)-5} ta"

    if WEBAPP_URL:
        url = f"{WEBAPP_URL}/offer/{batch_id}"
        kb  = ik(
            [ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))],
            [ib("⏭ Keyinroq", "skip_notify")],
        )
    else:
        bot_info = await bot.get_me()
        kb = ik(
            [ib("📤 Taklif yuborish", url=f"https://t.me/{bot_info.username}?start=offer_{needs[0]['id']}")],
            [ib("⏭ Keyinroq", "skip_notify")],
        )

    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    txt = (
        f"📦 *{len(needs)} ta buyurtma!*\n\n"
        f"{preview}\n\n"
        f"📍 {owner['region'] or ''}\n"
        f"⏱ {dl_map.get(needs[0]['deadline_hours'], '?')} ichida"
    )
    # Batch dagi rasmni topamiz
    photo_id = None
    for n in needs:
        if n.get("photo_file_id"):
            photo_id = n["photo_file_id"]
            break

    sent = 0
    for s in sellers:
        try:
            if photo_id:
                await bot.send_photo(s["id"], photo_id, caption=txt, reply_markup=kb)
            else:
                await bot.send_message(s["id"], txt, reply_markup=kb)
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass
    log.info(f"Batch notify: {sent}/{len(sellers)} sotuvchiga")

# ── EHTIYOJLARIM ─────────────────────────────────────────────────────────────
@router.message(F.text == "📋 Ehtiyojlarim")
async def my_needs(msg: Message):
    needs = await db_all(
        "SELECT n.*, "
        "(SELECT COUNT(*) FROM offers o WHERE o.need_id=n.id AND o.status='pending') as offer_cnt "
        "FROM needs n WHERE n.owner_id=? ORDER BY n.created_at DESC LIMIT 20",
        (msg.from_user.id,),
    )
    if not needs:
        await msg.answer(
            "📭 Hali ehtiyoj yo'q.",
            reply_markup=ik([ib("✏️ Ehtiyoj yozish", "new_need")]),
        )
        return

    await msg.answer(f"📋 *Ehtiyojlarim:* {len(needs)} ta")
    for n in needs:
        st    = {"active": "🟢", "paused": "⏸", "done": "✅", "cancelled": "❌"}.get(n["status"], "📋")
        cnt   = n["offer_cnt"] or 0
        badge = f" | 📩 *{cnt} taklif*" if cnt > 0 else ""
        rows  = []
        if cnt > 0:
            rows.append([ib(f"📩 {cnt} ta taklif ko'rish", f"view_offers_{n['id']}")])
        if n.get("batch_id"):
            rows.append([ib("📊 Jadval ko'rish", f"view_batch_{n['batch_id']}")])
        rows.append([
            ib("🔄 Qayta", f"repost_{n['id']}"),
            ib("⏸", f"pause_{n['id']}"),
            ib("✅ Tugat", f"done_{n['id']}"),
        ])
        await msg.answer(
            f"{st} *{n['product_name']}* — {n['quantity']} {n['unit']}{badge}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
        )

@router.callback_query(F.data == "new_need")
async def new_need_btn(call: CallbackQuery, state: FSMContext):
    await call.answer()
    await need_start(call.message, state)

@router.callback_query(F.data.startswith("view_batch_"))
async def view_batch_offers(call: CallbackQuery):
    batch_id = int(call.data[11:])
    # Taklif soni
    cnt = (await db_get(
        "SELECT COUNT(DISTINCT seller_id) as c FROM offers WHERE batch_id=? AND price>0",
        (batch_id,)))["c"] or 0
    if cnt == 0:
        await call.message.answer("📭 Hali taklif kelmagan. Kutishda davom eting.")
        await call.answer()
        return
    # Mini App bor bo'lsa — jadval sahifasi
    if WEBAPP_URL:
        url = f"{WEBAPP_URL}/compare/{batch_id}"
        await call.message.answer(
            f"📊 *{cnt} ta sotuvchidan taklif keldi!*\n\nJadvalda taqqoslang:",
            reply_markup=ik(
                [ib("📊 Jadval ko'rish →", web_app=WebAppInfo(url=url))],
                [ib("📋 Bot ichida ko'rish", f"batch_text_{batch_id}")],
            )
        )
    else:
        await _show_batch_table(call.message, batch_id)
    await call.answer()

@router.callback_query(F.data.startswith("batch_text_"))
async def batch_text_view(call: CallbackQuery):
    batch_id = int(call.data[11:])
    await _show_batch_table(call.message, batch_id)
    await call.answer()

async def _show_batch_table(target_msg, batch_id: int):
    """Batch dagi barcha takliflarni jadval ko'rinishida ko'rsatadi."""
    needs = await db_all(
        "SELECT * FROM needs WHERE batch_id=? AND status != 'cancelled' ORDER BY id",
        (batch_id,),
    )
    if not needs:
        await target_msg.answer("📭 Bu buyurtmada mahsulot yo'q.")
        return

    # Barcha sotuvchilarni topamiz
    sellers_map = {}  # seller_id → name
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? AND o.price > 0 ORDER BY o.price ASC",
            (n["id"],),
        )
        for o in offs:
            sid = o["seller_id"]
            if sid not in sellers_map:
                sellers_map[sid] = o["clinic_name"] or o["full_name"] or f"Sotuvchi{sid}"

    has_offers = len(sellers_map) > 0
    txt = f"📊 *Jadval #{batch_id}*\n_{len(needs)} ta mahsulot"
    txt += f", {len(sellers_map)} ta taklif_\n\n" if has_offers else " — taklif kutilmoqda_\n\n"

    rows_for_accept = []  # (need_id, best_offer_id, best_price, seller_name)

    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? AND o.price > 0 ORDER BY o.price ASC",
            (n["id"],),
        )
        st = {"active":"🟢","paused":"⏸","done":"✅","cancelled":"❌"}.get(n["status"],"📋")
        txt += f"{st} *{n['product_name']}* — {n['quantity']} {n['unit']}\n"
        if offs:
            for i, o in enumerate(offs, 1):
                sname  = o["clinic_name"] or o["full_name"] or "Sotuvchi"
                marker = "✅ " if i == 1 else "   "
                note   = f" _{o['note']}_" if o.get("note") and o["note"] != "mavjud_emas" else ""
                txt   += f"  {marker}{i}. {sname} — {o['price']:,.0f} so'm{note}\n"
            if n["status"] == "active":
                best = offs[0]
                rows_for_accept.append((n["id"], best["id"], best["price"],
                                        best["clinic_name"] or best["full_name"] or "Sotuvchi"))
        else:
            txt += "  _taklif kelmagan_\n"
        txt += "\n"

    # Qabul tugmalari
    kb_rows = []
    if rows_for_accept:
        txt += "─────────────────\n"
        txt += "_Eng arzon taklif qabul qilish:_\n"
        for nid, oid, price, sname in rows_for_accept[:8]:
            short = sname[:12] + ("…" if len(sname) > 12 else "")
            kb_rows.append([ib(f"✅ {short} {price:,.0f}", f"acc_{oid}")])

    # Excel tugmasi
    kb_rows.append([ib("📥 Excel yuklab olish", f"xlsx_{batch_id}")])
    await target_msg.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))

@router.callback_query(F.data.startswith("xlsx_"))
async def download_xlsx(call: CallbackQuery):
    batch_id = int(call.data[5:])
    await call.answer("⏳ Excel tayyorlanmoqda...")
    path = await build_excel(batch_id)
    if not path:
        await call.message.answer("❌ Excel yaratib bo'lmadi (openpyxl o'rnatilmagan?)")
        return
    import aiofiles
    async with aiofiles.open(path, "rb") as f:
        data = await f.read()
    await call.message.answer_document(
        document=BufferedInputFile(data, filename="jadval.xlsx"),
        caption=f"📊 Jadval #{batch_id}",
    )
    try:
        import os as _os
        _os.remove(path)
    except Exception:
        pass

@router.callback_query(F.data.startswith("view_offers_"))
async def view_offers(call: CallbackQuery):
    nid  = int(call.data[12:])
    nd   = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
    offs = await db_all(
        "SELECT o.*, u.full_name, u.phone, u.clinic_name "
        "FROM offers o JOIN users u ON o.seller_id=u.id "
        "WHERE o.need_id=? AND o.status='pending' ORDER BY o.price ASC",
        (nid,),
    )
    if not offs:
        await call.answer("Taklif yo'q hali", show_alert=True)
        return

    txt = f"📩 *{nd['product_name']}* uchun takliflar\n_{len(offs)} ta, arzondan:_\n\n"
    for i, o in enumerate(offs, 1):
        name  = o["clinic_name"] or o["full_name"] or "Sotuvchi"
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        txt  += f"{medal} *{name}*\n   💰 {o['price']:,.0f} so'm/{nd['unit']}\n   🚚 {o['delivery_hours']} soat\n\n"

    rows = [[ib(f"✅ {i}. Qabul — {o['price']:,.0f} so'm", f"acc_{o['id']}")] for i, o in enumerate(offs, 1)]
    rows.append([ib("◀️ Orqaga", "back")])
    await call.message.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await call.answer()

@router.callback_query(F.data.startswith("acc_"))
async def accept_offer(call: CallbackQuery):
    oid = int(call.data[4:])
    o   = await db_get(
        "SELECT o.*, u.full_name, u.phone, u.clinic_name, u.region, u.address "
        "FROM offers o JOIN users u ON o.seller_id=u.id WHERE o.id=?", (oid,)
    )
    if not o:
        await call.answer("Taklif topilmadi", show_alert=True)
        return

    await db_run("UPDATE offers SET status=\'accepted\' WHERE id=?", (oid,))
    await db_run("UPDATE needs SET status=\'paused\' WHERE id=?", (o["need_id"],))

    sname = o["clinic_name"] or o["full_name"] or "Sotuvchi"
    phone = o["phone"] or "—"
    await call.message.edit_text(
        f"✅ *Qabul qilindi!*\n\n🏪 {sname}\n📞 {phone}\n\n_Sotuvchi siz bilan bog\'lanadi._"
    )

    clinic = await get_user(call.from_user.id)
    nd     = await db_get("SELECT * FROM needs WHERE id=?", (o["need_id"],))
    if clinic and nd:
        await _notify_winner(
            seller_id=o["seller_id"],
            clinic=clinic,
            items=[(nd["product_name"], nd["quantity"], nd["unit"], o["price"])],
        )
    # Yutqazganlarga
    if nd:
        for loser in await db_all(
            "SELECT * FROM offers WHERE need_id=? AND seller_id!=? AND price>0",
            (nd["id"], o["seller_id"])
        ):
            await _notify_loser(loser["seller_id"], nd["product_name"], o["price"], loser["price"], nd["unit"])
    await call.answer("✅")

@router.callback_query(F.data.startswith("acc_batch_"))
async def accept_batch_offer(call: CallbackQuery):
    """Bitta do'kondan hamma narsani qabul qilish."""
    parts     = call.data[10:].split("_")
    batch_id  = int(parts[0])
    seller_id = int(parts[1])
    clinic    = await get_user(call.from_user.id)

    needs = await db_all(
        "SELECT * FROM needs WHERE batch_id=? AND status=\'active\'", (batch_id,)
    )
    winner_items = []
    accepted = 0
    for n in needs:
        best = await db_get(
            "SELECT * FROM offers WHERE need_id=? AND seller_id=? AND price>0",
            (n["id"], seller_id),
        )
        if not best:
            continue
        await db_run("UPDATE offers SET status=\'accepted\' WHERE id=?", (best["id"],))
        await db_run("UPDATE needs SET status=\'paused\' WHERE id=?", (n["id"],))
        winner_items.append((n["product_name"], n["quantity"], n["unit"], best["price"]))
        accepted += 1
        # Yutqazganlarga
        for loser in await db_all(
            "SELECT * FROM offers WHERE need_id=? AND seller_id!=? AND price>0",
            (n["id"], seller_id)
        ):
            await _notify_loser(loser["seller_id"], n["product_name"],
                                best["price"], loser["price"], n["unit"])

    seller_info = await db_get("SELECT clinic_name, full_name FROM users WHERE id=?", (seller_id,))
    sname = (seller_info["clinic_name"] or seller_info["full_name"] if seller_info else None) or "Sotuvchi"

    if clinic and winner_items:
        await _notify_winner(seller_id, clinic, winner_items)

    await call.message.edit_text(
        f"✅ *{sname} dan {accepted} ta taklif qabul qilindi!*\n\n_Sotuvchi siz bilan bog\'lanadi._"
    )
    await call.answer("✅")

@router.callback_query(F.data.startswith("repost_"))
async def repost_need(call: CallbackQuery):
    nid     = int(call.data[7:])
    nd      = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
    expires = (datetime.now() + timedelta(hours=nd["deadline_hours"])).isoformat()
    await db_run("UPDATE needs SET status='active', expires_at=? WHERE id=?", (expires, nid))
    mid = await post_to_channel(nid, dict(nd))
    if mid:
        await db_run("UPDATE needs SET channel_message_id=? WHERE id=?", (mid, nid))
    chan = CHANNEL_ID.lstrip("@") if isinstance(CHANNEL_ID, str) else str(CHANNEL_ID)
    link = f"[Ko'rish](https://t.me/{chan}/{mid})" if mid else ""
    await call.message.answer(f"🔄 *{nd['product_name']}* qayta joylashtirildi! {link}")
    await call.answer("✅")

@router.callback_query(F.data.startswith("pause_"))
async def pause_need(call: CallbackQuery):
    await db_run("UPDATE needs SET status='paused' WHERE id=?", (int(call.data[6:]),))
    await call.answer("⏸ Pauza")

@router.callback_query(F.data.startswith("done_"))
async def done_need(call: CallbackQuery):
    await db_run("UPDATE needs SET status='done' WHERE id=?", (int(call.data[5:]),))
    await call.answer("✅ Yakunlandi")

# ── TAKLIFLAR (klinika uchun) ─────────────────────────────────────────────────
@router.message(F.text == "📩 Takliflar")
async def clinic_offers(msg: Message):
    uid  = msg.from_user.id
    offs = await db_all(
        "SELECT o.*, n.product_name as np, n.unit as nu, u.clinic_name, u.full_name, u.phone "
        "FROM offers o "
        "JOIN needs n ON o.need_id=n.id "
        "JOIN users u ON o.seller_id=u.id "
        "WHERE n.owner_id=? AND o.status='pending' ORDER BY o.price ASC LIMIT 30",
        (uid,),
    )
    if not offs:
        await msg.answer("📭 Hali taklif kelmagan.")
        return
    await msg.answer(f"📩 *Kelgan takliflar:* {len(offs)} ta\n_(arzondan qimmatga)_")
    for i, o in enumerate(offs, 1):
        name  = o["clinic_name"] or o["full_name"] or "Sotuvchi"
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        note  = f"\n📝 _{o['note']}_" if o.get("note") else ""
        await msg.answer(
            f"{medal} *{o['np']}*\n"
            f"💰 {o['price']:,.0f} so'm/{o['nu']}\n"
            f"🚚 {o['delivery_hours']} soat\n"
            f"🏪 {name}{note}",
            reply_markup=ik(
                [ib("✅ Qabul", f"acc_{o['id']}"), ib("❌ Rad", f"rej_{o['id']}")],
            ),
        )

@router.callback_query(F.data.startswith("rej_"))
async def reject_offer(call: CallbackQuery):
    await db_run("UPDATE offers SET status='rejected' WHERE id=?", (int(call.data[4:]),))
    await call.answer("❌ Rad etildi")

@router.callback_query(F.data == "back")
async def cb_back(call: CallbackQuery):
    await call.answer()

# ── JADVAL ────────────────────────────────────────────────────────────────────
async def build_table(batch_id: int) -> str:
    """Batch dagi ehtiyojlar va takliflar jadvalini matn sifatida qaytaradi."""
    needs = await db_all(
        "SELECT * FROM needs WHERE batch_id=? ORDER BY id", (batch_id,)
    )
    if not needs:
        return "Bo'sh jadval."

    lines = [f"📊 *Jadval #{batch_id}*\n"]
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? ORDER BY o.price ASC",
            (n["id"],),
        )
        lines.append(f"🦷 *{n['product_name']}* — {n['quantity']} {n['unit']}")
        if not offs:
            lines.append("   _Taklif kelmagan_")
        else:
            for i, o in enumerate(offs, 1):
                name   = o["clinic_name"] or o["full_name"] or "Sotuvchi"
                marker = "✅ " if i == 1 else ""
                lines.append(f"   {marker}{i}. {name} — {o['price']:,.0f} so'm")
        lines.append("")
    return "\n".join(lines)

async def build_excel(batch_id: int):
    """Excel fayl yaratadi, path qaytaradi."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    needs = await db_all("SELECT * FROM needs WHERE batch_id=?", (batch_id,))
    sellers_set = set()
    need_offers = {}
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? ORDER BY o.price ASC",
            (n["id"],),
        )
        need_offers[n["id"]] = list(offs)
        for o in offs:
            sellers_set.add(o["clinic_name"] or o["full_name"] or f"Sotuvchi{o['seller_id']}")

    sellers = sorted(sellers_set)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadval #{batch_id}"

    # Header
    header = ["Mahsulot", "Miqdor", "Birlik"] + [f"{s}\n(1 ta narx)" for s in sellers] + ["Eng arzon jami"]
    for col, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_i, n in enumerate(needs, 2):
        ws.cell(row=row_i, column=1, value=n["product_name"])
        ws.cell(row=row_i, column=2, value=n["quantity"])
        ws.cell(row=row_i, column=3, value=n["unit"])
        offs      = need_offers[n["id"]]
        qty       = float(n["quantity"])
        min_total = None
        for o in offs:
            seller_name = o["clinic_name"] or o["full_name"] or f"Sotuvchi{o['seller_id']}"
            if seller_name in sellers:
                col_i   = sellers.index(seller_name) + 4
                unit_p  = o["price"]          # 1 ta uchun
                total_p = unit_p * qty        # jami
                # Katakda: "1 ta: 45,000" yozamiz, ustun sarlavhasida miqdor ko'rinadi
                ws.cell(row=row_i, column=col_i, value=unit_p)
                if min_total is None or total_p < min_total:
                    min_total = total_p
        if min_total is not None:
            last_col = len(sellers) + 4
            cell = ws.cell(row=row_i, column=last_col, value=min_total)
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
            cell.font = Font(bold=True)

    # Column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = os.path.join(BASE_DIR, f"jadval_{batch_id}.xlsx")
    wb.save(path)
    return path

# ── SOTUVCHI: EHTIYOJLAR ──────────────────────────────────────────────────────
@router.message(F.text == "🔔 Ehtiyojlar")
async def seller_feed(msg: Message):
    needs = await db_all(
        "SELECT n.*, u.region FROM needs n JOIN users u ON n.owner_id=u.id "
        "WHERE n.status='active' ORDER BY n.created_at DESC LIMIT 20"
    )
    if not needs:
        await msg.answer("📭 Hozircha aktiv ehtiyoj yo'q.")
        return

    await msg.answer(f"🔔 *Aktiv ehtiyojlar:* {len(needs)} ta")
    for n in needs:
        existing = await db_get(
            "SELECT id FROM offers WHERE need_id=? AND seller_id=?", (n["id"], msg.from_user.id)
        )
        if existing:
            kb = ik([ib("✅ Taklif yuborilgan", "noop")])
        elif WEBAPP_URL:
            url = f"{WEBAPP_URL}/offer/{n['batch_id'] or n['id']}"
            kb  = ik([ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))])
        else:
            kb = ik([ib("📤 Taklif yuborish", f"offer_{n['id']}")])

        dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
        await msg.answer(
            f"🦷 *{n['product_name']}*\n"
            f"📦 {n['quantity']} {n['unit']}\n"
            f"⏱ {dl_map.get(n['deadline_hours'], '?')} ichida\n"
            f"📍 {n['region'] or ''}",
            reply_markup=kb,
        )

@router.callback_query(F.data == "noop")
async def noop(call: CallbackQuery):
    await call.answer("Allaqachon taklif yubordingiz")

# ── SOTUVCHI: BOT ICHIDA TAKLIF (Mini App yo'q bo'lsa) ───────────────────────
@router.callback_query(F.data.startswith("offer_"))
async def offer_start(call: CallbackQuery, state: FSMContext):
    nid = int(call.data[6:])
    await _start_offer_bot(call, state, nid)
    await call.answer()

async def _start_offer_bot(msg_or_call, state: FSMContext, nid: int):
    uid = msg_or_call.from_user.id
    nd  = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
    if not nd or nd["status"] != "active":
        txt = "⚠️ Bu ehtiyoj yopilgan yoki topilmadi."
        if hasattr(msg_or_call, "answer"):
            await msg_or_call.answer(txt)
        else:
            await msg_or_call.message.answer(txt)
        return

    exists = await db_get("SELECT id FROM offers WHERE need_id=? AND seller_id=?", (nid, uid))
    if exists:
        txt = "⚠️ Bu ehtiyojga allaqachon taklif yubordingiz!"
        if hasattr(msg_or_call, "answer"):
            await msg_or_call.answer(txt)
        else:
            await msg_or_call.message.answer(txt)
        return

    await state.update_data(need_id=nid, need_unit=nd["unit"], need_name=nd["product_name"])
    await state.set_state(OfferState.price)
    txt = (
        f"📦 *{nd['product_name']}* — {nd['quantity']} {nd['unit']}\n\n"
        f"💰 Narxingiz? _(1 {nd['unit']} uchun, so'mda)_"
    )
    no_stock = ik([ib("❌ Mavjud emas", f"no_stock_{nid}")])
    if hasattr(msg_or_call, "answer"):
        await msg_or_call.answer(txt, reply_markup=no_stock)
    else:
        await msg_or_call.message.answer(txt, reply_markup=no_stock)

@router.callback_query(F.data.startswith("no_stock_"))
async def no_stock(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("❌ *Mavjud emas* deb belgilandi. Rahmat!")
    await call.answer()

@router.message(OfferState.price)
async def offer_price(msg: Message, state: FSMContext):
    try:
        price = float(msg.text.replace(" ", "").replace(",", ""))
    except Exception:
        await msg.answer("❌ Faqat raqam kiriting! _Masalan: 285000_")
        return
    await state.update_data(price=price)
    await state.set_state(OfferState.note)
    await msg.answer(
        "📝 Izoh? _(ixtiyoriy: brend, sifat, muddat...)_",
        reply_markup=ik([ib("⏭ Izohsiz yuborish", "offer_no_note")]),
    )

@router.callback_query(F.data == "offer_no_note", OfferState.note)
async def offer_no_note(call: CallbackQuery, state: FSMContext):
    await _save_offer(call, state, note=None)
    await call.answer()

@router.message(OfferState.note)
async def offer_note(msg: Message, state: FSMContext):
    await _save_offer(msg, state, note=msg.text)

async def _save_offer(obj, state: FSMContext, note):
    uid = obj.from_user.id
    d   = await state.get_data()
    u   = await get_user(uid)

    # Delivery default 24 soat
    await db_insert(
        "INSERT INTO offers(need_id,seller_id,product_name,price,unit,delivery_hours,note) VALUES(?,?,?,?,?,?,?)",
        (d["need_id"], uid, d["need_name"], d["price"], d["need_unit"], 24, note),
    )

    nd     = await db_get(
        "SELECT n.*, u2.id as cid FROM needs n JOIN users u2 ON n.owner_id=u2.id WHERE n.id=?",
        (d["need_id"],),
    )
    shop   = await db_get("SELECT shop_name FROM shops WHERE owner_id=? AND status='active'", (uid,))
    sname  = (shop["shop_name"] if shop else None) or u["clinic_name"] or u["full_name"] or "Sotuvchi"
    note_t = f"\n📝 _{note}_" if note else ""

    try:
        await bot.send_message(
            nd["cid"],
            f"📩 *Yangi taklif!*\n\n"
            f"🦷 {d['need_name']}\n"
            f"💰 *{d['price']:,.0f} so'm*/{d['need_unit']}\n"
            f"🏪 {sname}{note_t}",
            reply_markup=ik([ib(f"📩 Barcha takliflarni ko'rish", f"view_offers_{d['need_id']}")]),
        )
    except Exception as e:
        log.error(f"Klinikaga xabar xato: {e}")

    await state.clear()
    txt = f"✅ *Taklif yuborildi!*\n\n🦷 {d['need_name']}\n💰 {d['price']:,.0f} so'm/{d['need_unit']}"
    if note:
        txt += f"\n📝 {note}"
    if hasattr(obj, "answer"):
        await obj.answer(txt)
    else:
        await obj.message.answer(txt)

# ── SOTUVCHI: TAKLIFLARIM ─────────────────────────────────────────────────────
@router.message(F.text == "📤 Takliflarim")
async def my_offers(msg: Message):
    uid  = msg.from_user.id
    offs = await db_all(
        "SELECT o.*, n.product_name as np, n.quantity as nqty, n.unit as nunit "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? ORDER BY o.created_at DESC LIMIT 30",
        (uid,),
    )
    if not offs:
        await msg.answer("📭 Hali taklif yubormagansiz.")
        return

    # Jami savdo summasi
    total_won = sum(
        o["price"] * o["nqty"]
        for o in offs if o["status"] == "accepted"
    )
    won_count = sum(1 for o in offs if o["status"] == "accepted")
    pend_count= sum(1 for o in offs if o["status"] == "pending")

    summary = (
        f"📤 *Takliflarim* ({len(offs)} ta)\n\n"
        f"✅ Qabul: *{won_count} ta*\n"
        f"⏳ Kutmoqda: *{pend_count} ta*\n"
        f"💰 Jami savdo: *{total_won:,.0f} so'm*"
    )
    await msg.answer(summary, reply_markup=ik(
        [ib("📊 Batafsil statistika", "seller_stats")],
        [ib("📥 Excel yuklab olish", "seller_excel")],
    ))
    # So'nggi 10 ta
    for o in offs[:10]:
        st = {"pending":"⏳","accepted":"✅","rejected":"❌"}.get(o["status"],"📤")
        total_line = o["price"] * o["nqty"]
        await msg.answer(
            f"{st} *{o['np']}* — {o['nqty']} {o['nunit']}\n"
            f"💰 {o['price']:,.0f} × {o['nqty']} = *{total_line:,.0f} so'm*"
        )

# ── SOTUVCHI STATISTIKA ──────────────────────────────────────────────────────
@router.message(F.text == "📊 Statistika")
async def seller_stats_btn(msg: Message):
    await _show_seller_stats(msg.from_user.id, msg)

@router.callback_query(F.data == "seller_stats")
async def seller_stats_cb(call: CallbackQuery):
    await _show_seller_stats(call.from_user.id, call.message)
    await call.answer()

async def _show_seller_stats(uid: int, target_msg):
    now   = datetime.now()
    today = now.strftime("%Y-%m-%d")
    week  = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    month = now.strftime("%Y-%m")
    year  = now.strftime("%Y")

    async def won_sum(since=None, period=None):
        if period:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\' AND o.created_at LIKE ?",
                (uid, f"{period}%")
            )
        elif since:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\' AND o.created_at >= ?",
                (uid, since)
            )
        else:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\'",
                (uid,)
            )
        return sum(r["price"] * r["quantity"] for r in rows), len(rows)

    day_sum,   day_cnt   = await won_sum(since=today)
    week_sum,  week_cnt  = await won_sum(since=week)
    month_sum, month_cnt = await won_sum(period=month)
    year_sum,  year_cnt  = await won_sum(period=year)
    total_sum, total_cnt = await won_sum()

    # Jami takliflar
    all_offs  = await db_get("SELECT COUNT(*) as c FROM offers WHERE seller_id=?", (uid,))
    all_count = all_offs["c"] if all_offs else 0
    rate      = f"{total_cnt/all_count*100:.0f}%" if all_count else "—"

    # Top 5 mahsulot
    top_prods = await db_all(
        "SELECT n.product_name, COUNT(*) as cnt, SUM(o.price*n.quantity) as total "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY n.product_name ORDER BY total DESC LIMIT 5",
        (uid,)
    )

    txt = (
        f"📊 *Savdo statistikasi*\n\n"
        f"📅 Bugun:      *{day_sum:>12,.0f} so'm* ({day_cnt} ta)\n"
        f"📅 Bu hafta:   *{week_sum:>12,.0f} so'm* ({week_cnt} ta)\n"
        f"📅 Bu oy:      *{month_sum:>12,.0f} so'm* ({month_cnt} ta)\n"
        f"📅 Bu yil:     *{year_sum:>12,.0f} so'm* ({year_cnt} ta)\n"
        f"📅 Jami:       *{total_sum:>12,.0f} so'm* ({total_cnt} ta)\n\n"
        f"📤 Jami taklif: {all_count} ta | ✅ Qabul: {rate}\n"
    )

    if top_prods:
        txt += "\n🏆 *Top mahsulotlar:*\n"
        for i, p in enumerate(top_prods, 1):
            txt += f"  {i}. {p['product_name']} — {p['total']:,.0f} so'm ({p['cnt']} ta)\n"

    await target_msg.answer(txt, reply_markup=ik(
        [ib("📥 Excel yuklab olish", "seller_excel")],
        [ib("◀️ Orqaga", "seller_stats_back")],
    ))

@router.callback_query(F.data == "seller_stats_back")
async def seller_stats_back(call: CallbackQuery):
    await call.message.delete()
    await call.answer()

@router.callback_query(F.data == "seller_excel")
async def seller_excel_cb(call: CallbackQuery):
    await call.answer("⏳ Excel tayyorlanmoqda...")
    uid  = call.from_user.id
    path = await _build_seller_excel(uid)
    if not path:
        await call.message.answer("❌ Excel yaratib bo\'lmadi")
        return
    import aiofiles
    async with aiofiles.open(path, "rb") as f:
        data = await f.read()
    fname = f"savdo_{datetime.now().strftime('%Y%m%d')}.xlsx"
    await call.message.answer_document(
        document=BufferedInputFile(data, filename=fname),
        caption=f"📊 Savdo hisoboti — {datetime.now().strftime('%d.%m.%Y')}"
    )
    try:
        import os as _os; _os.remove(path)
    except Exception: pass

async def _build_seller_excel(uid: int) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb  = openpyxl.Workbook()
    now = datetime.now()

    def hdr(ws, cols, color="4472C4"):
        fill = PatternFill("solid", fgColor=color)
        font = Font(bold=True, color="FFFFFF")
        for i, v in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=v)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center")

    def aw(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)

    # ── 1. Umumiy savdo ──────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Jami savdo"
    hdr(ws1, ["Sana", "Mahsulot", "Miqdor", "Birlik", "Narx (1 ta)", "Jami", "Klinika"])
    offs = await db_all(
        "SELECT o.*, n.product_name, n.quantity, n.unit, "
        "COALESCE(u.clinic_name, u.full_name) as clinic "
        "FROM offers o "
        "JOIN needs n ON o.need_id=n.id "
        "JOIN users u ON n.owner_id=u.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "ORDER BY o.created_at DESC",
        (uid,)
    )
    for i, o in enumerate(offs, 2):
        ws1.cell(row=i, column=1, value=o["created_at"][:10] if o["created_at"] else "")
        ws1.cell(row=i, column=2, value=o["product_name"])
        ws1.cell(row=i, column=3, value=o["quantity"])
        ws1.cell(row=i, column=4, value=o["unit"])
        ws1.cell(row=i, column=5, value=o["price"])
        ws1.cell(row=i, column=6, value=o["price"] * o["quantity"])
        ws1.cell(row=i, column=7, value=o["clinic"] or "—")
    # Jami
    if offs:
        row = len(offs) + 2
        ws1.cell(row=row, column=5, value="JAMI:").font = Font(bold=True)
        total = sum(o["price"]*o["quantity"] for o in offs)
        ws1.cell(row=row, column=6, value=total).font = Font(bold=True)
    aw(ws1)

    # ── 2. Oylik ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Oylik")
    hdr(ws2, ["Oy", "Bitimlar", "Jami savdo (so'm)"], "2E7D32")
    monthly = await db_all(
        "SELECT SUBSTR(o.created_at,1,7) as month, "
        "COUNT(*) as cnt, SUM(o.price*n.quantity) as total "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY month ORDER BY month DESC LIMIT 24",
        (uid,)
    )
    for i, m in enumerate(monthly, 2):
        ws2.cell(row=i, column=1, value=m["month"])
        ws2.cell(row=i, column=2, value=m["cnt"])
        ws2.cell(row=i, column=3, value=m["total"])
    aw(ws2)

    # ── 3. Mahsulotlar ───────────────────────────────────────────
    ws3 = wb.create_sheet("Mahsulotlar")
    hdr(ws3, ["Mahsulot", "Bitimlar", "Jami (so'm)", "O'rtacha narx"], "1565C0")
    prods = await db_all(
        "SELECT n.product_name, COUNT(*) as cnt, "
        "SUM(o.price*n.quantity) as total, AVG(o.price) as avg_p "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY n.product_name ORDER BY total DESC",
        (uid,)
    )
    for i, p in enumerate(prods, 2):
        ws3.cell(row=i, column=1, value=p["product_name"])
        ws3.cell(row=i, column=2, value=p["cnt"])
        ws3.cell(row=i, column=3, value=p["total"])
        ws3.cell(row=i, column=4, value=round(p["avg_p"], 0) if p["avg_p"] else 0)
    aw(ws3)

    path = os.path.join(BASE_DIR, f"seller_{uid}_{now.strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(path)
    return path

# ── MAHSULOTLARIM ────────────────────────────────────────────────────────────
@router.message(F.text == "📦 Mahsulotlarim")
async def my_products(msg: Message, state: FSMContext):
    uid  = msg.from_user.id
    rows = await db_all(
        "SELECT * FROM clinic_products WHERE owner_id=? ORDER BY sort_order, id",
        (uid,)
    )
    if not rows:
        await msg.answer(
            "📦 *Mahsulotlar ro\'yxati*\n\nRo\'yxat bo\'sh.\n"
            "Tez-tez buyurtma beradigan mahsulotlarni qo\'shing — "
            "keyingi buyurtmada avtomatik chiqadi.",
            reply_markup=ik([ib("➕ Mahsulot qo\'shish", "prod_add")])
        )
        return
    txt = f"📦 *Mahsulotlarim* ({len(rows)} ta)\n\n"
    for i, r in enumerate(rows, 1):
        txt += f"{i}. *{r['name']}* — {r['unit']}\n"
    await msg.answer(txt, reply_markup=ik(
        [ib("➕ Qo\'shish", "prod_add"), ib("❌ O\'chirish", "prod_del")],
    ))

@router.callback_query(F.data == "prod_add")
async def prod_add(call: CallbackQuery, state: FSMContext):
    await state.set_state(MyProductsState.editing)
    await state.update_data(prod_action="add")
    await call.message.answer(
        "✏️ Mahsulot nomini kiriting:\n\n"
        "_Masalan: GC Fuji IX, Xarizma A2, Spirt_\n\n"
        "/cancel — bekor qilish"
    )
    await call.answer()

@router.callback_query(F.data == "prod_del")
async def prod_del(call: CallbackQuery, state: FSMContext):
    uid  = call.from_user.id
    rows = await db_all(
        "SELECT * FROM clinic_products WHERE owner_id=? ORDER BY sort_order, id", (uid,)
    )
    if not rows:
        await call.answer("Ro\'yxat bo\'sh", show_alert=True)
        return
    kb_rows = []
    for r in rows:
        kb_rows.append([ib(f"❌ {r['name']}", f"prod_del_{r['id']}")])
    kb_rows.append([ib("◀️ Orqaga", "prod_back")])
    await call.message.answer(
        "O\'chirmoqchi bo\'lgan mahsulotni tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows)
    )
    await call.answer()

@router.callback_query(F.data.startswith("prod_del_"))
async def prod_del_item(call: CallbackQuery):
    pid = int(call.data[9:])
    row = await db_get("SELECT * FROM clinic_products WHERE id=? AND owner_id=?",
                       (pid, call.from_user.id))
    if not row:
        await call.answer("Topilmadi", show_alert=True)
        return
    await db_run("DELETE FROM clinic_products WHERE id=?", (pid,))
    await call.message.edit_text(f"✅ *{row['name']}* o\'chirildi.")
    await call.answer()

@router.callback_query(F.data == "prod_back")
async def prod_back(call: CallbackQuery):
    await call.message.delete()
    await call.answer()

@router.message(MyProductsState.editing)
async def prod_add_name(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear()
        return
    name = msg.text.strip() if msg.text else ""
    if not name or len(name) < 2:
        await msg.answer("❌ Kamida 2 ta harf kiriting.")
        return
    await state.update_data(prod_name=name)
    await msg.answer(
        f"*{name}* — birligini tanlang:",
        reply_markup=ik(
            [ib("dona","pu_dona"), ib("kg","pu_kg"), ib("litr","pu_litr")],
            [ib("quti","pu_quti"), ib("paket","pu_paket"), ib("ml","pu_ml")],
        )
    )

@router.callback_query(F.data.startswith("pu_"), MyProductsState.editing)
async def prod_add_unit(call: CallbackQuery, state: FSMContext):
    unit = call.data[3:]
    d    = await state.get_data()
    name = d.get("prod_name","")
    uid  = call.from_user.id
    # Mavjudmi?
    ex = await db_get(
        "SELECT id FROM clinic_products WHERE owner_id=? AND name=?", (uid, name)
    )
    if ex:
        await call.message.edit_text(f"⚠️ *{name}* allaqachon ro\'yxatda bor.")
        await state.clear(); await call.answer(); return
    await db_insert(
        "INSERT INTO clinic_products(owner_id,name,unit) VALUES(?,?,?)",
        (uid, name, unit)
    )
    await state.clear()
    await call.message.edit_text(f"✅ *{name}* ({unit}) ro\'yxatga qo\'shildi!")
    await call.answer()

# ── DO'KON ────────────────────────────────────────────────────────────────────
@router.message(F.text == "🏪 Do'konim")
async def my_shop(msg: Message):
    uid  = msg.from_user.id
    shop = await db_get("SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,))
    if not shop:
        await msg.answer(
            "🏪 Do'koningiz yo'q yoki tasdiqlanmagan.",
            reply_markup=ik([ib("➕ Do'kon ochish", "open_shop")]),
        )
        return
    prod_count = (await db_get("SELECT COUNT(*) as c FROM products WHERE shop_id=? AND is_active=1", (shop["id"],)))["c"]
    catalog_url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller" if WEBAPP_URL else None
    add_url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller" if WEBAPP_URL else None
    kb_rows = []
    if catalog_url:
        kb_rows.append([ib("🛍 Katalog", web_app=WebAppInfo(url=catalog_url))])
    if add_url:
        kb_rows.append([ib("➕ Mahsulot qo\'shish", web_app=WebAppInfo(url=add_url+"&action=add"))])
    kb_rows.append([ib("📦 Mahsulotlarim (" + str(prod_count) + " ta)", "shop_products")])
    # Guruh holati
    group_id = shop.get("group_chat_id")
    if group_id:
        kb_rows.append([ib("👥 Guruh bog\'langan ✅", "group_info")])
    else:
        kb_rows.append([ib("👥 Guruh ulash (buyurtmalar uchun)", "group_howto")])
    kb = InlineKeyboardMarkup(inline_keyboard=kb_rows)
    rating = float(shop["rating"] or 0)
    stars = ""
    if rating >= 4.5: stars = "⭐⭐⭐⭐⭐"
    elif rating >= 3.5: stars = "⭐⭐⭐⭐"
    elif rating >= 2.5: stars = "⭐⭐⭐"
    elif rating >= 1.5: stars = "⭐⭐"
    elif rating > 0:    stars = "⭐"
    else: stars = "_Hali baholashlar yo\'q_"
    # Baholar soni
    review_count = (await db_get("SELECT COUNT(*) as c FROM reviews WHERE seller_id=?", (uid,)))["c"]
    rate_txt = f"{stars} ({rating:.1f}) · {review_count} ta sharh" if rating > 0 else stars
    await msg.answer(
        f"🏪 *{shop['shop_name']}*\n"
        f"{rate_txt}\n\n"
        f"📂 {shop['category']}\n"
        f"📦 Mahsulotlar: *{prod_count} ta*\n"
        f"🤝 Jami xaridlar: *{shop['total_deals'] or 0} ta*",
        reply_markup=kb
    )

@router.callback_query(F.data == "open_shop")
async def open_shop(call: CallbackQuery, state: FSMContext):
    if not await has_profile(call.from_user.id):
        await call.message.answer(
            "⚠️ Avval profilingizni to'ldiring!",
            reply_markup=ik([ib("✏️ To'ldirish", "edit_profile")]),
        )
        await call.answer()
        return
    await state.set_state(ShopState.cat)
    await call.message.answer("📂 Do'kon kategoriyasini tanlang:", reply_markup=kb_shop_cats())
    await call.answer()

@router.callback_query(F.data.startswith("cat_"), ShopState.cat)
async def shop_cat(call: CallbackQuery, state: FSMContext):
    cats = {
        "cat_1": "🦷 Terapevtik",
        "cat_2": "⚙️ Jarrohlik & Implant",
        "cat_3": "🔬 Zubtexnik",
        "cat_4": "🧪 Dezinfeksiya",
        "cat_5": "💡 Asbob-uskunalar",
    }
    await state.update_data(cat=cats.get(call.data, "Stomatologiya"))
    await state.set_state(ShopState.name)
    await call.message.answer(
        "🏪 Do'kon nomini kiriting:\n\n_Masalan: DentalPlus Toshkent_",
        reply_markup=ReplyKeyboardRemove()
    )
    await call.answer()

@router.message(ShopState.name)
async def shop_name(msg: Message, state: FSMContext):
    d   = await state.get_data()
    u   = await get_user(msg.from_user.id)
    uid = msg.from_user.id
    sname = (msg.text or "").strip()
    if not sname:
        await msg.answer("⚠️ Do'kon nomini kiriting.")
        return
    # Mavjud do'konni yangilaymiz yoki yangisini yaratamiz
    existing = await db_get("SELECT id FROM shops WHERE owner_id=?", (uid,))
    if existing:
        await db_run(
            "UPDATE shops SET shop_name=?, category=?, status='active' WHERE owner_id=?",
            (sname, d.get("cat","Stomatologiya"), uid)
        )
    else:
        await db_insert(
            "INSERT INTO shops(owner_id,shop_name,category,phone,region,status) "
            "VALUES(?,?,?,?,?,'active')",
            (uid, sname, d.get("cat","Stomatologiya"),
             u.get("phone","") if u else "",
             u.get("region","") if u else "")
        )
    await state.clear()
    await msg.answer(
        f"✅ *Do'kon yangilandi!*\n\n🏪 {sname}",
        reply_markup=kb_seller(u["lang"] or "uz", uid=uid, webapp_url=WEBAPP_URL) if u else None
    )
    # Adminga xabar
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(
                aid,
                f"🔄 *Do'kon yangilandi*\n\n"
                f"🏪 {sname}\n"
                f"👤 {u.get('full_name','') if u else ''}\n"
                f"📞 {u.get('phone','') if u else ''}\n"
                f"📍 {u.get('region','') if u else ''}\n"
                f"🆔 `{uid}`"
            )
        except Exception:
            pass

@router.callback_query(F.data == "group_info")
async def cb_group_info(call: CallbackQuery):
    uid  = call.from_user.id
    shop = await db_get("SELECT * FROM shops WHERE owner_id=?", (uid,))
    gid  = shop["group_chat_id"] if shop else None
    await call.message.answer(
        f"👥 *Guruh bog\'langan*\n\n"
        f"Guruh ID: `{gid}`\n\n"
        f"Guruh bog\'liqligini o\'chirish uchun:\n"
        f"Guruhda /unsetgroup yozing"
    )
    await call.answer()

@router.callback_query(F.data == "group_howto")
async def cb_group_howto(call: CallbackQuery):
    await call.message.answer(
        "👥 *Guruh ulash — Qo\'llanma*\n\n"
        "1️⃣ Telegram da yangi guruh oching\n"
        "2️⃣ @XazdentBot ni guruhga qo\'shing\n"
        "3️⃣ Botni *Admin* qiling\n"
        "4️⃣ Guruhda /setgroup yozing\n\n"
        "✅ Shundan keyin barcha buyurtmalar guruhga chiqadi.\n"
        "Jamoangizdаn birinchi javob bergan xodim buyurtmani oladi."
    )
    await call.answer()

@router.callback_query(F.data.startswith("shopok_"))
async def shop_ok(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS:
        return
    uid = int(call.data[7:])
    await db_run("UPDATE shops SET status='active' WHERE owner_id=?", (uid,))
    try:
        await bot.send_message(uid, "✅ Do'koningiz faollashdi!")
    except Exception:
        pass
    await call.message.edit_text(call.message.text + "\n\n✅ TASDIQLANDI", reply_markup=None)
    await call.answer("✅")

@router.callback_query(F.data.startswith("shoprej_"))
async def shop_rej(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS:
        return
    uid = int(call.data[8:])
    await db_run("UPDATE shops SET status='rejected' WHERE owner_id=?", (uid,))
    await call.message.edit_text(call.message.text + "\n\n❌ RAD ETILDI", reply_markup=None)
    await call.answer("❌")

# ── BALANS ────────────────────────────────────────────────────────────────────
@router.message(F.text == "💰 Hisobim")
async def show_balance(msg: Message):
    u    = await get_user(msg.from_user.id)
    card = await get_setting("card_number") or "9860020138100068"
    await msg.answer(
        f"💰 *Hisobingiz*\n\n"
        f"⚡️ Ball: *{u['balance'] or 0:.1f}*\n\n"
        f"_(Hozircha e'lon bepul)_",
        reply_markup=ik([ib("➕ Hisob to'ldirish", "topup")]),
    )

@router.callback_query(F.data == "topup")
async def topup_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(TopupState.amount)
    await call.message.answer(
        "💰 *Hisob to'ldirish*\n\n"
        "Qancha so'm o'tkazmoqchisiz?\n_Faqat raqam kiriting._"
    )
    await call.answer()

@router.message(TopupState.amount)
async def topup_amount(msg: Message, state: FSMContext):
    try:
        amount    = float(msg.text.replace(" ", "").replace(",", ""))
        ballprice = float(await get_setting("ball_price") or 1000)
        balls     = amount / ballprice
    except Exception:
        await msg.answer("❌ Faqat raqam kiriting!")
        return
    card = await get_setting("card_number") or "9860020138100068"
    await state.update_data(amount=amount, balls=balls)
    await state.set_state(TopupState.receipt)
    await msg.answer(
        f"✅ *{amount:,.0f} so'm → {balls:.1f} ball*\n\n"
        f"💳 Ushbu kartaga P2P o'tkazing:\n\n"
        f"`{card}`\n_Komilova M_\n\n"
        f"📸 O'tkazma screenshotini yuboring:"
    )

@router.message(TopupState.receipt, F.photo)
async def topup_receipt(msg: Message, state: FSMContext):
    d   = await state.get_data()
    fid = msg.photo[-1].file_id
    tid = await db_insert(
        "INSERT INTO transactions(user_id,amount,balls,type,receipt_file_id) VALUES(?,?,?,'topup',?)",
        (msg.from_user.id, d["amount"], d["balls"], fid),
    )
    u    = await get_user(msg.from_user.id)
    name = u["clinic_name"] or u["full_name"] or str(msg.from_user.id)
    for aid in ADMIN_IDS:
        try:
            await bot.send_photo(
                aid, fid,
                caption=f"💳 *Yangi chek #{tid}*\n\n👤 {name}\n💰 {d['amount']:,.0f} so'm → {d['balls']:.1f} ball",
                reply_markup=ik(
                    [ib("✅ Tasdiqlash", f"adm_ok_{tid}_{msg.from_user.id}_{d['balls']}"),
                     ib("❌ Rad", f"adm_rej_{tid}_{msg.from_user.id}")],
                ),
            )
        except Exception:
            pass
    await state.clear()
    await msg.answer("✅ Chek yuborildi! Admin 15-30 daqiqada tasdiqlaydi.")

@router.callback_query(F.data.startswith("adm_ok_"))
async def adm_ok(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS:
        return
    parts = call.data.split("_")
    tid, uid, balls = int(parts[2]), int(parts[3]), float(parts[4])
    await db_run("UPDATE transactions SET status='confirmed',confirmed_by=? WHERE id=?", (call.from_user.id, tid))
    await add_balance(uid, balls)
    try:
        await bot.send_message(uid, f"🎉 *Hisobingiz to'ldirildi!*\n\n+{balls:.1f} ball")
    except Exception:
        pass
    await call.message.edit_caption(call.message.caption + "\n\n✅ TASDIQLANDI", reply_markup=None)
    await call.answer("✅")

@router.callback_query(F.data.startswith("adm_rej_"))
async def adm_rej(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS:
        return
    parts = call.data.split("_")
    tid, uid = int(parts[2]), int(parts[3])
    await db_run("UPDATE transactions SET status='rejected' WHERE id=?", (tid,))
    try:
        await bot.send_message(uid, "❌ Chekingiz rad etildi. Admin bilan bog'laning.")
    except Exception:
        pass
    await call.message.edit_caption(call.message.caption + "\n\n❌ RAD ETILDI", reply_markup=None)
    await call.answer("❌")

@router.message(Command("admin"))
async def cmd_admin(msg: Message):
    if msg.from_user.id not in ADMIN_IDS: return
    if not WEBAPP_URL:
        await msg.answer("⚠️ WEBAPP_URL sozlanmagan"); return
    uid = msg.from_user.id
    # uid ni URL ga qo'shamiz — Mini App ichida initDataUnsafe ishlamasa ham ishlaydi
    url = f"{WEBAPP_URL}/admin?uid={uid}"
    await msg.answer(
        "👨‍💼 *Admin panel*",
        reply_markup=ik([ib("🖥 Ochish →", web_app=WebAppInfo(url=url))])
    )

@router.message(Command("setball"))
async def cmd_setball(msg: Message):
    if msg.from_user.id not in ADMIN_IDS: return
    try:
        val = msg.text.split()[1]
        await update_setting("ball_price", val)
        await msg.answer(f"✅ 1 ball = *{val} so\'m*")
    except Exception: await msg.answer("❌ /setball 2000")

@router.message(Command("setelon"))
async def cmd_setelon(msg: Message):
    if msg.from_user.id not in ADMIN_IDS: return
    try:
        val = msg.text.split()[1]
        await update_setting("elon_price", val)
        await msg.answer(f"✅ 1 e\'lon = *{val} ball*")
    except Exception: await msg.answer("❌ /setelon 1")

@router.message(Command("setcard"))
async def cmd_setcard(msg: Message):
    if msg.from_user.id not in ADMIN_IDS: return
    try:
        val = msg.text.split()[1]
        await update_setting("card_number", val)
        await msg.answer(f"✅ Karta: `{val}`")
    except Exception: await msg.answer("❌ /setcard 9860...")

@router.message(Command("broadcast"))
async def cmd_broadcast(msg: Message):
    if msg.from_user.id not in ADMIN_IDS: return
    parts = msg.text.split(maxsplit=1)
    if len(parts) < 2: await msg.answer("Format: /broadcast Matn"); return
    users = await db_all("SELECT id FROM users WHERE is_blocked=0")
    sent = 0
    for u in users:
        try: await bot.send_message(u["id"], parts[1]); sent += 1; await asyncio.sleep(0.05)
        except Exception: pass
    await msg.answer(f"✅ Yuborildi: *{sent}* ta")

@router.message(Command("debug"))
async def cmd_debug(msg: Message):
    if msg.from_user.id not in ADMIN_IDS:
        return
    try:
        total_u   = (await db_get("SELECT COUNT(*) as c FROM users"))["c"]
        total_p   = (await db_get("SELECT COUNT(*) as c FROM products"))["c"]
        active_p  = (await db_get("SELECT COUNT(*) as c FROM products WHERE COALESCE(is_active,1)=1"))["c"]
        total_s   = (await db_get("SELECT COUNT(*) as c FROM shops"))["c"]
        active_s  = (await db_get("SELECT COUNT(*) as c FROM shops WHERE status='active'"))["c"]
        # is_active ustuni bormi?
        try:
            test = await db_get("SELECT is_active FROM products LIMIT 1")
            ia_col = f"✅ is_active: {test['is_active'] if test else 'NULL'}"
        except Exception as e:
            ia_col = f"❌ is_active ustun yo'q: {e}"
        # price ustuni variantlarda bormi?
        try:
            test2 = await db_get("SELECT price FROM product_variants LIMIT 1")
            pv_col = f"✅ variant.price: {test2['price'] if test2 else 'NULL'}"
        except Exception as e:
            pv_col = f"❌ variant.price yo'q: {e}"
        # Catalog query test
        test_rows = await db_all(
            "SELECT p.id, p.name, COALESCE(p.is_active,1) as ia, s.status "
            "FROM products p JOIN shops s ON p.shop_id=s.id LIMIT 3"
        )
        prod_sample = "\n".join([
            f"  #{r['id']} {r['name'][:20]} ia={r['ia']} shop={r['status']}"
            for r in test_rows
        ]) or "  (bo'sh)"

        await msg.answer(
            f"🔍 *Debug Info*\n\n"
            f"👥 Users: {total_u}\n"
            f"📦 Products: {total_p} (aktiv: {active_p})\n"
            f"🏪 Shops: {total_s} (aktiv: {active_s})\n\n"
            f"{ia_col}\n"
            f"{pv_col}\n\n"
            f"📋 Mahsulot namuna:\n{prod_sample}"
        )
    except Exception as e:
        await msg.answer(f"❌ Debug xato: {e}")

@router.message(Command("help"))
async def cmd_help(msg: Message):
    uid = msg.from_user.id
    u   = await get_user(uid)
    if not WEBAPP_URL:
        await msg.answer(
            "📖 *Yordam*\n\n"
            "Klinika: ehtiyoj yozing → takliflar keling → eng arzonni tanlang\n"
            "Sotuvchi: buyurtmalarni ko\'ring → narx kiriting → qabul kuting"
        )
        return
    role = (u["role"] if u else "") or "clinic"
    url  = f"{WEBAPP_URL}/help?role={role}"
    await msg.answer(
        "📖 *Qo\'llanma*\n\nBot qanday ishlashini o\'rganing:",
        reply_markup=ik([ib("📖 Qo\'llanmani ochish →", web_app=WebAppInfo(url=url))])
    )

# ── /debug ───────────────────────────────────────────────────────────────────
# ── REKLAMA E'LON TIZIMI ─────────────────────────────────────────────────────
# Narxlar: Toshkent shahri 200 ball, boshqa 50 ball
# 2 auditoriya (klinika+zubtex) tanlansa — 2x

AD_REGION_PRICES = {
    "Toshkent shahri": 200,
}
AD_REGION_DEFAULT = 50

def calc_ad_price(regions: list, audiences: list) -> tuple:
    """Ball narxini hisoblaydi. (jami, izoh)"""
    multiplier = 2 if len(audiences) >= 2 else 1
    total = 0
    details = []
    for reg in regions:
        base = AD_REGION_PRICES.get(reg, AD_REGION_DEFAULT)
        price = base * multiplier
        total += price
        if multiplier == 2:
            details.append(f"{reg}: {base}×2={price} ball")
        else:
            details.append(f"{reg}: {price} ball")
    return total, details


@router.callback_query(F.data == "ad_start")
async def ad_start(call: CallbackQuery, state: FSMContext):
    uid = call.from_user.id
    u   = await get_user(uid)
    bal = u["balance"] or 0 if u else 0
    await state.set_state(AdState.audience)
    await state.update_data(ad_audiences=[], ad_regions=[])
    await call.message.answer(
        f"📢 *E\'lon berish*\n\n"
        f"💰 Balansingiz: *{bal:.0f} ball*\n\n"
        f"*1-qadam: Kimga e\'lon berasiz?*",
        reply_markup=_ad_audience_kb([])
    )
    await call.answer()

@router.callback_query(F.data.startswith("adaud_"), AdState.audience)
async def ad_aud_toggle(call: CallbackQuery, state: FSMContext):
    key = call.data[6:]
    d   = await state.get_data()
    sel = list(d.get("ad_audiences", []))
    if key in sel: sel.remove(key)
    else: sel.append(key)
    await state.update_data(ad_audiences=sel)
    await call.message.edit_reply_markup(reply_markup=_ad_audience_kb(sel))
    await call.answer()

@router.callback_query(F.data == "adaud_next", AdState.audience)
async def ad_aud_next(call: CallbackQuery, state: FSMContext):
    d   = await state.get_data()
    sel = d.get("ad_audiences", [])
    if not sel:
        await call.answer("⚠️ Kamida 1 ta tanlang!", show_alert=True)
        return
    await state.set_state(AdState.regions)
    await call.message.answer(
        "📍 *2-qadam: Qaysi hududlar?*\n\n"
        "_🏙 Toshkent shahri = 200 ball\n"
        "Boshqa viloyatlar = 50 ball\n"
        "2 auditoriya = narx × 2_",
        reply_markup=_ad_regions_kb([])
    )
    await call.answer()

@router.callback_query(F.data.startswith("adreg_"), AdState.regions)
async def ad_reg_toggle(call: CallbackQuery, state: FSMContext):
    key = call.data[6:]
    d   = await state.get_data()
    sel = list(d.get("ad_regions", []))

    if key == "all":
        sel = [r for r,_ in _AD_REGIONS]
    elif key == "clear":
        sel = []
    elif key == "next":
        if not sel:
            await call.answer("⚠️ Kamida 1 ta hudud tanlang!", show_alert=True)
            return
        await state.update_data(ad_regions=sel)
        await state.set_state(AdState.content)
        auds  = d.get("ad_audiences", [])
        total, reason = _calc_ad_price(auds, sel)
        aud_n = {"aud_clinic":"Klinikalar","aud_zubtex":"Zubtexniklar","aud_seller":"Sotuvchilar"}
        aud_txt = ", ".join(aud_n.get(a,"?") for a in auds)
        await call.message.answer(
            f"✍️ *3-qadam: E\'lon mazmunini yuboring*\n\n"
            f"• Matn (reklama matni)\n"
            f"• Rasm + izoh\n"
            f"• @username yoki t.me/... link\n\n"
            f"⚠️ _Faqat Telegram linklarga ruxsat_\n\n"
            f"📊 *Narx:*\n"
            f"👥 {aud_txt}\n"
            f"📍 {len(sel)} ta hudud · {reason}\n"
            f"💰 *Jami: {total} ball*",
            reply_markup=ik([ib("◀️ Orqaga", "adreg_back")])
        )
        await call.answer()
        return
    elif key == "back":
        await state.set_state(AdState.audience)
        d2  = await state.get_data()
        sel2 = d2.get("ad_audiences", [])
        await call.message.answer(
            "*1-qadam: Kimga e\'lon berasiz?*",
            reply_markup=_ad_audience_kb(sel2)
        )
        await call.answer()
        return
    else:
        if key in sel: sel.remove(key)
        else: sel.append(key)

    await state.update_data(ad_regions=sel)
    await call.message.edit_reply_markup(reply_markup=_ad_regions_kb(sel))
    await call.answer()

@router.message(AdState.content, F.text | F.photo)
async def ad_content_input(msg: Message, state: FSMContext):
    d        = await state.get_data()
    audiences= d.get("ad_audiences", [])
    regions  = d.get("ad_regions", [])
    total, reason = _calc_ad_price(audiences, regions)

    # Kontent olish
    if msg.photo:
        photo_id = msg.photo[-1].file_id
        raw_text = msg.caption or ""
    else:
        photo_id = None
        raw_text = msg.text or ""

    # Link himoya
    clean_text = _sanitize_ad_text(raw_text)
    was_cleaned = clean_text != raw_text

    await state.update_data(
        ad_photo=photo_id,
        ad_text=clean_text,
        ad_raw_text=raw_text,
    )

    uid = msg.from_user.id
    u   = await get_user(uid)
    bal = u["balance"] or 0 if u else 0

    aud_n   = {"aud_clinic":"🏥 Klinikalar","aud_zubtex":"🔬 Zubtexniklar","aud_seller":"🛒 Sotuvchilar"}
    aud_txt = ", ".join(aud_n.get(a,"?") for a in audiences)

    warn = "\n⚠️ _Xavfli linklar o\'chirildi_" if was_cleaned else ""
    bal_warn = f"\n❌ *Balans yetarli emas!* ({bal:.0f}/{total} ball)" if bal < total else ""

    preview_txt = (
        f"👁 *Ko\'rib chiqing:*{warn}\n\n"
        f"📢 *Reklama*\n\n"
        f"{clean_text}\n\n"
        f"👥 {aud_txt}\n"
        f"📍 {len(regions)} ta hudud\n"
        f"💡 {reason}\n"
        f"💰 *{total} ball* (sizda: {bal:.0f}){bal_warn}"
    )

    if bal < total:
        kb = ik(
            [ib("➕ Hisob to\'ldirish", "topup")],
            [ib("◀️ Orqaga", "ad_edit"), ib("✖️ Bekor", "ad_cancel")],
        )
    else:
        kb = ik(
            [ib("✅ Adminga yuborish", "ad_confirm")],
            [ib("✏️ Tahrirlash", "ad_edit"), ib("✖️ Bekor", "ad_cancel")],
        )

    await state.set_state(AdState.confirm)
    if photo_id:
        await msg.answer_photo(photo_id, caption=preview_txt, reply_markup=kb)
    else:
        await msg.answer(preview_txt, reply_markup=kb)

@router.callback_query(F.data == "ad_confirm", AdState.confirm)
async def ad_confirm(call: CallbackQuery, state: FSMContext):
    d         = await state.get_data()
    audiences = d.get("ad_audiences", [])
    regions   = d.get("ad_regions", [])
    total, reason = _calc_ad_price(audiences, regions)
    uid       = call.from_user.id
    u         = await get_user(uid)
    bal       = u["balance"] or 0 if u else 0

    if bal < total:
        await call.answer("❌ Balans yetarli emas!", show_alert=True)
        return

    photo_id = d.get("ad_photo")
    ad_text  = d.get("ad_text", "")

    aud_n   = {"aud_clinic":"Klinikalar","aud_zubtex":"Zubtexniklar","aud_seller":"Sotuvchilar"}
    aud_txt = ", ".join(aud_n.get(a,"?") for a in audiences)

    # Adminga preview + tasdiqlash tugmasi
    ad_id = await db_insert(
        "INSERT INTO transactions(user_id,amount,balls,type,note,status) VALUES(?,?,?,\'ad\',?,\'pending\')",
        (uid, total, total, f"audiences:{','.join(audiences)}|regions:{','.join(regions)}")
    )

    admin_txt = (
        f"📢 *Yangi reklama so\'rovi #{ad_id}*\n\n"
        f"👤 {u['clinic_name'] or u['full_name'] or str(uid)}\n"
        f"👥 {aud_txt}\n"
        f"📍 {', '.join(regions[:3])}{'...' if len(regions)>3 else ''}\n"
        f"💰 {total} ball\n\n"
        f"Mazmun:\n{ad_text}"
    )
    admin_kb = ik(
        [ib("✅ Tasdiqlash", f"adm_ad_ok_{ad_id}_{uid}"),
         ib("❌ Rad etish",  f"adm_ad_rej_{ad_id}_{uid}")],
    )

    for aid in ADMIN_IDS:
        try:
            if photo_id:
                await bot.send_photo(aid, photo_id, caption=admin_txt, reply_markup=admin_kb)
            else:
                await bot.send_message(aid, admin_txt, reply_markup=admin_kb)
        except Exception: pass

    await state.clear()
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        f"⏳ *E\'lon adminga yuborildi!*\n\n"
        f"Admin tasdiqlagandan keyin:\n"
        f"1. @xazdent kanalga chiqadi\n"
        f"2. Keyin {aud_txt} ga tarqatiladi\n\n"
        f"💰 Tasdiqlanganda *{total} ball* hisobdan ayiriladi."
    )
    await call.answer("✅")

@router.callback_query(F.data.startswith("adm_ad_ok_"))
async def adm_ad_ok(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS: return
    parts  = call.data.split("_")
    ad_id  = int(parts[3])
    uid    = int(parts[4])

    tx = await db_get("SELECT * FROM transactions WHERE id=?", (ad_id,))
    if not tx:
        await call.answer("Topilmadi", show_alert=True); return

    u   = await get_user(uid)
    bal = u["balance"] or 0 if u else 0
    total = tx["balls"]

    if bal < total:
        await call.message.edit_caption(
            (call.message.caption or "") + "\n\n❌ Foydalanuvchi balansi yetarli emas!",
            reply_markup=None)
        await call.answer(); return

    # Baldan ayiramiz
    await db_run("UPDATE users SET balance=balance-? WHERE id=?", (total, uid))
    await db_run("UPDATE transactions SET status=\'confirmed\',confirmed_by=? WHERE id=?",
                 (call.from_user.id, ad_id))

    # Note dan audiences va regions olamiz
    note  = tx["note"] or ""
    auds  = []
    regs  = []
    for part in note.split("|"):
        if part.startswith("audiences:"):
            auds = part[10:].split(",") if part[10:] else []
        elif part.startswith("regions:"):
            regs = part[8:].split(",") if part[8:] else []

    # Kanalga post
    photo_id = None
    ad_text  = ""
    # Caption dan olamiz
    orig_caption = call.message.caption or call.message.text or ""
    # Mazmun qismini olamiz
    if "Mazmun:" in orig_caption:
        ad_text = orig_caption.split("Mazmun:")[-1].strip()
    # Rasm bor bo'lsa
    if call.message.photo:
        photo_id = call.message.photo[-1].file_id

    chan_txt = f"📢 *Reklama*\n\n{ad_text}\n\n_@XazdentBot orqali_"
    chan_msg = None
    try:
        if photo_id:
            chan_msg = await bot.send_photo(CHANNEL_ID, photo_id, caption=chan_txt)
        else:
            chan_msg = await bot.send_message(CHANNEL_ID, chan_txt)
    except Exception as e:
        log.error(f"Kanal reklama post xato: {e}")

    chan_msg_id = chan_msg.message_id if chan_msg else None

    # Kanalga borgandan keyin forward qilamiz
    aud_roles = {
        "aud_clinic": ("clinic", "zubtex"),
        "aud_zubtex": ("zubtex",),
        "aud_seller": ("seller",),
    }
    roles = list(set(r for a in auds for r in aud_roles.get(a, [])))
    sent = 0
    if roles and regs:
        role_ph  = ",".join(["?" for _ in roles])
        reg_ph   = ",".join(["?" for _ in regs])
        recipients = await db_all(
            f"SELECT id FROM users WHERE role IN ({role_ph}) "
            f"AND region IN ({reg_ph}) AND is_blocked=0 AND id!=?",
            (*roles, *regs, uid)
        )
        for r in recipients:
            try:
                if chan_msg:
                    await bot.forward_message(r["id"], CHANNEL_ID, chan_msg_id)
                else:
                    if photo_id:
                        await bot.send_photo(r["id"], photo_id, caption=chan_txt)
                    else:
                        await bot.send_message(r["id"], chan_txt)
                sent += 1
                await asyncio.sleep(0.05)
            except Exception: pass

    # Reklama beruvchiga xabar
    try:
        await bot.send_message(uid,
            f"✅ *E\'loningiz tasdiqlandi!*\n\n"
            f"📢 Kanal: @xazdent\n"
            f"📨 {sent} ta foydalanuvchiga yuborildi\n"
            f"💰 -{total:.0f} ball ayirildi\n\n"
            f"{'👁 Kanal postini ko\'rishlar soni uchun: @xazdent' if chan_msg_id else ''}"
        )
    except Exception: pass

    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(f"✅ Yuborildi: {sent} kishi | Kanal post: {chan_msg_id}")
    await call.answer("✅")

@router.callback_query(F.data.startswith("adm_ad_rej_"))
async def adm_ad_rej(call: CallbackQuery):
    if call.from_user.id not in ADMIN_IDS: return
    parts = call.data.split("_")
    ad_id = int(parts[3])
    uid   = int(parts[4])
    await db_run("UPDATE transactions SET status=\'rejected\',confirmed_by=? WHERE id=?",
                 (call.from_user.id, ad_id))
    try:
        await bot.send_message(uid,
            "❌ *E\'loningiz rad etildi.*\n\n"
            "Sabab: admin tomonidan qabul qilinmadi.\n"
            "Ball ayirilmadi.")
    except Exception: pass
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer("❌ Rad etildi.")
    await call.answer()

@router.callback_query(F.data == "ad_edit")
async def ad_edit(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdState.content)
    await call.message.answer(
        "✍️ *Yangi mazmunni yuboring:*\n"
        "Matn, rasm yoki rasm + izoh"
    )
    await call.answer()

@router.callback_query(F.data == "ad_cancel")
async def ad_cancel(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("✖️ E\'lon bekor qilindi.")
    await call.answer()

# ── TAKLIF QABUL HELPER FUNKSIYALAR ─────────────────────────────────────────

async def _send_order_to_group(shop, order_id: int, msg_txt: str,
                               buyer_id: int, total: float):
    """Agar do'konda guruh bog'liq bo'lsa, guruhga ham yuboradi."""
    if not shop or not shop.get("group_chat_id"):
        return
    group_id = shop["group_chat_id"]
    claim_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="✋ Men olaman",
            callback_data=f"claim_order_{order_id}"
        )
    ]])
    try:
        await bot.send_message(group_id, msg_txt, reply_markup=claim_kb)
    except Exception as e:
        log.error(f"Guruhga yuborish xato: {e}")

async def _notify_winner(seller_id: int, clinic: dict, items: list):
    """G'olib sotuvchiga to'liq xabar: mahsulotlar + klinika ma'lumoti."""
    cname   = clinic["clinic_name"] or clinic["full_name"] or "Klinika"
    cphone  = clinic["phone"] or "—"
    cregion = clinic["region"] or "—"
    caddr   = clinic["address"] or "—"
    clat    = clinic.get("latitude")
    clon    = clinic.get("longitude")

    # Mahsulotlar ro'yxati
    lines  = []
    total  = 0
    for i, (name, qty, unit, price) in enumerate(items, 1):
        subtotal = price * qty
        total   += subtotal
        lines.append(
            f"{i}. *{name}* — {qty} {unit}\n"
            f"   {price:,.0f} × {qty} = *{subtotal:,.0f} so'm*"
        )
    items_txt = "\n".join(lines)

    # Sotuvchining to'lov usullari
    seller_u = await get_user(seller_id)
    pay_icons = {"p2p":"💳 P2P","cash":"💵 Naqd","bank":"🏦 Hisob raqam"}
    spm_raw = (seller_u.get("payment_methods") or "") if seller_u else ""
    spm_txt = " · ".join(pay_icons[p] for p in spm_raw.split(",") if p in pay_icons)
    spm_line = f"\n💳 To\'lov: {spm_txt}" if spm_txt else ""

    txt = (
        f"🎉 *Taklifingiz qabul qilindi!*\n\n"
        f"📦 *Buyurtma:*\n{items_txt}\n\n"
        f"💰 *Jami: {total:,.0f} so\'m*{spm_line}\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🏥 *{cname}*\n"
        f"📞 {cphone}\n"
        f"📍 {cregion}\n"
        f"🏠 {caddr}"
    )

    try:
        await bot.send_message(seller_id, txt)
    except Exception as e:
        log.error(f"Winner notify xato: {e}")
        return

    # Lokatsiya yuboramiz (agar bo'lsa)
    if clat and clon:
        try:
            await bot.send_location(seller_id, latitude=clat, longitude=clon)
        except Exception:
            pass

async def _notify_loser(seller_id: int, product_name: str, win_price: float,
                        my_price: float, unit: str):
    """Yutqazgan sotuvchiga faqat narx statistikasi."""
    diff = my_price - win_price
    pct  = abs(diff) / win_price * 100 if win_price else 0
    try:
        await bot.send_message(
            seller_id,
            f"📊 *{product_name}* bo\'yicha boshqa taklif qabul qilindi.\n\n"
            f"Qabul qilingan narx: *{win_price:,.0f} so\'m/{unit}*\n"
            f"Sizning narxingiz: *{my_price:,.0f} so\'m/{unit}*\n"
            f"Farq: *{diff:+,.0f} so\'m ({pct:.0f}%)*\n\n"
            f"_Xaridor ma\'lumotlari maxfiy._"
        )
    except Exception:
        pass

# ── TEJASH STATISTIKASI ──────────────────────────────────────────────────────
@router.message(F.text == "📊 Tejash")
async def savings_stats(msg: Message):
    uid = msg.from_user.id
    await _show_savings(uid, msg)

async def _show_savings(uid: int, target_msg):
    now   = datetime.now()
    month = now.strftime("%Y-%m")
    year  = now.strftime("%Y")

    # Qabul qilingan takliflar + shu buyurtmadagi barcha takliflar
    accepted = await db_all(
        "SELECT o.need_id, o.price as won_price, n.quantity, n.product_name, n.unit, "
        "o.created_at "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE n.owner_id=? AND o.status=\'accepted\'",
        (uid,)
    )
    if not accepted:
        await target_msg.answer(
            "📊 *Tejash statistikasi*\n\nHali qabul qilingan taklif yo\'q.",
            reply_markup=ik([ib("📥 Excel", "savings_excel")])
        )
        return

    # Har qabul qilingan taklif uchun o'sha need dagi MAX narxni topamiz
    rows = []
    total_saved_month = 0
    total_saved_year  = 0
    total_saved_all   = 0

    for a in accepted:
        max_row = await db_get(
            "SELECT MAX(price) as max_p FROM offers WHERE need_id=? AND price>0",
            (a["need_id"],)
        )
        max_p = max_row["max_p"] if max_row and max_row["max_p"] else a["won_price"]
        saved_per_unit = max_p - a["won_price"]
        saved_total    = saved_per_unit * a["quantity"]
        pct = saved_per_unit / max_p * 100 if max_p > 0 else 0
        created = a["created_at"] or ""
        rows.append({
            "name":       a["product_name"],
            "qty":        a["quantity"],
            "unit":       a["unit"],
            "won":        a["won_price"],
            "max":        max_p,
            "saved":      saved_total,
            "pct":        pct,
            "month":      created[:7],
            "year":       created[:4],
        })
        if created.startswith(month): total_saved_month += saved_total
        if created.startswith(year):  total_saved_year  += saved_total
        total_saved_all += saved_total

    # Top 5 tejash
    top5 = sorted(rows, key=lambda x: x["saved"], reverse=True)[:5]

    txt = (
        f"📊 *Tejash statistikasi*\n\n"
        f"💰 Bu oy:   *{total_saved_month:>12,.0f} so\'m*\n"
        f"💰 Bu yil:  *{total_saved_year:>12,.0f} so\'m*\n"
        f"💰 Jami:    *{total_saved_all:>12,.0f} so\'m*\n\n"
        f"_Bot orqali eng arzon taklifni tanlab tejadingiz_\n\n"
        f"🏆 *Top tejamlar:*\n"
    )
    for i, r in enumerate(top5, 1):
        txt += (
            f"{i}. *{r['name']}*\n"
            f"   {r['won']:,.0f} vs {r['max']:,.0f} so\'m/ta\n"
            f"   Tejaldi: *{r['saved']:,.0f} so\'m ({r['pct']:.0f}%)*\n"
        )

    await target_msg.answer(txt, reply_markup=ik(
        [ib("📥 Excel yuklab olish", "savings_excel")],
    ))

@router.callback_query(F.data == "savings_excel")
async def savings_excel(call: CallbackQuery):
    await call.answer("⏳ Tayyorlanmoqda...")
    uid  = call.from_user.id
    path = await _build_savings_excel(uid)
    if not path:
        await call.message.answer("❌ Excel yaratib bo\'lmadi")
        return
    import aiofiles
    async with aiofiles.open(path, "rb") as f:
        data = await f.read()
    fname = f"tejash_{datetime.now().strftime('%Y%m%d')}.xlsx"
    await call.message.answer_document(
        document=BufferedInputFile(data, filename=fname),
        caption=f"📊 Tejash hisoboti — {datetime.now().strftime('%d.%m.%Y')}"
    )
    try:
        import os as _os; _os.remove(path)
    except Exception: pass

async def _build_savings_excel(uid: int) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb  = openpyxl.Workbook()
    now = datetime.now()

    def hdr(ws, cols, color="1565C0"):
        fill = PatternFill("solid", fgColor=color)
        font = Font(bold=True, color="FFFFFF")
        for i, v in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=v)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center")
    def aw(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)

    # Qabul qilingan takliflar
    accepted = await db_all(
        "SELECT o.need_id, o.price as won_price, n.quantity, n.product_name, "
        "n.unit, o.created_at "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE n.owner_id=? AND o.status=\'accepted\' ORDER BY o.created_at DESC",
        (uid,)
    )

    ws1 = wb.active; ws1.title = "Tejash tarixi"
    hdr(ws1, ["Sana","Mahsulot","Miqdor","Birlik",
              "To\'langan narx","Eng qimmat narx","Tejaldi (so\'m)","Tejaldi (%)"])
    green = PatternFill("solid", fgColor="E2EFDA")
    for i, a in enumerate(accepted, 2):
        max_row = await db_get(
            "SELECT MAX(price) as m FROM offers WHERE need_id=? AND price>0",
            (a["need_id"],)
        )
        max_p   = max_row["m"] if max_row and max_row["m"] else a["won_price"]
        saved   = (max_p - a["won_price"]) * a["quantity"]
        pct     = (max_p - a["won_price"]) / max_p * 100 if max_p > 0 else 0
        ws1.cell(row=i, column=1, value=(a["created_at"] or "")[:10])
        ws1.cell(row=i, column=2, value=a["product_name"])
        ws1.cell(row=i, column=3, value=a["quantity"])
        ws1.cell(row=i, column=4, value=a["unit"])
        ws1.cell(row=i, column=5, value=a["won_price"])
        ws1.cell(row=i, column=6, value=max_p)
        c7 = ws1.cell(row=i, column=7, value=round(saved))
        c8 = ws1.cell(row=i, column=8, value=round(pct, 1))
        if saved > 0:
            c7.fill = green; c8.fill = green

    # Jami qator
    if accepted:
        r = len(accepted) + 2
        ws1.cell(row=r, column=6, value="JAMI:").font = Font(bold=True)
        total = sum(
            ((await db_get("SELECT MAX(price) as m FROM offers WHERE need_id=? AND price>0",
                           (a["need_id"],)) or {}).get("m", a["won_price"]) - a["won_price"])
            * a["quantity"] for a in accepted
        )
        ws1.cell(row=r, column=7, value=round(total)).font = Font(bold=True)
    aw(ws1)

    # Oylik tejash
    ws2 = wb.create_sheet("Oylik")
    hdr(ws2, ["Oy","Bitimlar","Jami tejaldi (so\'m)"], "2E7D32")
    monthly = await db_all(
        "SELECT SUBSTR(o.created_at,1,7) as month, COUNT(*) as cnt "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE n.owner_id=? AND o.status=\'accepted\' "
        "GROUP BY month ORDER BY month DESC LIMIT 24",
        (uid,)
    )
    for i, m in enumerate(monthly, 2):
        ws2.cell(row=i, column=1, value=m["month"])
        ws2.cell(row=i, column=2, value=m["cnt"])
        ws2.cell(row=i, column=3, value="—")
    aw(ws2)

    path = os.path.join(BASE_DIR, f"tejash_{uid}_{now.strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(path)
    return path

@router.message(F.text == "📖 Yordam")
async def menu_help(msg: Message):
    uid  = msg.from_user.id
    u    = await get_user(uid)
    role = (u["role"] if u else "") or "clinic"
    if not WEBAPP_URL:
        await msg.answer("📖 Yordam: /help buyrug'ini yuboring")
        return
    url = f"{WEBAPP_URL}/help?role={role}"
    await msg.answer(
        "📖 *Qo\'llanma*",
        reply_markup=ik([ib("📖 Ochish →", web_app=WebAppInfo(url=url))])
    )

# ── SUPPORT TIZIMI ───────────────────────────────────────────────────────────
class SupportState(StatesGroup):
    waiting_message = State()
    waiting_reply   = State()

class PhotoOrderState(StatesGroup):
    waiting_photo   = State()
    waiting_caption = State()

class ReviewState(StatesGroup):
    waiting_comment = State()

class ComplaintState(StatesGroup):
    waiting_reason = State()

@router.callback_query(F.data == "support_start")
async def support_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(SupportState.waiting_message)
    await call.message.answer(
        "💬 *Yordam xizmati*\n\n"
        "Savolingizni yoki muammoingizni yozing.\n"
        "Admin imkon qadar tez javob beradi.\n\n"
        "_/cancel — bekor qilish_"
    )
    await call.answer()

@router.message(SupportState.waiting_message)
async def support_message_received(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear()
        await msg.answer("❌ Bekor qilindi.")
        return

    text = msg.text or (msg.caption if msg.caption else "")
    if not text:
        await msg.answer("⚠️ Faqat matn yuboring.")
        return

    uid = msg.from_user.id
    u   = await get_user(uid)
    name = (u["clinic_name"] or u["full_name"] or str(uid)) if u else str(uid)

    # DB ga saqlaymiz
    mid = await db_insert(
        "INSERT INTO support_messages(user_id, message) VALUES(?,?)",
        (uid, text)
    )

    await state.clear()
    await msg.answer(
        "✅ *Xabaringiz qabul qilindi!*\n\n"
        "Admin tez orada javob beradi."
    )

    # Adminga xabar
    admin_txt = (
        f"💬 *Yangi support xabari #{mid}*\n\n"
        f"👤 {name} (ID: {uid})\n"
        f"📱 {u['phone'] or '—' if u else '—'}\n\n"
        f"📝 {text}"
    )
    kb = ik([ib(f"💬 Javob berish", f"reply_{mid}_{uid}")])
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(aid, admin_txt, reply_markup=kb)
        except Exception:
            pass

@router.callback_query(F.data.startswith("reply_"))
async def admin_reply_start(call: CallbackQuery, state: FSMContext):
    if call.from_user.id not in ADMIN_IDS: return
    parts  = call.data.split("_")
    msg_id = int(parts[1])
    uid    = int(parts[2])
    await state.set_state(SupportState.waiting_reply)
    await state.update_data(support_msg_id=msg_id, support_user_id=uid)
    await call.message.answer(
        f"✍️ *Javobingizni yozing* (xabar #{msg_id}):\n"
        f"_/cancel — bekor qilish_"
    )
    await call.answer()

@router.message(SupportState.waiting_reply)
async def admin_reply_send(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear()
        await msg.answer("❌ Bekor qilindi.")
        return

    d      = await state.get_data()
    mid    = d.get("support_msg_id")
    uid    = d.get("support_user_id")
    reply  = msg.text or ""
    admin  = msg.from_user.id

    await db_run(
        "UPDATE support_messages SET admin_reply=?, status='replied', "
        "replied_at=to_char(now(),'YYYY-MM-DD HH24:MI:SS'), admin_id=? WHERE id=?",
        (reply, admin, mid)
    )
    await state.clear()

    # Foydalanuvchiga javob
    try:
        await bot.send_message(
            uid,
            f"💬 *Yordam xizmatidan javob*\n\n{reply}"
        )
        await msg.answer(f"✅ Javob yuborildi.")
    except Exception as e:
        await msg.answer(f"❌ Yuborib bo'lmadi: {e}")

# Profildan support ochish
@router.callback_query(F.data == "open_support")
async def open_support(call: CallbackQuery, state: FSMContext):
    await state.set_state(SupportState.waiting_message)
    await call.message.answer(
        "💬 *Yordam xizmati*\n\n"
        "Savolingizni yozing — admin javob beradi.",
        reply_markup=ik([ib("✖️ Bekor", "cancel_support")])
    )
    await call.answer()

@router.callback_query(F.data == "cancel_support")
async def cancel_support(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.delete()
    await call.answer()

# ── RASM BILAN BUYURTMA (bot orqali) ─────────────────────────────────────────
@router.message(F.photo, F.func(lambda msg: True))
async def handle_photo_message(msg: Message, state: FSMContext):
    """Foydalanuvchi botga rasm yuborsa — buyurtma berish imkonini taklif qilamiz."""
    cur_state = await state.get_state()
    # Boshqa state da bo'lsa — bu handlerni ishlatmaymiz
    if cur_state and cur_state not in (None, ""):
        return

    uid = msg.from_user.id
    u   = await get_user(uid)
    if not u or u["role"] not in ("clinic", "zubtex"):
        return

    file_id = msg.photo[-1].file_id
    await state.set_state(PhotoOrderState.waiting_caption)
    await state.update_data(order_photo_id=file_id)
    await msg.answer(
        "📷 *Rasm qabul qilindi!*\n\n"
        "Endi qisqacha yozing:\n"
        "_Mahsulot nomi, miqdori — masalan:_\n"
        "`GC Fuji 2 kg, Xarizma A2 5 dona`\n\n"
        "/cancel — bekor qilish",
        reply_markup=ik([ib("✖️ Bekor qilish", "cancel_photo_order")])
    )

@router.message(PhotoOrderState.waiting_caption, F.text)
async def photo_order_caption(msg: Message, state: FSMContext):
    if msg.text.startswith("/"):
        await state.clear()
        await msg.answer("❌ Bekor qilindi.")
        return

    d        = await state.get_data()
    photo_id = d.get("order_photo_id")
    uid      = msg.from_user.id
    u        = await get_user(uid)

    # Matndan mahsulotlarni oddiy parse qilamiz
    caption  = msg.text.strip()
    items    = []
    for line in caption.replace(",", "\n").split("\n"):
        line = line.strip()
        if not line: continue
        parts = line.split()
        if len(parts) >= 2:
            # Raqamni topamiz
            qty = None
            unit = "dona"
            name_parts = []
            for p in parts:
                p_clean = p.replace(",","").replace(".","")
                if p_clean.isdigit() and qty is None:
                    qty = float(p_clean)
                elif p_clean in ["kg","litr","quti","paket","ml","gr","dona"] and qty is not None:
                    unit = p_clean
                else:
                    name_parts.append(p)
            if name_parts and qty:
                items.append({"name":" ".join(name_parts),"qty":qty,"unit":unit})
            elif name_parts:
                items.append({"name":" ".join(name_parts),"qty":1,"unit":"dona"})
        elif line:
            items.append({"name":line,"qty":1,"unit":"dona"})

    if not items:
        await msg.answer("⚠️ Mahsulot nomini aniqlab bo'lmadi. Qayta yuboring.")
        return

    # Deadline tanlash
    await state.update_data(photo_items=items, photo_caption=caption)
    preview = "\n".join([f"• {it['qty']} {it['unit']} — {it['name']}" for it in items])
    await msg.answer(
        f"✅ *Topildi:*\n{preview}\n\n"
        f"Qachongacha kerak?",
        reply_markup=ik(
            [ib("⚡️ 2 soat",  "pod_2"),   ib("🕐 24 soat", "pod_24")],
            [ib("📅 3 kun",    "pod_72"),   ib("🗓 1 hafta", "pod_168")],
            [ib("📆 10 kun",   "pod_240")],
            [ib("✖️ Bekor",    "cancel_photo_order")],
        )
    )

@router.callback_query(F.data.startswith("pod_"), PhotoOrderState.waiting_caption)
async def photo_order_deadline(call: CallbackQuery, state: FSMContext):
    deadline = int(call.data[4:])
    d        = await state.get_data()
    items    = d.get("photo_items", [])
    photo_id = d.get("order_photo_id")
    uid      = call.from_user.id
    u        = await get_user(uid)

    if not items or not u:
        await call.answer("Xato", show_alert=True)
        await state.clear()
        return

    room    = await get_or_create_room(uid)
    expires = (datetime.now() + timedelta(hours=deadline)).isoformat()
    batch_id = await db_insert(
        "INSERT INTO batches(owner_id,deadline_hours,expires_at) VALUES(?,?,?)",
        (uid, deadline, expires)
    )
    saved = []
    for item in items:
        nid = await db_insert(
            "INSERT INTO needs(batch_id,room_id,owner_id,product_name,quantity,unit,"
            "deadline_hours,expires_at,photo_file_id) VALUES(?,?,?,?,?,?,?,?,?)",
            (batch_id, room["id"], uid, item["name"],
             float(item["qty"]), item["unit"], deadline, expires, photo_id)
        )
        need = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
        saved.append(dict(need))

    mid = await post_batch_to_channel(batch_id, saved, dict(u), photo_id)
    if mid:
        for n in saved:
            await db_run("UPDATE needs SET channel_message_id=? WHERE id=?", (mid, n["id"]))

    dl_map = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
    await state.clear()
    chan = CHANNEL_ID.lstrip("@") if isinstance(CHANNEL_ID, str) else str(CHANNEL_ID)
    link = f"\n[Kanalda ko\'rish](https://t.me/{chan}/{mid})" if mid else ""
    await call.message.edit_text(
        f"✅ *{len(saved)} ta mahsulot joylashtirildi!*{link}\n\n"
        f"⏱ {dl_map.get(deadline, str(deadline)+' soat')} ichida"
    )
    asyncio.create_task(notify_sellers_batch(batch_id, uid))
    await call.answer("✅")

@router.callback_query(F.data == "cancel_photo_order")
async def cancel_photo_order(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("❌ Bekor qilindi.")
    await call.answer()

@router.message(F.text == "🛍 Dental Market")
async def dental_market_btn(msg: Message):
    """Dental Market — WebApp reply keyboard orqali ochiladi.
    Bu handler faqat WebApp URL yo'q holatda ishlaydi (fallback)."""
    uid  = msg.from_user.id
    u    = await get_user(uid)
    role = u["role"] if u else "clinic"
    if not WEBAPP_URL:
        await msg.answer("🛍 Dental Market hozircha ishlamayapti.")
        return
    url = f"{WEBAPP_URL}/catalog?uid={uid}&role={role}"
    await msg.answer(
        "🛍 *XazDent Market*",
        reply_markup=ik([ib("🛍 XazDent Market →", web_app=WebAppInfo(url=url))])
    )

@router.message(F.text == "✏️ Buyurtma yozish")
async def buyurtma_yozish_btn(msg: Message):
    """Yangi menyu uchun — eski Ehtiyoj yozish bilan bir xil."""
    uid = msg.from_user.id
    u   = await get_user(uid)
    if not WEBAPP_URL:
        await msg.answer("Buyurtma berish uchun /start bosing.")
        return
    url = f"{WEBAPP_URL}/order?uid={uid}"
    await msg.answer(
        "✏️ *Buyurtma yozish*\n\nKerakli materiallarni kiriting:",
        reply_markup=ik([ib("✏️ Buyurtma yozish →", web_app=WebAppInfo(url=url))])
    )

@router.message(F.text == "🔔 Buyurtmalar")
async def seller_orders_btn(msg: Message):
    """Buyurtmalar — WebApp da ko'rinadi."""
    uid = msg.from_user.id
    if WEBAPP_URL:
        url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller#orders"
        await msg.answer(
            "🔔 *Buyurtmalar*",
            reply_markup=ik([ib("🔔 Buyurtmalarni ko\'rish →",
                               web_app=WebAppInfo(url=url))])
        )
    else:
        await msg.answer("Buyurtmalar bo\'limi hozircha ishlamayapti.")

@router.message(F.text == "➕ Mahsulot qo\'shish")
async def seller_add_product_btn(msg: Message):
    """Mahsulot qo'shish — WebApp reply keyboard orqali ochiladi (fallback)."""
    uid  = msg.from_user.id
    shop = await db_get("SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,))
    if not shop:
        await msg.answer(
            "🏪 Avval do\'kon oching!",
            reply_markup=ik([ib("➕ Do\'kon ochish", "open_shop")])
        )
        return
    if not WEBAPP_URL:
        await msg.answer("Hozircha ishlamayapti.")
        return
    url = f"{WEBAPP_URL}/catalog?uid={uid}&role=seller&action=add"
    await msg.answer(
        "➕ *Mahsulot qo\'shish*",
        reply_markup=ik([ib("➕ Mahsulot qo\'shish →", web_app=WebAppInfo(url=url))])
    )

# ── BUYURTMA TASDIQLASH VA BAHOLASH TIZIMI ───────────────────────────────────

@router.callback_query(F.data.startswith("co_confirm_"))
async def catalog_order_confirm(call: CallbackQuery):
    """Sotuvchi buyurtmani qabul qildi."""
    parts    = call.data.split("_")
    order_id = int(parts[2])
    buyer_id = int(parts[3])
    seller   = call.from_user.id

    await db_run(
        "UPDATE catalog_orders SET status='confirmed', confirmed_at=to_char(now(),'YYYY-MM-DD HH24:MI:SS') "
        "WHERE id=? AND seller_id=?",
        (order_id, seller)
    )

    # Sotuvchiga tasdiqlash xabari
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        f"✅ *Buyurtma #{order_id} qabul qilindi!*\n\n"
        f"Xaridorga xabar yuborildi. 48 soatdan keyin\n"
        f"baholash so\'rovi avtomatik ketadi."
    )

    # Xaridorga xabar
    u = await get_user(seller)
    shop = await db_get("SELECT shop_name FROM shops WHERE owner_id=?", (seller,))
    sname = (shop["shop_name"] if shop else None) or (u["clinic_name"] if u else None) or "Sotuvchi"
    try:
        await bot.send_message(
            buyer_id,
            f"✅ *Buyurtmangiz qabul qilindi!*\n\n"
            f"🏪 *{sname}* buyurtmangizni tasdiqladi va\n"
            f"jo\'natmoqda.\n\n"
            f"_48 soatdan keyin yetib kelganini so\'raymiz._"
        )
    except Exception as e:
        log.error(f"Buyer notify xato: {e}")
    await call.answer("✅ Tasdiqlandi!")

@router.callback_query(F.data.startswith("co_reject_"))
async def catalog_order_reject(call: CallbackQuery):
    """Sotuvchi buyurtmani rad etdi."""
    parts    = call.data.split("_")
    order_id = int(parts[2])
    buyer_id = int(parts[3])

    await db_run(
        "UPDATE catalog_orders SET status='rejected' WHERE id=?", (order_id,)
    )
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(f"❌ Buyurtma #{order_id} rad etildi.")

    try:
        await bot.send_message(
            buyer_id,
            f"❌ *Kechirasiz!*\n\n"
            f"Sotuvchida bu mahsulot hozir mavjud emas.\n"
            f"Boshqa sotuvchilardan qidirishingiz mumkin:\n\n"
            f"👉 @XazdentBot → 🛍 Dental Market"
        )
    except Exception: pass
    await call.answer("❌ Rad etildi")

@router.callback_query(F.data.startswith("co_delivered_"))
async def catalog_order_delivered(call: CallbackQuery):
    """Xaridor yetib kelganini tasdiqladi."""
    parts    = call.data.split("_")
    order_id = int(parts[2])
    seller_id= int(parts[3])
    buyer_id = call.from_user.id

    await db_run(
        "UPDATE catalog_orders SET status='delivered', "
        "delivered_at=to_char(now(),'YYYY-MM-DD HH24:MI:SS') WHERE id=?",
        (order_id,)
    )
    await call.message.edit_reply_markup(reply_markup=None)

    # Baholash so'rovini yuboramiz
    rating_kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⭐", callback_data=f"rate_{order_id}_{seller_id}_1"),
            InlineKeyboardButton(text="⭐⭐", callback_data=f"rate_{order_id}_{seller_id}_2"),
            InlineKeyboardButton(text="⭐⭐⭐", callback_data=f"rate_{order_id}_{seller_id}_3"),
            InlineKeyboardButton(text="⭐⭐⭐⭐", callback_data=f"rate_{order_id}_{seller_id}_4"),
            InlineKeyboardButton(text="⭐⭐⭐⭐⭐", callback_data=f"rate_{order_id}_{seller_id}_5"),
        ]
    ])
    await call.message.answer(
        f"✅ *Ajoyib!*\n\nSotuvchini baholang:",
        reply_markup=rating_kb
    )
    await call.answer()

@router.callback_query(F.data.startswith("co_problem_"))
async def catalog_order_problem(call: CallbackQuery, state: FSMContext):
    """Xaridor muammo bildirdi."""
    parts    = call.data.split("_")
    order_id = int(parts[2])
    seller_id= int(parts[3])

    await state.set_state(ComplaintState.waiting_reason)
    await state.update_data(order_id=order_id, seller_id=seller_id)
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        "😔 *Muammoni tavsiflang:*\n\n"
        "_Masalan: Mahsulot kelmadi, sifat yomon, soni kam..._\n\n"
        "/cancel — bekor qilish"
    )
    await call.answer()

@router.message(ComplaintState.waiting_reason, F.text)
async def complaint_reason(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear(); await msg.answer("Bekor qilindi."); return

    d         = await state.get_data()
    order_id  = d.get("order_id")
    seller_id = d.get("seller_id")
    reason    = msg.text.strip()
    buyer_id  = msg.from_user.id
    u         = await get_user(buyer_id)
    uname     = (u["clinic_name"] or u["full_name"] if u else None) or str(buyer_id)

    # DB ga shikoyat
    await db_insert(
        "INSERT INTO complaints(from_user_id,against_user_id,reason) VALUES(?,?,?)",
        (buyer_id, seller_id, reason)
    )
    await db_run(
        "UPDATE catalog_orders SET status='disputed' WHERE id=?", (order_id,)
    )
    await state.clear()
    await msg.answer(
        "✅ *Shikoyatingiz qabul qilindi.*\n\n"
        "Admin 24 soat ichida ko\'rib chiqadi va\n"
        "siz bilan bog\'lanadi."
    )

    # Adminga xabar
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(
                aid,
                f"🚨 *Yangi shikoyat!*\n\n"
                f"👤 Xaridor: {uname} (ID: {buyer_id})\n"
                f"🏪 Sotuvchi ID: {seller_id}\n"
                f"📋 Buyurtma #{order_id}\n\n"
                f"📝 {reason}",
                reply_markup=ik(
                    [ib("👥 Users da ko\'rish", "admin_users_check")]
                )
            )
        except Exception: pass

@router.callback_query(F.data.startswith("rate_"))
async def rate_seller(call: CallbackQuery, state: FSMContext):
    """Xaridor yulduz berdi."""
    parts     = call.data.split("_")
    order_id  = int(parts[1])
    seller_id = int(parts[2])
    rating    = int(parts[3])
    buyer_id  = call.from_user.id

    # Avval baholagan bo'lsa qayta baholamasin
    existing = await db_get(
        "SELECT id FROM reviews WHERE order_id=? AND buyer_id=?",
        (order_id, buyer_id)
    )
    if existing:
        await call.answer("Allaqachon baholagansiz!", show_alert=True)
        return

    await state.set_state(ReviewState.waiting_comment)
    await state.update_data(order_id=order_id, seller_id=seller_id, rating=rating)

    stars = "⭐" * rating
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        f"Baho: {stars}\n\n"
        f"*Izoh qoldiring* (ixtiyoriy):\n"
        f"_Mahsulot sifati, yetkazish tezligi..._",
        reply_markup=ik([ib("➡️ Izohlarsiz yuborish", "review_skip")])
    )
    await call.answer()

@router.callback_query(F.data == "review_skip", ReviewState.waiting_comment)
async def review_skip(call: CallbackQuery, state: FSMContext):
    await _save_review(call.from_user.id, state, comment="")
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer("✅ Bahoingiz qabul qilindi! Rahmat!")
    await call.answer()

@router.message(ReviewState.waiting_comment, F.text)
async def review_comment(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear(); return
    await _save_review(msg.from_user.id, state, comment=msg.text.strip())
    await msg.answer("✅ Bahoingiz va izohingiz qabul qilindi! Rahmat!")

async def _save_review(buyer_id: int, state: FSMContext, comment: str):
    d         = await state.get_data()
    order_id  = d.get("order_id")
    seller_id = d.get("seller_id")
    rating    = d.get("rating", 5)
    await state.clear()

    await db_insert(
        "INSERT INTO reviews(order_id,buyer_id,seller_id,rating,comment) VALUES(?,?,?,?,?)",
        (order_id, buyer_id, seller_id, rating, comment)
    )
    # Do'kon reytingini yangilaymiz
    avg = await db_get(
        "SELECT AVG(rating) as avg FROM reviews WHERE seller_id=?", (seller_id,)
    )
    if avg and avg["avg"]:
        await db_run(
            "UPDATE shops SET rating=? WHERE owner_id=?",
            (round(float(avg["avg"]), 1), seller_id)
        )

# ── 48 SOAT CHECKER ───────────────────────────────────────────────────────────
async def delivery_checker():
    """Har soat ishlaydigan — 48 soat o'tgan buyurtmalarni tekshiradi."""
    while True:
        await asyncio.sleep(3600)  # Har soat
        try:
            # 48 soat o'tgan, hali notify yuborilmagan confirmed buyurtmalar
            cutoff = (datetime.now() - timedelta(hours=48)).isoformat()
            orders = await db_all(
                "SELECT * FROM catalog_orders "
                "WHERE status='confirmed' AND notify_sent=0 "
                "AND confirmed_at <= ?",
                (cutoff,)
            )
            for order in orders:
                try:
                    delivery_kb = InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(
                            text="✅ Ha, yetib keldi!",
                            callback_data=f"co_delivered_{order['id']}_{order['seller_id']}"
                        ),
                        InlineKeyboardButton(
                            text="❌ Yo'q, muammo bor",
                            callback_data=f"co_problem_{order['id']}_{order['seller_id']}"
                        )
                    ]])
                    # Mahsulotlar ro'yxati
                    import json as _pj
                    try:
                        items = _pj.loads(order["products_json"] or "[]")
                        prod_txt = ", ".join([f"{i['name']} x{i['qty']}" for i in items[:3]])
                    except Exception:
                        prod_txt = "mahsulotlar"

                    await bot.send_message(
                        order["buyer_id"],
                        f"📦 *Buyurtmangiz yetib keldimi?*\n\n"
                        f"_{prod_txt}_\n\n"
                        f"Iltimos, tasdiqlang:",
                        reply_markup=delivery_kb
                    )
                    await db_run(
                        "UPDATE catalog_orders SET notify_sent=1 WHERE id=?",
                        (order["id"],)
                    )
                    await asyncio.sleep(0.1)
                except Exception as e:
                    log.error(f"Delivery notify xato order {order['id']}: {e}")
        except Exception as e:
            log.error(f"delivery_checker xato: {e}")

# ── MAHSULOT ULASHISH VA TEZKOR BUYURTMA ─────────────────────────────────────

async def _show_product_start(msg: Message, pid: int):
    """Deep link orqali kelgan foydalanuvchiga mahsulotni ko'rsatish."""
    prod = await db_get(
        "SELECT p.*, s.shop_name, s.owner_id as seller_id, u.region "
        "FROM products p "
        "JOIN shops s ON p.shop_id=s.id "
        "JOIN users u ON s.owner_id=u.id "
        "WHERE p.id=? AND p.is_active=1",
        (pid,)
    )
    if not prod:
        await msg.answer("❌ Mahsulot topilmadi yoki o\'chirilgan.")
        return

    stars = ""
    avg = await db_get(
        "SELECT AVG(rating) as a, COUNT(*) as c FROM reviews WHERE seller_id=?",
        (prod["seller_id"],)
    )
    if avg and avg["a"]:
        rating = float(avg["a"])
        stars = "⭐" * round(rating) + f" ({rating:.1f})"

    txt = (
        f"🦷 *{prod['name']}*\n\n"
        f"💰 *{prod['price']:,.0f} so\'m / {prod['unit']}*\n"
        f"🏪 {prod['shop_name']}\n"
        f"📍 {prod['region'] or '—'}\n"
        + (f"⭐ {stars}\n" if stars else "") +
        (f"\n_{prod['description']}_" if prod.get("description") else "")
    )
    kb = ik(
        [ib("⚡ Tezkor buyurtma", f"qorder_{pid}_{prod['seller_id']}")],
        [ib("🔳 QR kod olish", f"get_qr_{pid}"),
         ib("📤 Ulashish", f"share_prod_{pid}")],
    )
    # Rasm bor bo'lsa
    photo = await db_get(
        "SELECT file_id FROM product_photos WHERE product_id=? ORDER BY sort_order LIMIT 1",
        (pid,)
    )
    if photo:
        try:
            await msg.answer_photo(photo["file_id"], caption=txt, reply_markup=kb)
            return
        except Exception: pass
    await msg.answer(txt, reply_markup=kb)

class QuickOrderState(StatesGroup):
    waiting_qty = State()

@router.callback_query(F.data.startswith("qorder_"))
async def quick_order_start(call: CallbackQuery, state: FSMContext):
    """Tezkor buyurtma — savatga solmasdan."""
    parts     = call.data.split("_")
    pid       = int(parts[1])
    seller_id = int(parts[2])
    uid       = call.from_user.id
    u         = await get_user(uid)

    if not u or u["role"] in (None, "none", ""):
        await call.answer("⚠️ Avval ro\'yxatdan o\'ting!", show_alert=True)
        return
    if u["role"] == "seller":
        await call.answer("⚠️ Sotuvchilar buyurtma bera olmaydi!", show_alert=True)
        return

    prod = await db_get("SELECT * FROM products WHERE id=?", (pid,))
    if not prod:
        await call.answer("Mahsulot topilmadi", show_alert=True)
        return

    await state.set_state(QuickOrderState.waiting_qty)
    await state.update_data(qo_pid=pid, qo_seller=seller_id,
                            qo_name=prod["name"], qo_price=prod["price"],
                            qo_unit=prod["unit"])
    await call.message.answer(
        f"⚡ *Tezkor buyurtma*\n\n"
        f"📦 *{prod['name']}*\n"
        f"💰 {prod['price']:,.0f} so\'m/{prod['unit']}\n\n"
        f"Nechta kerak?",
        reply_markup=ik(
            [ib("1", "qoq_1"), ib("2", "qoq_2"), ib("3", "qoq_3"),
             ib("5", "qoq_5"), ib("10", "qoq_10")],
            [ib("✏️ O\'zim kiritaman", "qoq_custom")],
            [ib("✖️ Bekor", "qoq_cancel")],
        )
    )
    await call.answer()

@router.callback_query(F.data.startswith("qoq_"), QuickOrderState.waiting_qty)
async def quick_order_qty(call: CallbackQuery, state: FSMContext):
    action = call.data[4:]
    if action == "cancel":
        await state.clear()
        await call.message.edit_text("❌ Bekor qilindi.")
        await call.answer()
        return
    if action == "custom":
        await call.message.answer(
            "Miqdorni kiriting (raqam):\n/cancel — bekor"
        )
        await call.answer()
        return
    qty = int(action)
    await _send_quick_order(call.message, state, qty, call.from_user.id)
    await call.answer()

@router.message(QuickOrderState.waiting_qty, F.text)
async def quick_order_qty_text(msg: Message, state: FSMContext):
    if msg.text and msg.text.startswith("/"):
        await state.clear()
        await msg.answer("❌ Bekor qilindi.")
        return
    try:
        qty = float(msg.text.replace(",", "."))
        if qty <= 0: raise ValueError
    except ValueError:
        await msg.answer("⚠️ To\'g\'ri raqam kiriting.")
        return
    await _send_quick_order(msg, state, qty, msg.from_user.id)

async def _send_quick_order(target_msg, state: FSMContext, qty: float, buyer_id: int):
    d         = await state.get_data()
    pid       = d.get("qo_pid")
    seller_id = d.get("qo_seller")
    name      = d.get("qo_name", "?")
    price     = d.get("qo_price", 0)
    unit      = d.get("qo_unit", "dona")
    total     = price * qty
    await state.clear()

    u      = await get_user(buyer_id)
    uname  = (u["clinic_name"] or u["full_name"] or str(buyer_id)) if u else str(buyer_id)
    uphone = u["phone"] if u else "—"
    uregion= u["region"] if u else "—"
    uaddr  = u["address"] if u else "—"

    # catalog_orders ga yozamiz
    import json as _pj
    items = [{"name": name, "qty": qty, "price": price, "unit": unit, "subtotal": total}]
    order_id = await db_insert(
        "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) VALUES(?,?,?,?)",
        (buyer_id, seller_id, _pj.dumps(items, ensure_ascii=False), total)
    )

    msg_txt = (
        f"⚡ *Tezkor buyurtma #{order_id}!*\n\n"
        f"📦 *{name}* — {qty} {unit}\n"
        f"💰 {price:,.0f} × {qty} = *{total:,.0f} so\'m*\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🏥 *{uname}*\n"
        f"📞 {uphone}\n"
        f"📍 {uregion}\n"
        f"🏠 {uaddr}"
    )
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Qabul qildim",
                             callback_data=f"co_confirm_{order_id}_{buyer_id}"),
        InlineKeyboardButton(text="❌ Mavjud emas",
                             callback_data=f"co_reject_{order_id}_{buyer_id}")
    ]])
    try:
        await bot.send_message(seller_id, msg_txt, reply_markup=confirm_kb)
    except Exception as e:
        log.error(f"Quick order notify xato: {e}")

    # Guruhga ham
    shop_g = await db_get("SELECT * FROM shops WHERE owner_id=?", (seller_id,))
    await _send_order_to_group(
        dict(shop_g) if shop_g else None,
        order_id, msg_txt, buyer_id, total
    )

    await target_msg.answer(
        f"✅ *Buyurtma #{order_id} yuborildi!*\n\n"
        f"📦 {name} — {qty} {unit}\n"
        f"💰 *{total:,.0f} so\'m*\n\n"
        f"Sotuvchi tez orada bog\'lanadi."
    )

# ── ARTIKUL VA QR KOD TIZIMI ─────────────────────────────────────────────────

async def _generate_article_code() -> str:
    """XD00001 dan boshlanadi, to'lganda avtomatik uzayadi."""
    row = await db_get(
        "SELECT article_code FROM products "
        "WHERE article_code IS NOT NULL "
        "ORDER BY LENGTH(article_code) DESC, article_code DESC LIMIT 1"
    )
    if row and row["article_code"]:
        try:
            last_num = int(row["article_code"][2:])
            next_num = last_num + 1
            # Minimum 5 xona, kerak bo'lsa uzayadi
            digits = max(5, len(str(next_num)))
            return f"XZ{next_num:0{digits}d}"
        except Exception:
            pass
    return "XZ00001"

def _generate_qr_bytes(url: str, label: str = "") -> bytes:
    """QR kod PNG — ichida artikul kodi yozilgan."""
    try:
        import qrcode
        import io
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)

        try:
            from PIL import Image, ImageDraw, ImageFont
            qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
            if label:
                qr_w, qr_h = qr_img.size
                txt_h = 52
                final = Image.new("RGB", (qr_w, qr_h + txt_h), "white")
                final.paste(qr_img, (0, 0))
                draw = ImageDraw.Draw(final)
                # Font qidirish
                font = None
                font_paths = [
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                    "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
                ]
                for fp in font_paths:
                    try:
                        font = ImageFont.truetype(fp, 26)
                        break
                    except Exception:
                        continue
                if font is None:
                    font = ImageFont.load_default()
                try:
                    bbox = draw.textbbox((0, 0), label, font=font)
                    txt_w = bbox[2] - bbox[0]
                except Exception:
                    txt_w = len(label) * 14
                x = max(0, (qr_w - txt_w) // 2)
                y = qr_h + 10
                draw.text((x, y), label, fill="#090979", font=font)
                # XAZDENT yozuvi
                try:
                    bbox2 = draw.textbbox((0, 0), "XAZDENT", font=font)
                    tw2 = bbox2[2] - bbox2[0]
                except Exception:
                    tw2 = 80
                draw.text(((qr_w - tw2)//2, qr_h - 2), "XAZDENT", fill="#444DCF", font=font)
            else:
                final = qr_img
        except Exception as pil_e:
            log.warning(f"PIL xato, oddiy QR: {pil_e}")
            qr_img = qr.make_image(fill_color="black", back_color="white")
            final = qr_img

        buf = io.BytesIO()
        final.save(buf, format="PNG")
        return buf.getvalue()
    except ImportError as e:
        raise Exception(f"qrcode kutubxonasi o'rnatilmagan: {e}")

async def _send_product_qr(user_id: int, product_id: int):
    """Foydalanuvchiga mahsulot QR kodini yuboradi."""
    prod = await db_get("SELECT * FROM products WHERE id=?", (product_id,))
    if not prod:
        return
    code = prod["article_code"] or f"p_{product_id}"
    url  = f"https://t.me/XazdentBot?start=xd_{code}"
    try:
        qr_bytes = _generate_qr_bytes(url, label=code)
        caption  = (
            f"🔳 *{prod['name']}* QR kodi\n\n"
            f"📌 Artikul: *{code}*\n"
            f"🔗 Havola: `{url}`\n\n"
            f"_Instagram postga joylashtiring — "
            f"mijozlar skan qilib buyurtma beradi_"
        )
        await bot.send_photo(
            user_id,
            BufferedInputFile(qr_bytes, filename=f"{code}_qr.png"),
            caption=caption
        )
    except Exception as e:
        log.error(f"QR yuborish xato: {e}")

# ── HASHTAG QIDIRISH (#XD00005 yoki XD00005) ─────────────────────────────────
@router.message(F.text.regexp(r'(?i)^#?X[DZ]\d+$'))
async def search_by_article(msg: Message):
    """XD00005 yoki #XD00005 yozilganda mahsulotni topadi."""
    code = msg.text.strip().lstrip("#").upper()
    # XD ham XZ sifatida qabul qilish
    if code.startswith("XD"): code = "XZ" + code[2:]
    prod = await db_get(
        "SELECT p.*, s.shop_name, s.owner_id as seller_id, u.region "
        "FROM products p "
        "JOIN shops s ON p.shop_id=s.id "
        "JOIN users u ON s.owner_id=u.id "
        "WHERE UPPER(p.article_code)=? AND p.is_active=1",
        (code,)
    )
    if not prod:
        await msg.answer(
            f"❌ *{code}* artikul topilmadi.\n\n"
            f"To\'g\'ri yozing: XD00005"
        )
        return
    await _show_product_start(msg, prod["id"])

# ── /start xd_XD00005 — QR skan ──────────────────────────────────────────────
# (bu _show_product_start ichida ishlaydi, start handlerda qo'shamiz)

# ── Sotuvchi QR buyrug'i ──────────────────────────────────────────────────────
@router.message(F.text.startswith("/qr_"))
async def cmd_qr(msg: Message):
    """Sotuvchi /qr_5 yozsa mahsulot QR kodini oladi."""
    if not msg.text: return
    try:
        pid = int(msg.text.split("_")[1])
    except Exception:
        await msg.answer("Format: /qr_5 (mahsulot ID)")
        return
    prod = await db_get("SELECT shop_id FROM products WHERE id=?", (pid,))
    if not prod: await msg.answer("❌ Topilmadi"); return
    shop = await db_get("SELECT owner_id FROM shops WHERE id=?", (prod["shop_id"],))
    if not shop or shop["owner_id"] != msg.from_user.id:
        await msg.answer("❌ Bu sizning mahsulotingiz emas"); return
    await msg.answer("⏳ QR kod tayyorlanmoqda...")
    await _send_product_qr(msg.from_user.id, pid)

@router.callback_query(F.data.startswith("get_qr_"))
async def cb_get_qr(call: CallbackQuery):
    """Mahsulot QR kodini botda yuboradi."""
    pid = int(call.data[7:])
    await call.answer("⏳ QR tayyorlanmoqda...")
    await _send_product_qr(call.from_user.id, pid)

@router.callback_query(F.data.startswith("share_prod_"))
async def cb_share_prod(call: CallbackQuery):
    """Mahsulot havolasi + QR ni ulashish."""
    pid  = int(call.data[11:])
    prod = await db_get("SELECT * FROM products WHERE id=?", (pid,))
    if not prod:
        await call.answer("Topilmadi", show_alert=True)
        return
    code = prod["article_code"] or f"p_{pid}"
    url  = f"https://t.me/XazdentBot?start=xd_{code}"

    # Havolani yuboradi
    await call.message.answer(
        f"📤 *Mahsulot havolasi:*\n\n"
        f"🦷 *{prod['name']}*\n"
        f"📌 Artikul: *{code}*\n"
        f"🔗 `{url}`\n\n"
        f"_Bu havolani do\'stlaringizga yuboring yoki\n"
        f"Instagram posta qo\'shing: #{code}_"
    )
    # QR ham yuboradi
    await _send_product_qr(call.from_user.id, pid)
    await call.answer()

# ── GURUH TIZIMI ─────────────────────────────────────────────────────────────

@router.message(F.text == "/setgroup")
async def cmd_setgroup(msg: Message):
    """Guruhda yozilganda shu guruhni do'konga bog'laydi."""
    if msg.chat.type not in ("group", "supergroup"):
        await msg.answer(
            "⚠️ Bu buyruq faqat guruhda ishlaydi.\n\n"
            "Qanday qilish:\n"
            "1. Guruh oching\n"
            "2. Botni guruhga qo'shing (@XazdentBot)\n"
            "3. Botni admin qiling\n"
            "4. Guruhda /setgroup yozing"
        )
        return
    uid      = msg.from_user.id
    group_id = msg.chat.id
    shop     = await db_get(
        "SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,)
    )
    if not shop:
        await msg.answer("❌ Sizning aktiv do'koningiz topilmadi.")
        return
    await db_run(
        "UPDATE shops SET group_chat_id=? WHERE owner_id=?",
        (group_id, uid)
    )
    await msg.answer(
        f"✅ *{shop['shop_name']}* do'koni shu guruhga bog'landi!\n\n"
        f"Bundan keyin barcha buyurtmalar shu guruhga ham chiqadi.\n"
        f"Guruh ID: `{group_id}`"
    )

@router.message(F.text == "/unsetgroup")
async def cmd_unsetgroup(msg: Message):
    """Guruh bog'liqligini o'chirish."""
    uid  = msg.from_user.id
    shop = await db_get("SELECT * FROM shops WHERE owner_id=?", (uid,))
    if not shop:
        await msg.answer("❌ Do'kon topilmadi.")
        return
    await db_run("UPDATE shops SET group_chat_id=NULL WHERE owner_id=?", (uid,))
    await msg.answer("✅ Guruh bog'liqligi o'chirildi.")

@router.callback_query(F.data.startswith("claim_"))
async def claim_order(call: CallbackQuery):
    """Guruh a'zosi buyurtmani o'z zimmasiga oladi."""
    parts    = call.data.split("_")
    order_id = int(parts[1])
    claimer  = call.from_user.id
    cname    = call.from_user.full_name or str(claimer)

    order = await db_get("SELECT * FROM catalog_orders WHERE id=?", (order_id,))
    if not order:
        await call.answer("Buyurtma topilmadi", show_alert=True)
        return
    if order["claimed_by"]:
        # Allaqachon birov olgan
        prev = await get_user(order["claimed_by"])
        pname = (prev["clinic_name"] or prev["full_name"] if prev else None) or "Boshqa xodim"
        await call.answer(f"❌ {pname} allaqachon olgan!", show_alert=True)
        return
    if order["status"] != "pending":
        await call.answer("Bu buyurtma allaqachon bajarilgan!", show_alert=True)
        return

    # Claim qilamiz
    await db_run(
        "UPDATE catalog_orders SET claimed_by=? WHERE id=? AND claimed_by IS NULL",
        (claimer, order_id)
    )
    # Tekshiramiz — race condition uchun
    updated = await db_get("SELECT claimed_by FROM catalog_orders WHERE id=?", (order_id,))
    if updated["claimed_by"] != claimer:
        prev = await get_user(updated["claimed_by"])
        pname = (prev["clinic_name"] or prev["full_name"] if prev else None) or "Boshqa xodim"
        await call.answer(f"❌ {pname} birozdan oldin oldi!", show_alert=True)
        return

    # Guruh xabarini yangilaymiz
    shop = await db_get("SELECT * FROM shops WHERE owner_id=?", (order["seller_id"],))
    shop_name = shop["shop_name"] if shop else "Do'kon"

    import json as _pj
    try:
        items = _pj.loads(order["products_json"] or "[]")
        prod_txt = "\n".join([
            "  • " + str(it.get("name","")) + " " + str(it.get("size","") or "") +
            " — " + str(it.get("qty",0)) + " " + str(it.get("unit","dona"))
            for it in items[:5]
        ])
    except Exception:
        prod_txt = "mahsulotlar"

    try:
        await call.message.edit_text(
            call.message.text + f"\n\n✅ *{cname}* qabul qildi!",
            reply_markup=None
        )
    except Exception: pass

    # Xaridorga kompaniya nomidan xabar
    buyer_id = order["buyer_id"]
    u = await get_user(buyer_id)
    uname = (u["clinic_name"] or u["full_name"] or str(buyer_id)) if u else str(buyer_id)

    try:
        await bot.send_message(
            buyer_id,
            f"✅ *Buyurtmangiz qabul qilindi!*\n\n"
            f"🏪 *{shop_name}* jamoasi\n"
            f"tez orada siz bilan bog\'lanadi.\n\n"
            f"📦 {prod_txt}"
        )
    except Exception as e:
        log.error(f"Claim buyer notify xato: {e}")

    # Guruh a'zosiga xaridor kontakti
    try:
        await bot.send_message(
            claimer,
            f"📋 *Buyurtma #{order_id} — Sizning zimmaingizda*\n\n"
            f"👤 Xaridor: *{uname}*\n"
            f"📞 {u['phone'] if u else '—'}\n"
            f"📍 {u['region'] if u else '—'}\n"
            f"🏠 {u['address'] if u else '—'}\n\n"
            f"📦 {prod_txt}\n\n"
            f"💰 Jami: *{order['total_amount']:,.0f} so\'m*\n\n"
            f"_Kompaniya nomidan muloqot qiling!_"
        )
    except Exception as e:
        log.error(f"Claim claimer notify xato: {e}")

    await call.answer(f"✅ Buyurtma #{order_id} sizda!")


async def _post_order_to_group(order_id: int, shop: dict,
                               products_txt: str, total: float,
                               buyer_name: str, buyer_region: str):
    """Buyurtmani do'kon guruhiga yuboradi."""
    group_id = shop.get("group_chat_id")
    if not group_id:
        return
    shop_name = shop.get("shop_name", "Do'kon")
    msg_txt = (
        f"🛒 *Yangi buyurtma #{order_id}!*\n\n"
        f"📦 {products_txt}\n\n"
        f"💰 *Jami: {total:,.0f} so\'m*\n\n"
        f"🏥 *{buyer_name}*\n"
        f"📍 {buyer_region}\n\n"
        f"_Kim qabul qiladi?_"
    )
    claim_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="✋ Men olaman",
            callback_data=f"claim_{order_id}"
        )
    ]])
    try:
        sent = await bot.send_message(group_id, msg_txt, reply_markup=claim_kb)
        await db_run(
            "UPDATE catalog_orders SET group_message_id=? WHERE id=?",
            (sent.message_id, order_id)
        )
    except Exception as e:
        log.error(f"Guruhga post xato (group={group_id}): {e}")

# ── GURUH TIZIMI ─────────────────────────────────────────────────────────────

@router.message(F.text == "/setgroup")
async def cmd_setgroup(msg: Message):
    """Guruhda yozilganda shu guruhni do'konga bog'laydi."""
    if msg.chat.type not in ("group", "supergroup"):
        await msg.answer(
            "⚠️ Bu buyruq faqat guruhda ishlaydi!\n\n"
            "1. Guruh oching\n"
            "2. Botni guruhga qo\'shing (@XazdentBot)\n"
            "3. Botni admin qiling\n"
            "4. Guruhda /setgroup yozing"
        )
        return

    uid      = msg.from_user.id
    group_id = msg.chat.id
    group_name = msg.chat.title or "Guruh"

    # Bu user ning do'koni bormi?
    shop = await db_get(
        "SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,))
    if not shop:
        await msg.answer(
            "❌ Sizning faol do\'koningiz topilmadi.\n"
            "Avval @XazdentBot da do\'kon oching."
        )
        return

    await db_run(
        "UPDATE shops SET group_chat_id=? WHERE owner_id=?",
        (group_id, uid)
    )
    await msg.answer(
        f"✅ *{group_name}* guruhi\n"
        f"🏪 *{shop['shop_name']}* do\'koniga bog\'landi!\n\n"
        f"Bundan keyin barcha buyurtmalar shu guruhga ham chiqadi.\n"
        f"_Buyurtmani qabul qilish uchun [✋ Men olaman] tugmasini bosing_"
    )

@router.message(F.text == "/removegroup")
async def cmd_removegroup(msg: Message):
    """Guruhni do'kondan ajratadi."""
    uid = msg.from_user.id
    shop = await db_get("SELECT * FROM shops WHERE owner_id=?", (uid,))
    if not shop:
        await msg.answer("❌ Do\'kon topilmadi")
        return
    await db_run("UPDATE shops SET group_chat_id=NULL WHERE owner_id=?", (uid,))
    await msg.answer("✅ Guruh bog\'liqlik o\'chirildi")

@router.callback_query(F.data.startswith("claim_order_"))
async def claim_order(call: CallbackQuery):
    """Guruh a'zosi buyurtmani o'z zimmasiga oladi."""
    order_id  = int(call.data[12:])
    claimer   = call.from_user.id
    claimer_name = call.from_user.full_name or "Xodim"

    # Allaqachon qabul qilinganmi?
    order = await db_get(
        "SELECT * FROM catalog_orders WHERE id=?", (order_id,))
    if not order:
        await call.answer("Buyurtma topilmadi", show_alert=True)
        return
    if order["claimed_by"]:
        await call.answer(
            "❌ Bu buyurtmani allaqachon boshqasi qabul qildi!",
            show_alert=True)
        return

    # Claim qilamiz
    await db_run(
        "UPDATE catalog_orders SET claimed_by=?, "
        "claimed_at=to_char(now(),'YYYY-MM-DD HH24:MI:SS') "
        "WHERE id=? AND claimed_by IS NULL",
        (claimer, order_id)
    )

    # Tekshiramiz — race condition bo'lmasin
    updated = await db_get(
        "SELECT claimed_by FROM catalog_orders WHERE id=?", (order_id,))
    if not updated or updated["claimed_by"] != claimer:
        await call.answer(
            "❌ Bir soniya kech qoldingiz, boshqasi oldi!",
            show_alert=True)
        return

    # Guruh xabarini yangilaymiz
    import json as _pj
    try:
        items = _pj.loads(order["products_json"] or "[]")
    except Exception:
        items = []

    lines_txt = ""
    for it in items:
        size = it.get("size") or it.get("variant") or ""
        name = it.get("name","?")
        qty  = it.get("qty", 1)
        unit = it.get("unit","dona")
        sub  = it.get("subtotal", 0)
        if size:
            lines_txt += f"• {name} ({size}) — {qty} {unit} · {sub:,.0f} so'm\n"
        else:
            lines_txt += f"• {name} — {qty} {unit} · {sub:,.0f} so'm\n"

    # Guruhga yangilangan xabar
    new_txt = (
        f"📦 *Buyurtma #{order_id}*\n\n"
        f"{lines_txt}\n"
        f"💰 *Jami: {order['total_amount']:,.0f} so'm*\n\n"
        f"✅ *{claimer_name}* qabul qildi!"
    )
    try:
        await call.message.edit_text(new_txt, reply_markup=None)
    except Exception:
        pass

    # Qabul qiluvchiga xaridor kontakti yuboramiz
    buyer  = await get_user(order["buyer_id"])
    shop   = await db_get(
        "SELECT * FROM shops WHERE owner_id=?", (order["seller_id"],))
    sname  = shop["shop_name"] if shop else "Do'kon"

    if buyer:
        uname   = buyer["clinic_name"] or buyer["full_name"] or str(order["buyer_id"])
        uphone  = buyer["phone"] or "—"
        uregion = buyer["region"] or "—"
        uaddr   = buyer["address"] or "—"

        contact_txt = (
            f"✅ *Buyurtma #{order_id} sizga biriktirildi!*\n\n"
            f"📦 *Buyurtma:*\n{lines_txt}\n"
            f"💰 *Jami: {order['total_amount']:,.0f} so'm*\n\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"👤 *Xaridor:* {uname}\n"
            f"📞 *Telefon:* {uphone}\n"
            f"📍 *Hudud:* {uregion}\n"
            f"🏠 *Manzil:* {uaddr}\n\n"
            f"_Iltimos, {sname} kompaniyasi nomidan bog\'laning_"
        )
        try:
            await bot.send_message(claimer, contact_txt)
        except Exception as e:
            log.error(f"Claim contact xato: {e}")

    # Xaridorga ham xabar — kompaniya nomi bilan
    try:
        await bot.send_message(
            order["buyer_id"],
            f"✅ *Buyurtmangiz #{order_id} qabul qilindi!*\n\n"
            f"🏪 *{sname}* xodimi tez orada siz bilan bog\'lanadi."
        )
    except Exception:
        pass

    await call.answer(f"✅ Qabul qildingiz! Xaridor kontakti yuborildi.")

@router.callback_query(F.data.startswith("co_partial_"))
async def catalog_order_partial(call: CallbackQuery, state: FSMContext):
    """Sotuvchi qisman qabul qildi."""
    parts    = call.data.split("_")
    order_id = int(parts[2])
    buyer_id = int(parts[3])
    seller   = call.from_user.id

    order = await db_get("SELECT * FROM catalog_orders WHERE id=?", (order_id,))
    if not order:
        await call.answer("Topilmadi", show_alert=True)
        return

    import json as _pj
    try:
        lines = _pj.loads(order["products_json"] or "[]")
    except Exception:
        lines = []

    await state.set_state(ComplaintState.waiting_reason)
    await state.update_data(
        partial_order_id=order_id,
        partial_buyer_id=buyer_id,
        partial_seller_id=seller
    )

    items_list = "\n".join([
        "  • " + str(it.get("size","?")) + " — " +
        str(int(it.get("qty",1))) + " " + str(it.get("unit","dona"))
        for it in lines
    ])
    await call.message.answer(
        f"⚠️ *Qisman qabul #{order_id}*\n\n"
        f"Buyurtma:\n{items_list}\n\n"
        f"Qaysi razmerlar yoki nechta yo\'qligini yozing:\n"
        f"_Masalan: 5510 razmer yo\'q, 4008 dan faqat 1 ta bor_\n\n"
        f"/cancel — bekor qilish"
    )
    await call.answer()

@router.message(ComplaintState.waiting_reason, F.text)
async def partial_or_complaint_text(msg: Message, state: FSMContext):
    """Qisman qabul yoki shikoyat matni."""
    if msg.text and msg.text.startswith("/"):
        await state.clear()
        await msg.answer("Bekor qilindi.")
        return

    d = await state.get_data()

    # Qisman qabul
    if d.get("partial_order_id"):
        order_id  = d["partial_order_id"]
        buyer_id  = d["partial_buyer_id"]
        seller_id = d["partial_seller_id"]
        reason    = msg.text.strip()
        await state.clear()

        await db_run(
            "UPDATE catalog_orders SET status='partial' WHERE id=?", (order_id,))

        u = await get_user(seller_id)
        shop = await db_get("SELECT shop_name FROM shops WHERE owner_id=?", (seller_id,))
        sname = (shop["shop_name"] if shop else None) or                 (u["clinic_name"] if u else None) or "Sotuvchi"

        # Xaridorga
        try:
            await bot.send_message(
                buyer_id,
                f"⚠️ *Buyurtma #{order_id} — Qisman qabul*\n\n"
                f"🏪 *{sname}* dan xabar:\n"
                f"_{reason}_\n\n"
                f"Sotuvchi siz bilan bog\'lanib aniqlashtiradi."
            )
        except Exception: pass

        await msg.answer(
            f"✅ Xaridorga qisman qabul haqida xabar yuborildi.\n"
            f"Buyurtma #{order_id} aktiv holatda qoldi.")
        return

    # Shikoyat
    order_id  = d.get("order_id")
    seller_id = d.get("seller_id")
    reason    = msg.text.strip()
    buyer_id  = msg.from_user.id
    u         = await get_user(buyer_id)
    uname     = (u["clinic_name"] or u["full_name"] or str(buyer_id)) if u else str(buyer_id)

    await db_insert(
        "INSERT INTO complaints(from_user_id,against_user_id,reason) VALUES(?,?,?)",
        (buyer_id, seller_id, reason))
    await db_run(
        "UPDATE catalog_orders SET status='disputed' WHERE id=?", (order_id,))
    await state.clear()
    await msg.answer(
        "✅ *Shikoyatingiz qabul qilindi.*\n\n"
        "Admin 24 soat ichida ko\'rib chiqadi.")

    for aid in ADMIN_IDS:
        try:
            await bot.send_message(
                aid,
                f"🚨 *Yangi shikoyat!*\n\n"
                f"👤 {uname} (ID: {buyer_id})\n"
                f"🏪 Sotuvchi ID: {seller_id}\n"
                f"📋 Buyurtma #{order_id}\n\n"
                f"📝 {reason}")
        except Exception: pass

# ── FALLBACK ─────────────────────────────────────────────────────────────────
@router.message()
async def fallback(msg: Message, state: FSMContext):
    current = await state.get_state()
    if current:
        return  # FSM davom etayotgan bo'lsa ignore
    u  = await get_user(msg.from_user.id)
    lg = (u["lang"] if u else None) or "uz"
    if u and u["role"] in ("clinic", "zubtex"):
        await msg.answer("🏥 *Klinika paneli*", reply_markup=kb_clinic(lg, uid=uid, webapp_url=WEBAPP_URL))
    elif u and u["role"] == "seller":
        uid2 = msg.from_user.id
        await msg.answer("🛒 *Sotuvchi paneli*", reply_markup=kb_seller(lg, uid=uid2, webapp_url=WEBAPP_URL))
    else:
        await msg.answer(t(lg, "welcome"), reply_markup=kb_lang())

# ── MAIN ─────────────────────────────────────────────────────────────────────
# ── WEB SERVER (Mini App uchun) ───────────────────────────────────────────────
async def handle_order_page(request):
    path = os.path.join(BASE_DIR, "webapp", "order.html")
    if not os.path.exists(path):
        return _web.Response(text="order.html topilmadi", status=404)
    return _web.FileResponse(path)

async def handle_api_needs(request):
    """GET /api/needs/{batch_id}"""
    try:
        batch_id = int(request.match_info.get("batch_id", 0))
    except Exception:
        batch_id = 0
    if batch_id <= 0:
        return _web.Response(
            text=_json.dumps([]),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin": "*"},
        )
    needs = await db_all(
        "SELECT id, product_name, quantity, unit FROM needs "
        "WHERE batch_id=? AND status != 'cancelled' ORDER BY id",
        (batch_id,),
    )
    log.info(f"API needs: batch={batch_id} -> {len(needs)} ta")
    data = [{"id": n["id"], "name": n["product_name"],
             "qty": n["quantity"], "unit": n["unit"]} for n in needs]
    return _web.Response(
        text=_json.dumps(data, ensure_ascii=False),
        content_type="application/json",
        headers={"Access-Control-Allow-Origin": "*"},
    )

async def handle_api_products(request):
    """GET /api/products/{uid} — klinikaning mahsulotlar ro'yxati"""
    uid      = int(request.match_info.get("uid", 0))
    products = await db_all(
        "SELECT name, unit FROM clinic_products WHERE owner_id=? ORDER BY sort_order, id",
        (uid,),
    )
    data = [{"name": p["name"], "unit": p["unit"]} for p in products]
    return _web.Response(
        text=_json.dumps(data, ensure_ascii=False),
        content_type="application/json",
        headers={"Access-Control-Allow-Origin": "*"},
    )

# offer.html inline (fayl topilmasa ham ishlaydi)
_OFFER_HTML_PATH = os.path.join(BASE_DIR, "webapp", "offer.html")

async def handle_offer_page(request):
    path = _OFFER_HTML_PATH
    if os.path.exists(path):
        return _web.FileResponse(path)
    # Fallback — inline HTML
    log.warning(f"offer.html topilmadi: {path}, inline version ishlatilmoqda")
    return _web.Response(
        text=_get_offer_html_inline(),
        content_type="text/html",
        charset="utf-8",
    )

def _get_offer_html_inline():
    """offer.html inline — fayl yo\'q bo\'lsa ishlatiladi."""
    return """<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>Narx kiriting</title>
<script src="https://telegram.org/js/telegram-web-app.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
  background:var(--tg-theme-bg-color,#fff);color:var(--tg-theme-text-color,#000);padding-bottom:90px}
.hdr{background:var(--tg-theme-secondary-bg-color,#f2f2f7);padding:14px 16px 10px;
  position:sticky;top:0;z-index:10;border-bottom:1px solid rgba(0,0,0,.08)}
.hdr h1{font-size:17px;font-weight:600}
.hdr p{font-size:12px;opacity:.55;margin-top:2px}
.pb{height:3px;background:var(--tg-theme-secondary-bg-color,#f2f2f7)}
.pf{height:100%;background:var(--tg-theme-button-color,#007aff);transition:width .3s;width:0%}
.item{border-bottom:1px solid rgba(0,0,0,.07);padding:14px 16px}
.item.unav{opacity:.4}
.irow{display:flex;align-items:center;gap:10px;margin-bottom:10px}
.num{width:28px;height:28px;border-radius:50%;
  background:var(--tg-theme-secondary-bg-color,#f2f2f7);
  color:var(--tg-theme-hint-color,#888);font-size:12px;font-weight:600;
  display:flex;align-items:center;justify-content:center;flex-shrink:0}
.num.done{background:#E2EFDA;color:#3B6D11}
.iname{font-size:15px;font-weight:600;flex:1}
.iqty{font-size:12px;color:var(--tg-theme-hint-color,#888);white-space:nowrap;
  background:var(--tg-theme-secondary-bg-color,#f2f2f7);padding:3px 8px;border-radius:8px}
.prow{display:flex;gap:8px;align-items:center;margin-bottom:6px}
.pwrap{flex:1;position:relative}
.pinput{width:100%;height:42px;border:1.5px solid rgba(0,0,0,.12);border-radius:10px;
  font-size:17px;font-weight:600;padding:0 52px 0 12px;
  background:var(--tg-theme-bg-color,#fff);color:var(--tg-theme-text-color,#000);
  outline:none;-webkit-appearance:none;appearance:none}
.pinput:focus{border-color:var(--tg-theme-button-color,#007aff)}
.pinput.ok{border-color:#3B6D11;background:#f8fff5}
.psuffix{position:absolute;right:10px;top:50%;transform:translateY(-50%);
  font-size:11px;color:var(--tg-theme-hint-color,#888);pointer-events:none}
.ubtn{height:42px;padding:0 12px;border:1.5px solid rgba(0,0,0,.10);border-radius:10px;
  font-size:12px;background:transparent;color:var(--tg-theme-hint-color,#888);
  cursor:pointer;white-space:nowrap;flex-shrink:0;-webkit-tap-highlight-color:transparent}
.ubtn.on{background:#FCEBEB;border-color:#E24B4A;color:#A32D2D}
.hint{font-size:12px;color:#3B6D11;font-weight:500;margin-bottom:6px;padding-left:2px}
.noteinp{width:100%;border:1px solid rgba(0,0,0,.08);border-radius:8px;
  font-size:13px;padding:7px 10px;background:var(--tg-theme-secondary-bg-color,#f2f2f7);
  color:var(--tg-theme-text-color,#000);outline:none;resize:none;font-family:inherit}
.dlrow{padding:12px 16px;border-bottom:1px solid rgba(0,0,0,.07)}
.dllabel{font-size:13px;opacity:.55;margin-bottom:8px}
.chips{display:flex;gap:8px;flex-wrap:wrap}
.chip{padding:7px 14px;border-radius:20px;border:1.5px solid rgba(0,0,0,.12);
  font-size:13px;font-weight:500;cursor:pointer;background:transparent;
  color:var(--tg-theme-text-color,#000);-webkit-tap-highlight-color:transparent}
.chip.sel{background:var(--tg-theme-button-color,#007aff);
  color:var(--tg-theme-button-text-color,#fff);border-color:transparent}
.bot{position:fixed;bottom:0;left:0;right:0;padding:10px 16px 14px;
  background:var(--tg-theme-bg-color,#fff);border-top:1px solid rgba(0,0,0,.08);z-index:20}
.binfo{display:flex;justify-content:space-between;font-size:13px;
  color:var(--tg-theme-hint-color,#888);margin-bottom:8px}
.btotal{font-weight:600;color:var(--tg-theme-button-color,#007aff)}
.sbtn{width:100%;padding:14px;border:none;border-radius:12px;
  background:var(--tg-theme-button-color,#007aff);
  color:var(--tg-theme-button-text-color,#fff);
  font-size:16px;font-weight:600;cursor:pointer;-webkit-tap-highlight-color:transparent}
.sbtn:active{opacity:.85}
.sbtn:disabled{opacity:.35;cursor:not-allowed}
.errmsg{padding:24px 16px;text-align:center;color:var(--tg-theme-hint-color,#888);
  font-size:14px;line-height:1.6}
.loading{display:flex;flex-direction:column;align-items:center;
  justify-content:center;padding:60px 16px;gap:14px;
  color:var(--tg-theme-hint-color,#888);font-size:14px}
.spin{width:30px;height:30px;border:3px solid var(--tg-theme-secondary-bg-color,#eee);
  border-top-color:var(--tg-theme-button-color,#007aff);
  border-radius:50%;animation:sp .7s linear infinite}
@keyframes sp{to{transform:rotate(360deg)}}
</style>
</head>
<body>

<div class="hdr">
  <h1>💰 Narx kiriting</h1>
  <p id="sub">Yuklanmoqda...</p>
</div>
<div class="pb"><div class="pf" id="pf"></div></div>

<div id="ct"><div class="loading"><div class="spin"></div>Yuklanmoqda...</div></div>

<div class="dlrow">
  <div class="dllabel">🚚 Yetkazib berish muddati:</div>
  <div class="chips">
    <button class="chip sel" data-val="2"   onclick="sDl(this)">⚡️ 2 soat</button>
    <button class="chip"     data-val="24"  onclick="sDl(this)">🕐 24 soat</button>
    <button class="chip"     data-val="48"  onclick="sDl(this)">📅 2 kun</button>
    <button class="chip"     data-val="168" onclick="sDl(this)">🗓 1 hafta</button>
  </div>
</div>

<div class="bot">
  <div class="binfo">
    <span id="fc">0/0 to'ldirildi</span>
    <span class="btotal" id="tt"></span>
  </div>
  <button class="sbtn" id="sb" onclick="doSend()" disabled>Narx kiriting</button>
</div>

<script>
var tg  = window.Telegram && window.Telegram.WebApp;
var dlv = 2;
var nds = [];
var bId = 0;

if (tg) { tg.ready(); tg.expand(); }

// batch_id ni URL dan olish
(function() {
  var path  = window.location.pathname;
  var parts = path.split('/');
  for (var i = 0; i < parts.length; i++) {
    if (parts[i] === 'offer' && parts[i+1]) {
      var n = parseInt(parts[i+1]);
      if (n > 0) { bId = n; break; }
    }
  }
  if (!bId) {
    var q = new URLSearchParams(window.location.search);
    bId = parseInt(q.get('batch_id') || q.get('bid') || '0') || 0;
  }
})();

// user_id
var userId = 0;
if (tg && tg.initDataUnsafe && tg.initDataUnsafe.user) {
  userId = tg.initDataUnsafe.user.id;
}

function esc(s) {
  return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── Yuklash ──────────────────────────────────────────────────────────────────
function loadNeeds() {
  if (!bId) {
    showErr('Buyurtma raqami topilmadi.\\nURL: ' + window.location.href);
    return;
  }
  fetch('/api/needs/' + bId)
    .then(function(r) {
      if (!r.ok) throw new Error('Server xatosi: ' + r.status);
      return r.json();
    })
    .then(function(data) {
      if (!data || data.length === 0) {
        showErr('Bu buyurtmada mahsulot topilmadi.\\n(batch #' + bId + ')');
        return;
      }
      nds = data;
      renderList();
    })
    .catch(function(e) {
      showErr('Yuklab bo\\'lmadi:\\n' + e.message);
    });
}

function showErr(msg) {
  document.getElementById('ct').innerHTML =
    '<div class="errmsg">' + msg.replace(/\\n/g,'<br>') + '</div>';
  document.getElementById('sub').textContent = 'Xato';
}

// ── Ro'yxatni ko'rsatish ─────────────────────────────────────────────────────
function renderList() {
  document.getElementById('sub').textContent = nds.length + ' ta mahsulot';
  var html = '';
  for (var i = 0; i < nds.length; i++) {
    var n = nds[i];
    html +=
      '<div class="item" id="row'+n.id+'">' +
        '<div class="irow">' +
          '<div class="num" id="num'+n.id+'">' + (i+1) + '</div>' +
          '<div class="iname">' + esc(n.name) + '</div>' +
          '<div class="iqty">' + n.qty + ' ' + n.unit + '</div>' +
        '</div>' +
        '<div class="prow">' +
          '<div class="pwrap">' +
            '<input class="pinput" id="p'+n.id+'" type="number" inputmode="decimal" ' +
              'placeholder="1 ' + n.unit + ' narxi..." min="0" step="any" ' +
              'oninput="onP('+n.id+','+n.qty+',\\''+n.unit+'\\')">' +
            '<span class="psuffix">so\\'m/' + n.unit + '</span>' +
          '</div>' +
          '<button class="ubtn" id="u'+n.id+'" onclick="togU('+n.id+')">' +
            'Mavjud emas' +
          '</button>' +
        '</div>' +
        '<div class="hint" id="h'+n.id+'" style="display:none"></div>' +
        '<textarea class="noteinp" id="nt'+n.id+'" rows="1" ' +
          'placeholder="Izoh (brend, sifat, muddati...)"></textarea>' +
      '</div>';
  }
  document.getElementById('ct').innerHTML = html;
  refreshBot();
}

// ── Narx o'zgarganda ─────────────────────────────────────────────────────────
function onP(id, qty, unit) {
  var inp = document.getElementById('p'+id);
  var hnt = document.getElementById('h'+id);
  var ub  = document.getElementById('u'+id);
  var v   = parseFloat(inp.value) || 0;

  if (v > 0) {
    inp.classList.add('ok');
    ub.classList.remove('on');
    ub.disabled = false;
    document.getElementById('row'+id).classList.remove('unav');
    if (hnt) {
      var total = v * qty;
      hnt.textContent = qty + ' ' + unit + ' × ' + fnum(v) + ' = ' + fnum(total) + ' so\\'m';
      hnt.style.display = 'block';
    }
  } else {
    inp.classList.remove('ok');
    if (hnt) hnt.style.display = 'none';
  }
  refreshBot();
  if (tg && tg.HapticFeedback && v > 0) tg.HapticFeedback.selectionChanged();
}

// ── Mavjud emas ───────────────────────────────────────────────────────────────
function togU(id) {
  var ub  = document.getElementById('u'+id);
  var inp = document.getElementById('p'+id);
  var row = document.getElementById('row'+id);
  var hnt = document.getElementById('h'+id);
  var on  = ub.classList.toggle('on');
  if (on) {
    inp.value = ''; inp.classList.remove('ok'); inp.disabled = true;
    row.classList.add('unav');
    if (hnt) hnt.style.display = 'none';
  } else {
    inp.disabled = false; row.classList.remove('unav');
    inp.focus();
  }
  refreshBot();
}

function fnum(n) {
  return Math.round(n).toLocaleString('uz-UZ');
}

// ── Pastki panel yangilash ────────────────────────────────────────────────────
function refreshBot() {
  var filled = 0;
  var total  = 0;
  for (var i = 0; i < nds.length; i++) {
    var n   = nds[i];
    var inp = document.getElementById('p'+n.id);
    var ub  = document.getElementById('u'+n.id);
    var num = document.getElementById('num'+n.id);
    var isU = ub && ub.classList.contains('on');
    var v   = inp ? (parseFloat(inp.value)||0) : 0;
    if (isU) {
      filled++;
      if (num) { num.textContent = '—'; num.classList.add('done'); }
    } else if (v > 0) {
      filled++;
      total += v * n.qty;
      if (num) num.classList.add('done');
    } else {
      if (num) { num.textContent = i+1; num.classList.remove('done'); }
    }
  }
  var pct = nds.length > 0 ? Math.round(filled/nds.length*100) : 0;
  document.getElementById('pf').style.width = pct + '%';
  document.getElementById('fc').textContent = filled + '/' + nds.length + ' to\\'ldirildi';
  var tt = document.getElementById('tt');
  tt.textContent = total > 0 ? 'Jami: ' + fnum(total) + ' so\\'m' : '';
  var sb = document.getElementById('sb');
  sb.disabled = filled === 0;
  sb.textContent = filled > 0 ? ('✅ ' + filled + ' ta taklif yuborish') : 'Narx kiriting';
}

function sDl(el) {
  document.querySelectorAll('.chip').forEach(function(c){ c.classList.remove('sel'); });
  el.classList.add('sel');
  dlv = parseInt(el.dataset.val);
}

// ── Yuborish ─────────────────────────────────────────────────────────────────
function doSend() {
  var offers = [];
  for (var i = 0; i < nds.length; i++) {
    var n   = nds[i];
    var inp = document.getElementById('p'+n.id);
    var ub  = document.getElementById('u'+n.id);
    var nt  = document.getElementById('nt'+n.id);
    var v   = inp ? (parseFloat(inp.value)||0) : 0;
    var isU = ub && ub.classList.contains('on');
    var note= nt ? (nt.value||'').trim() : '';
    if (isU) {
      offers.push({need_id:n.id, price:0, unavailable:true, note:note});
    } else if (v > 0) {
      offers.push({need_id:n.id, price:v, unavailable:false, note:note});
    }
  }
  if (!offers.length) return;

  var sb = document.getElementById('sb');
  sb.disabled = true;
  sb.textContent = '⏳ Yuklanmoqda...';

  fetch('/api/submit_offer', {
    method:  'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({
      payload:   JSON.stringify({type:'offer', batch_id:bId, offers:offers, delivery:dlv}),
      user_id:   userId,
      init_data: tg ? tg.initData : ''
    })
  })
  .then(function(r){ return r.json(); })
  .then(function(res){
    if (res.ok) {
      sb.textContent = '✅ Yuborildi!';
      if (tg) setTimeout(function(){ tg.close(); }, 1500);
    } else {
      sb.disabled = false;
      sb.textContent = '✅ Taklif yuborish';
      alert('Xato: ' + (res.error || 'nomalum'));
    }
  })
  .catch(function(e){
    sb.disabled = false;
    sb.textContent = '✅ Taklif yuborish';
    alert('Tarmoq xatosi: ' + e.message);
  });
}

loadNeeds();
</script>
</body>
</html>
"""


async def _accept_offers_handler(req):
    """POST /api/accept_offers"""
    try:
        body      = await req.json()
        offer_ids = body.get("offer_ids", [])
        batch_id  = int(body.get("batch_id", 0))
        user_id   = int(body.get("user_id", 0))
        if not offer_ids or not user_id:
            return _web.Response(
                text=_json.dumps({"ok": False, "error": "offer_ids yoki user_id yo'q"}),
                content_type="application/json")
        clinic = await get_user(user_id)
        if not clinic:
            return _web.Response(
                text=_json.dumps({"ok": False, "error": "foydalanuvchi topilmadi"}),
                content_type="application/json")

        accepted_by_seller = {}
        need_winners = {}

        for oid in offer_ids:
            o = await db_get("SELECT * FROM offers WHERE id=?", (oid,))
            if not o:
                continue
            await db_run("UPDATE offers SET status='accepted' WHERE id=?", (oid,))
            await db_run("UPDATE needs SET status='paused' WHERE id=?", (o["need_id"],))
            need_winners[o["need_id"]] = (oid, o["price"], o["seller_id"])
            accepted_by_seller.setdefault(o["seller_id"], []).append(o["need_id"])

        # G'olib sotuvchilarga: har biri uchun qabul qilingan mahsulotlar ro'yxati
        for sid, nids in accepted_by_seller.items():
            items = []
            for nid in nids:
                win_info = need_winners.get(nid)
                if not win_info:
                    continue
                win_oid, win_price, _ = win_info
                nd = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
                if nd:
                    items.append((nd["product_name"], nd["quantity"], nd["unit"], win_price))
            if items:
                await _notify_winner(sid, clinic, items)

        # Yutqazganlarga: faqat narx statistikasi
        for need_id, (win_oid, win_price, win_sid) in need_winners.items():
            nd = await db_get("SELECT * FROM needs WHERE id=?", (need_id,))
            if not nd:
                continue
            for loser in await db_all(
                "SELECT * FROM offers WHERE need_id=? AND seller_id!=? AND price>0",
                (need_id, win_sid)
            ):
                await _notify_loser(loser["seller_id"], nd["product_name"],
                                    win_price, loser["price"], nd["unit"])

        return _web.Response(
            text=_json.dumps({"ok": True, "accepted": len(offer_ids)}),
            content_type="application/json")
    except Exception as e:
        log.error(f"accept_offers xato: {e}")
        return _web.Response(
            text=_json.dumps({"ok": False, "error": str(e)}),
            content_type="application/json", status=500)

async def start_webserver():
    app = _web.Application(client_max_size=50*1024*1024)  # 50MB
    # ── GET /order ────────────────────────────────────────────────────
    app.router.add_get("/order", handle_order_page)
    app.router.add_get("/offer/{batch_id}", handle_offer_page)

    async def _help_page(req):
        path = os.path.join(BASE_DIR, "webapp", "help.html")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return _web.Response(text=f.read(), content_type="text/html", charset="utf-8")
        return _web.Response(text="help.html topilmadi", status=404, content_type="text/html")
    app.router.add_get("/help", _help_page)

    async def _photo_proxy(req):
        """Telegram file_id -> bytes -> browser."""
        file_id = req.match_info.get("file_id", "")
        if not file_id:
            return _web.Response(status=404)
        try:
            file = await bot.get_file(file_id)
            file_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file.file_path}"
            import aiohttp as _aiohttp
            async with _aiohttp.ClientSession() as session:
                async with session.get(file_url) as resp:
                    data = await resp.read()
                    ctype = resp.headers.get("Content-Type", "image/jpeg")
            return _web.Response(
                body=data, content_type=ctype,
                headers={
                    "Cache-Control": "public, max-age=86400",
                    "Access-Control-Allow-Origin": "*"
                })
        except Exception as e:
            log.error(f"Photo proxy xato: {e}")
            return _web.Response(status=404)
    app.router.add_get("/api/photo/{file_id}", _photo_proxy)

    async def _api_qr(req):
        """Mahsulot QR kodi PNG."""
        pid = int(req.match_info.get("product_id", 0))
        prod = await db_get("SELECT * FROM products WHERE id=?", (pid,))
        if not prod:
            return _web.Response(status=404)
        code = prod["article_code"] or f"p_{pid}"
        url  = f"https://t.me/XazdentBot?start=xd_{code}"
        try:
            qr_bytes = _generate_qr_bytes(url, label=code)
            return _web.Response(
                body=qr_bytes,
                content_type="image/png",
                headers={
                    "Cache-Control": "public, max-age=86400",
                    "Content-Disposition": f'inline; filename="{code}.png"',
                    "Access-Control-Allow-Origin": "*"
                })
        except Exception as e:
            log.error(f"QR API xato: {e}")
            return _web.Response(status=500)
    app.router.add_get("/api/catalog/qr/{product_id}", _api_qr)
    app.router.add_get("/api/products/{uid}", handle_api_products)

    # ── GET /api/needs/{batch_id} ──────────────────────────────────
    async def _api_needs(req):
        try: bid = int(req.match_info.get("batch_id", 0))
        except: bid = 0
        if bid <= 0:
            return _web.Response(text="[]", content_type="application/json",
                                 headers={"Access-Control-Allow-Origin": "*"})
        rows = await db_all(
            "SELECT id, product_name, quantity, unit FROM needs "
            "WHERE batch_id=? AND status != 'cancelled' ORDER BY id", (bid,))
        data = [{"id": r["id"], "name": r["product_name"],
                 "qty": r["quantity"], "unit": r["unit"]} for r in rows]
        log.info(f"API needs: batch={bid} -> {len(data)} ta")
        return _web.Response(text=_json.dumps(data, ensure_ascii=False),
                             content_type="application/json",
                             headers={"Access-Control-Allow-Origin": "*"})
    app.router.add_get("/api/needs/{batch_id}", _api_needs)

    # ── GET /compare/{batch_id} ────────────────────────────────────
    async def _compare_page(req):
        path = os.path.join(BASE_DIR, "webapp", "compare.html")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return _web.Response(text=f.read(), content_type="text/html", charset="utf-8")
        return _web.Response(text="compare.html topilmadi", status=404, content_type="text/html")
    app.router.add_get("/compare/{batch_id}", _compare_page)

    # ── GET /api/offers/{batch_id} ─────────────────────────────────
    async def _api_offers(req):
        try: bid = int(req.match_info.get("batch_id", 0))
        except: bid = 0
        rows = await db_all(
            "SELECT o.id, o.need_id, o.seller_id, o.price, o.unit, o.note, "
            "       COALESCE(s.shop_name, u.clinic_name, u.full_name, \'Sotuvchi\') as seller_name, "
            "       CASE WHEN o.note=\'mavjud_emas\' THEN 1 ELSE 0 END as unavail "
            "FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "LEFT JOIN shops s ON s.owner_id=o.seller_id AND s.status=\'active\' "
            "WHERE o.batch_id=? ORDER BY o.seller_id, o.need_id",
            (bid,))
        data = [{"id":r["id"],"need_id":r["need_id"],"seller_id":r["seller_id"],
                 "price":r["price"],"unit":r["unit"],"note":r["note"] or "",
                 "seller_name":r["seller_name"],"unavail":bool(r["unavail"])} for r in rows]
        return _web.Response(text=_json.dumps(data, ensure_ascii=False),
                             content_type="application/json",
                             headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/offers/{batch_id}", _api_offers)



    # ── POST /api/submit_order ─────────────────────────────────────
    async def _submit_order(req):
        try:
            body    = await req.json()
            payload = body.get("payload", "")
            user_id = body.get("user_id")
            if not payload or not user_id:
                return _web.Response(
                    text=_json.dumps({"ok": False, "error": "payload yoki user_id yo'q"}),
                    content_type="application/json")
            data     = _json.loads(payload)
            items    = data.get("items", [])
            deadline = int(data.get("deadline", 24))
            uid      = int(user_id)
            if not items:
                return _web.Response(
                    text=_json.dumps({"ok": False, "error": "items bo'sh"}),
                    content_type="application/json")
            u = await get_user(uid)
            if not u:
                return _web.Response(
                    text=_json.dumps({"ok": False, "error": "foydalanuvchi topilmadi"}),
                    content_type="application/json")
            room    = await get_or_create_room(uid)
            expires = (datetime.now() + timedelta(hours=deadline)).isoformat()
            batch_id = await db_insert(
                "INSERT INTO batches(owner_id,deadline_hours,expires_at) VALUES(?,?,?)",
                (uid, deadline, expires))
            saved = []
            for item in items:
                nid = await db_insert(
                    "INSERT INTO needs(batch_id,room_id,owner_id,product_name,quantity,unit,deadline_hours,expires_at) "
                    "VALUES(?,?,?,?,?,?,?,?)",
                    (batch_id, room["id"], uid, item["name"],
                     float(item["qty"]), item.get("unit","dona"), deadline, expires))
                need = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
                saved.append(dict(need))
            mid = await post_batch_to_channel(batch_id, saved, dict(u))
            if mid:
                for n in saved:
                    await db_run("UPDATE needs SET channel_message_id=? WHERE id=?", (mid, n["id"]))
            dl_map  = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
            preview = "\n".join([f"• {n['quantity']} {n['unit']} — {n['product_name']}" for n in saved[:5]])
            chan = CHANNEL_ID.lstrip("@") if isinstance(CHANNEL_ID, str) else str(CHANNEL_ID)
            link = f"\n[Kanalda ko'rish](https://t.me/{chan}/{mid})" if mid else ""
            try:
                await bot.send_message(uid,
                    f"✅ *{len(saved)} ta mahsulot joylashtirildi!*{link}\n\n"
                    f"{preview}\n\n⏱ {dl_map.get(deadline,str(deadline)+' soat')} ichida")
            except Exception as e:
                log.error(f"Buyurtmachiga xabar xato: {e}")
            asyncio.create_task(notify_sellers_batch(batch_id, uid))
            return _web.Response(
                text=_json.dumps({"ok": True, "batch_id": batch_id, "count": len(saved)}),
                content_type="application/json")
        except Exception as e:
            log.error(f"submit_order xato: {e}")
            return _web.Response(
                text=_json.dumps({"ok": False, "error": str(e)}),
                content_type="application/json", status=500)
    app.router.add_post("/api/submit_order", _submit_order)

    # ── POST /api/submit_offer ─────────────────────────────────────
    async def _submit_offer(req):
        try:
            body     = await req.json()
            payload  = body.get("payload", "")
            user_id  = body.get("user_id")
            if not payload or not user_id:
                return _web.Response(
                    text=_json.dumps({"ok": False, "error": "payload yoki user_id yo'q"}),
                    content_type="application/json")
            data     = _json.loads(payload)
            offers   = data.get("offers", [])
            batch_id = int(data.get("batch_id", 0))
            delivery = int(data.get("delivery", 24))
            uid      = int(user_id)
            u = await get_user(uid)
            if not u:
                return _web.Response(
                    text=_json.dumps({"ok": False, "error": "foydalanuvchi topilmadi"}),
                    content_type="application/json")
            shop  = await db_get("SELECT shop_name FROM shops WHERE owner_id=? AND status='active'", (uid,))
            sname = (shop["shop_name"] if shop else None) or u["clinic_name"] or u["full_name"] or "Sotuvchi"
            saved = 0
            unavail_c = 0
            for offer in offers:
                nid    = int(offer.get("need_id", 0))
                price  = float(offer.get("price", 0))
                is_unav= bool(offer.get("unavailable", False))
                note   = offer.get("note", "") or ""
                if not nid: continue
                nd = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
                if not nd: continue
                ex = await db_get("SELECT id FROM offers WHERE need_id=? AND seller_id=?", (nid, uid))
                if ex: continue
                if is_unav:
                    await db_insert(
                        "INSERT INTO offers(need_id,batch_id,seller_id,product_name,price,unit,delivery_hours,note) "
                        "VALUES(?,?,?,?,?,?,?,?)",
                        (nid, batch_id, uid, nd["product_name"], 0, nd["unit"], delivery, "mavjud_emas"))
                    unavail_c += 1
                elif price > 0:
                    await db_insert(
                        "INSERT INTO offers(need_id,batch_id,seller_id,product_name,price,unit,delivery_hours,note) "
                        "VALUES(?,?,?,?,?,?,?,?)",
                        (nid, batch_id, uid, nd["product_name"], price, nd["unit"], delivery, note))
                    saved += 1
            if saved > 0:
                owners = await db_all("SELECT DISTINCT owner_id FROM needs WHERE batch_id=?", (batch_id,))
                dl_map = {2:"2 soat",24:"24 soat",48:"2 kun",168:"1 hafta"}
                for row in owners:
                    try:
                        await bot.send_message(row["owner_id"],
                            f"📩 *Yangi taklif!*\n\n🏪 {sname}\n"
                            f"📦 {saved} ta mahsulotga narx berdi\n"
                            f"🚚 {dl_map.get(delivery,str(delivery)+' soat')} ichida",
                            reply_markup=ik([ib("📩 Takliflarni ko'rish", f"view_batch_{batch_id}")]))
                    except Exception: pass
            parts = []
            if saved: parts.append(f"{saved} ta narx")
            if unavail_c: parts.append(f"{unavail_c} ta mavjud emas")
            return _web.Response(
                text=_json.dumps({"ok": True, "result": " | ".join(parts) if parts else "0"}),
                content_type="application/json")
        except Exception as e:
            log.error(f"submit_offer xato: {e}")
            return _web.Response(
                text=_json.dumps({"ok": False, "error": str(e)}),
                content_type="application/json", status=500)
    app.router.add_post("/api/submit_offer", _submit_offer)
    app.router.add_post("/api/accept_offers", _accept_offers_handler)

    # ── KATALOG API ────────────────────────────────────────────────
    async def _catalog_page(req):
        path = os.path.join(BASE_DIR, "webapp", "catalog.html")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                html = f.read()
            return _web.Response(text=html, content_type="text/html", charset="utf-8")
        return _web.Response(text="catalog.html topilmadi", status=404, content_type="text/html")
    app.router.add_get("/catalog", _catalog_page)

    async def _api_catalog(req):
        cat  = req.query.get("cat", "")
        q    = req.query.get("q", "").lower().strip()
        params = []
        # Faqat aktiv do'kon va aktiv mahsulotlar
        # is_active NULL yoki 1 bo'lsa ko'rsatamiz
        where = "s.status='active' AND (p.is_active IS NULL OR p.is_active = 1)"
        if cat:
            where += " AND p.category_id=?"
            params.append(int(cat))
        if q:
            where += " AND (LOWER(p.name) LIKE ? OR LOWER(COALESCE(p.article_code,'')) LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%"])
        query = (
            "SELECT p.*, s.shop_name, s.owner_id as seller_id, u.region "
            "FROM products p "
            "JOIN shops s ON p.shop_id=s.id "
            "JOIN users u ON s.owner_id=u.id "
            "WHERE " + where +
            " ORDER BY p.id DESC LIMIT 80"
        )
        rows = await db_all(query, tuple(params))
        log.info(f"📋 Catalog API: cat={cat!r}, q={q!r}, natija={len(rows)} ta")
        cats = {1:"🦷 Terapevtik",2:"⚙️ Jarrohlik",3:"🔬 Zubtexnik",
                4:"🧪 Dezinfeksiya",5:"💡 Uskunalar",6:"📸 Rentgen",
                7:"🖥 CAD/CAM",8:"🦴 Implantlar",9:"💻 Stom Soft",10:"🎓 Kurslar"}
        data = []
        for r in rows:
            photos_rows = await db_all(
                "SELECT file_id FROM product_photos WHERE product_id=? ORDER BY sort_order",
                (r["id"],))
            vars_rows = await db_all(
                "SELECT size_name, article, stock, COALESCE(price,0) as price "
                "FROM product_variants WHERE product_id=? ORDER BY id",
                (r["id"],))
            data.append({
                "id": r["id"], "name": r["name"], "price": float(r["price"] or 0),
                "unit": r["unit"], "shop_name": r["shop_name"],
                "seller_id": r["seller_id"],
                "region": r["region"], "stock": r["stock"] or 0,
                "category": cats.get(r["category_id"] or 1,""),
                "category_id": r["category_id"] or 1,
                "description": r["description"] or "",
                "article_code": r.get("article_code") or "",
                "photos": [f"/api/photo/{p['file_id']}" for p in photos_rows],
                "variants": [{"size_name":v["size_name"],"article":v["article"] or "",
                               "stock":v["stock"] or 0,"price":float(v["price"] or 0)} for v in vars_rows]
            })
        return _web.Response(
            text=_json.dumps({"products": data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin": "*"})
    app.router.add_get("/api/catalog", _api_catalog)

    async def _api_catalog_my(req):
        uid  = int(req.query.get("uid", 0))
        shop = await db_get("SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,))
        if not shop:
            return _web.Response(
                text=_json.dumps({"products": [], "error": "do'kon yo'q"}),
                content_type="application/json",
                headers={"Access-Control-Allow-Origin": "*"})
        cats = {1:"🦷 Terapevtik",2:"⚙️ Jarrohlik",3:"🔬 Zubtexnik",
                4:"🧪 Dezinfeksiya",5:"💡 Uskunalar",6:"📸 Rentgen",
                7:"🖥 CAD/CAM",8:"🦴 Implantlar",9:"💻 Stom Soft",10:"🎓 Kurslar"}
        rows = await db_all(
            "SELECT * FROM products WHERE shop_id=? ORDER BY created_at DESC", (shop["id"],))
        data = []
        for r in rows:
            ph_rows = await db_all(
                "SELECT file_id FROM product_photos WHERE product_id=? ORDER BY sort_order",
                (r["id"],))
            vr_rows = await db_all(
                "SELECT size_name, article, stock FROM product_variants WHERE product_id=? ORDER BY id",
                (r["id"],))
            data.append({
                "id":r["id"],"name":r["name"],"price":r["price"],"unit":r["unit"],
                "stock":r["stock"] or 0,"is_active":r["is_active"],
                "category": cats.get(r["category_id"] or 1,""),
                "photos":[f"/api/photo/{p['file_id']}" for p in ph_rows],
                "variants":[{"size_name":v["size_name"],"article":v["article"],"stock":v["stock"]} for v in vr_rows]
            })
        return _web.Response(
            text=_json.dumps({"products": data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin": "*"})
    app.router.add_get("/api/catalog/my", _api_catalog_my)

    async def _api_catalog_product(req):
        pid  = int(req.match_info.get("product_id", 0))
        row  = await db_get(
            "SELECT p.*, s.shop_name, u.region "
            "FROM products p JOIN shops s ON p.shop_id=s.id "
            "JOIN users u ON s.owner_id=u.id WHERE p.id=?", (pid,))
        if not row:
            return _web.Response(
                text=_json.dumps({"ok":False,"error":"topilmadi"}),
                content_type="application/json")
        return _web.Response(
            text=_json.dumps({"ok":True,"product":{
                "id":row["id"],"name":row["name"],"price":row["price"],
                "unit":row["unit"],"description":row["description"],
                "shop_name":row["shop_name"],"region":row["region"],
                "stock":row["stock"] or 0,"photo":None
            }}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/catalog/product/{product_id}", _api_catalog_product)

    async def _api_add_product(req):
        try:
            body     = await req.json()
            uid      = int(body.get("user_id", 0))
            name     = body.get("name","").strip()
            price    = float(body.get("price", 0))
            cat_id   = int(body.get("category_id", 1))
            unit     = body.get("unit","dona")
            desc     = body.get("description","") or ""
            # JS dan "images" yoki "photos" key kelishi mumkin
            images = body.get("images") or body.get("photos") or []
            variants = body.get("variants", [])  # [{size_name,article,stock}]

            if not name or price <= 0:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"nom va narx kerak"}),
                    content_type="application/json")
            shop = await db_get("SELECT * FROM shops WHERE owner_id=? AND status='active'", (uid,))
            if not shop:
                u2 = await get_user(uid)
                if u2:
                    sname2 = u2.get("clinic_name") or u2.get("full_name") or "Do\'konim"
                    new_sid2 = await db_insert(
                        "INSERT INTO shops(owner_id,shop_name,category,phone,region,status) "
                        "VALUES(?,?,?,?,?,'active')",
                        (uid, sname2, "Stomatologiya",
                         u2.get("phone",""), u2.get("region",""))
                    )
                    shop = await db_get("SELECT * FROM shops WHERE id=?", (new_sid2,))
                    log.info(f"🏪 Auto-shop yaratildi uid={uid}")
                if not shop:
                    return _web.Response(
                        text=_json.dumps({
                            "ok": False,
                            "error": "Do\'kon topilmadi. Botni qayta oching: /start"
                        }),
                        content_type="application/json")

            # Asosiy mahsulot yozish
            # Avtomatik artikul kod
            art_code = await _generate_article_code()
            # Ko'p kategoriya bo'lsa birinchisini asosiy qilamiz
            categories = body.get("categories", [cat_id])
            if not categories: categories = [cat_id]
            main_cat = int(categories[0])
            pid = await db_insert(
                "INSERT INTO products(shop_id,name,price,unit,description,category_id,article_code,stock) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (shop["id"], name, price, unit, desc, main_cat, art_code, int(body.get("stock",0)))
            )

            # Rasmlar — Telegram ga yuborib file_id saqlaymiz
            import base64 as _b64
            first_photo_id = None
            all_photo_ids = []
            storage_target = uid

            log.info(f"📸 Rasm soni: {len(images)}, photos_count={body.get('photos_count',0)}")

            for i, img_b64 in enumerate(images[:5]):
                if not img_b64:
                    log.warning(f"⚠️ Rasm {i} bo'sh, o'tkazildi")
                    continue
                try:
                    # JS dan faqat base64 (header siz) keladi
                    # Lekin agar header bilan kelsa ham tozalaymiz
                    raw = img_b64
                    if "base64," in raw:
                        raw = raw.split("base64,")[1]
                    elif "," in raw and len(raw.split(",")[0]) < 60:
                        raw = raw.split(",")[1]
                    raw = raw.strip().replace(" ", "+")
                    # Padding
                    missing = len(raw) % 4
                    if missing:
                        raw += "=" * (4 - missing)
                    log.info(f"📸 Rasm {i}: raw uzunligi={len(raw)}")
                    img_bytes = _b64.b64decode(raw)
                    log.info(f"📸 Rasm {i}: bytes={len(img_bytes)}")
                    if len(img_bytes) < 500:
                        log.warning(f"⚠️ Rasm {i} juda kichik ({len(img_bytes)} bytes), o'tkazildi")
                        continue
                    buf = BufferedInputFile(img_bytes, filename=f"prod_{pid}_{i}.jpg")
                    sent = await bot.send_photo(
                        storage_target, buf,
                        caption=f"📦 #{art_code} — rasm {i+1}"
                    )
                    fid = sent.photo[-1].file_id
                    all_photo_ids.append(fid)
                    if not first_photo_id:
                        first_photo_id = fid
                        await db_run(
                            "UPDATE products SET photo_file_id=? WHERE id=?", (fid, pid))
                    await db_insert(
                        "INSERT INTO product_photos(product_id,file_id,sort_order) VALUES(?,?,?)",
                        (pid, fid, i)
                    )
                    log.info(f"✅ Rasm {i+1} saqlandi: {fid[:30]}...")
                except Exception as e:
                    log.error(f"❌ Photo upload xato ({i}): {e}", exc_info=True)

            # Variantlar — price bilan
            for v in variants:
                size = (v.get("size_name") or "").strip()
                if not size: continue
                vp = float(v.get("price") or 0)
                await db_insert(
                    "INSERT INTO product_variants(product_id,size_name,article,stock,price) "
                    "VALUES(?,?,?,?,?)",
                    (pid, size, v.get("article","") or "",
                     int(v.get("stock") or 0), vp)
                )

            # Kanalga post yuborish — barcha rasmlar bilan
            try:
                shop_info = await db_get("SELECT * FROM shops WHERE id=?", (shop["id"],))
                shop_name = shop_info["shop_name"] if shop_info else "?"
                region    = shop_info["region"] if shop_info else ""
                channel   = CHANNEL_ID if CHANNEL_ID else "@testxzd"

                # Bot username ni aniqlash
                try:
                    bot_info = await bot.get_me()
                    bot_username = bot_info.username
                except Exception:
                    bot_username = "XazdentBot"

                channel_kb = InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(
                        text="🛍 Ko\'rish va buyurtma →",
                        url=f"https://t.me/{bot_username}?start=xz_{art_code}"
                    )
                ]])
                caption = (
                    f"🆕 *Yangi mahsulot!*\n\n"
                    f"🦷 *{name}*\n"
                    f"📌 {art_code}\n"
                    f"💰 *{price:,.0f} so\'m/{unit}*\n"
                    f"🏪 {shop_name} · 📍 {region}\n\n"
                    f"👆 Ko\'rish va buyurtma berish:"
                )

                if len(all_photo_ids) > 1:
                    # Ko'p rasm — media group (bytes dan to'g'ridan yuborish)
                    from aiogram.types import InputMediaPhoto
                    # Avval rasmlarni bytes sifatida qayta yuklaymiz kanalga
                    media = []
                    import base64 as _b64ch
                    for idx, img_b64 in enumerate(images[:len(all_photo_ids)]):
                        try:
                            raw = img_b64
                            if "base64," in raw:
                                raw = raw.split("base64,")[1]
                            elif "," in raw and len(raw.split(",")[0]) < 50:
                                raw = raw.split(",")[1]
                            raw = raw.strip()
                            missing = len(raw) % 4
                            if missing:
                                raw += "=" * (4 - missing)
                            img_bytes = _b64ch.b64decode(raw)
                            buf = BufferedInputFile(img_bytes, filename=f"ch_{idx}.jpg")
                            if idx == 0:
                                media.append(InputMediaPhoto(
                                    media=buf, caption=caption, parse_mode="Markdown"))
                            else:
                                media.append(InputMediaPhoto(media=buf))
                        except Exception as me:
                            log.error(f"Media group rasm {idx} xato: {me}")
                    if media:
                        await bot.send_media_group(channel, media=media)
                        await bot.send_message(channel, "👆 Buyurtma berish uchun:",
                                              reply_markup=channel_kb)
                    else:
                        await bot.send_message(channel, caption, reply_markup=channel_kb)

                elif all_photo_ids:
                    # Bitta rasm — bytes dan yuborish
                    try:
                        import base64 as _b64ch2
                        raw = images[0]
                        if "base64," in raw:
                            raw = raw.split("base64,")[1]
                        elif "," in raw and len(raw.split(",")[0]) < 50:
                            raw = raw.split(",")[1]
                        raw = raw.strip()
                        missing = len(raw) % 4
                        if missing:
                            raw += "=" * (4 - missing)
                        img_bytes = _b64ch2.b64decode(raw)
                        buf = BufferedInputFile(img_bytes, filename="ch_0.jpg")
                        await bot.send_photo(channel, buf,
                                            caption=caption, reply_markup=channel_kb)
                    except Exception as se:
                        log.error(f"Single rasm kanal xato: {se}")
                        await bot.send_message(channel, caption, reply_markup=channel_kb)
                else:
                    # Rasmsiz
                    await bot.send_message(channel, caption, reply_markup=channel_kb)
            except Exception as ch_err:
                log.error(f"Kanal post xato: {ch_err}")

            return _web.Response(
                text=_json.dumps({"ok":True,"product_id":pid,"article_code":art_code}),
                content_type="application/json")
        except Exception as e:
            log.error(f"add_product xato: {e}")
            return _web.Response(
                text=_json.dumps({"ok":False,"error":str(e)}),
                content_type="application/json")
    app.router.add_post("/api/catalog/add_product", _api_add_product)

    async def _api_del_product(req):
        try:
            body = await req.json()
            uid  = int(body.get("user_id", 0))
            pid  = int(body.get("product_id", 0))
            shop = await db_get("SELECT id FROM shops WHERE owner_id=?", (uid,))
            if not shop: return _web.Response(
                text=_json.dumps({"ok":False}), content_type="application/json")
            await db_run("DELETE FROM products WHERE id=? AND shop_id=?", (pid, shop["id"]))
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/catalog/del_product", _api_del_product)

    async def _api_quick_order(req):
        try:
            body = await req.json()
            pid  = int(body.get("product_id", 0))
            uid  = int(body.get("user_id", 0))
            prod = await db_get(
                "SELECT p.*, s.owner_id as seller_id, s.shop_name "
                "FROM products p JOIN shops s ON p.shop_id=s.id WHERE p.id=?", (pid,))
            if not prod:
                return _web.Response(text=_json.dumps({"ok":False,"error":"topilmadi"}),
                                     content_type="application/json")
            u = await get_user(uid)
            uname = (u["clinic_name"] or u["full_name"] or str(uid)) if u else str(uid)
            # Sotuvchiga xabar
            try:
                await bot.send_message(
                    prod["seller_id"],
                    f"📦 *Katalogdan so\'rov!*\n\n"
                    f"🏥 {uname}\n"
                    f"📌 {prod['name']} — {prod['price']:,.0f} so\'m/{prod['unit']}\n\n"
                    f"📞 {u['phone'] if u else '—'}\n"
                    f"📍 {u['region'] if u else '—'}"
                )
            except Exception: pass
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/catalog/quick_order", _api_quick_order)

    async def _api_cart_order(req):
        """Savatdagi mahsulotlarni sotuvchilarga yuborish."""
        try:
            body = await req.json()
            uid  = int(body.get("user_id", 0))

            # JS dan: {items:[{product_id, seller_id, qty, price, name, unit, variant}], user_id}
            flat_items = body.get("items", [])

            if not flat_items or not uid:
                return _web.Response(
                    text=_json.dumps({"ok":False,
                        "error":f"Savat bo'sh yoki uid yo'q (uid={uid}, items={len(flat_items)})"}),
                    content_type="application/json")

            u = await get_user(uid)
            if not u:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"foydalanuvchi topilmadi"}),
                    content_type="application/json")

            uname   = u.get("clinic_name") or u.get("full_name") or str(uid)
            uphone  = u.get("phone") or "—"
            uregion = u.get("region") or "—"
            uaddr   = u.get("address") or "—"

            # Sotuvchi bo'yicha guruhlash
            seller_map = {}
            for it in flat_items:
                sid = int(it.get("seller_id") or 0)
                if not sid:
                    # seller_id yo'q bo'lsa product_id dan olamiz
                    prod_row = await db_get(
                        "SELECT s.owner_id as seller_id, s.shop_name "
                        "FROM products p JOIN shops s ON p.shop_id=s.id WHERE p.id=?",
                        (int(it.get("product_id",0)),))
                    if prod_row:
                        sid = prod_row["seller_id"]
                        it["shop_name"] = prod_row["shop_name"]
                if sid not in seller_map:
                    seller_map[sid] = []
                seller_map[sid].append(it)

            sent = 0
            import json as _pj
            for seller_id, items in seller_map.items():
                if not seller_id: continue
                # Xabar matni
                lines = []
                total = 0
                for i, item in enumerate(items, 1):
                    qty      = float(item.get("qty", 1) or 1)
                    price    = float(item.get("price", 0) or 0)
                    subtotal = price * qty
                    total   += subtotal
                    variant  = item.get("variant") or ""
                    name     = item.get("name","?")
                    unit     = item.get("unit","dona")
                    vstr     = f" ({variant})" if variant else ""
                    lines.append(
                        f"{i}. *{name}{vstr}*\n"
                        f"   {qty:.0f} {unit} × {price:,.0f} = *{subtotal:,.0f} so\'m*"
                    )
                items_txt = "\n".join(lines)
                # catalog_orders jadvaliga yozamiz
                import json as _pyjson
                order_id = await db_insert(
                    "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) "
                    "VALUES(?,?,?,?)",
                    (uid, seller_id, _pyjson.dumps(items, ensure_ascii=False), total)
                )
                msg = (
                    f"🛒 *Katalogdan yangi buyurtma!*\n\n"
                    f"📦 *Buyurtma #{order_id}:*\n{items_txt}\n\n"
                    f"💰 *Jami: {total:,.0f} so\'m*\n\n"
                    f"━━━━━━━━━━━━━━━━━━\n"
                    f"🏥 *{uname}*\n"
                    f"📞 {uphone}\n"
                    f"📍 {uregion}\n"
                    f"🏠 {uaddr}"
                )
                confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(
                        text="✅ Buyurtmani qabul qildim",
                        callback_data=f"co_confirm_{order_id}_{uid}"
                    ),
                    InlineKeyboardButton(
                        text="❌ Mavjud emas",
                        callback_data=f"co_reject_{order_id}_{uid}"
                    )
                ]])
                try:
                    await bot.send_message(seller_id, msg, reply_markup=confirm_kb)
                    sent += 1
                except Exception as e:
                    log.error(f"Cart order notify xato: {e}")
                # Guruhga ham yuborish
                shop_info = await db_get(
                    "SELECT * FROM shops WHERE owner_id=? AND status='active'", (seller_id,))
                if shop_info and shop_info.get("group_chat_id"):
                    import json as _pj2
                    try:
                        _items = _pj.loads(_json.dumps(items, ensure_ascii=False))
                        _prod_txt = "\n".join([
                            f"  • {it.get('name','')} — {it.get('qty',0)} {it.get('unit','dona')}"
                            for it in _items[:5]
                        ])
                    except Exception:
                        _prod_txt = "mahsulotlar"
                    asyncio.create_task(_post_order_to_group(
                        order_id=order_id,
                        shop=dict(shop_info),
                        products_txt=_prod_txt,
                        total=total,
                        buyer_name=uname,
                        buyer_region=uregion
                    ))
            return _web.Response(
                text=_json.dumps({"ok":True,"sent":sent}),
                content_type="application/json")
        except Exception as e:
            log.error(f"cart_order xato: {e}")
            return _web.Response(
                text=_json.dumps({"ok":False,"error":str(e)}),
                content_type="application/json")
    app.router.add_post("/api/catalog/cart_order", _api_cart_order)

    async def _api_quick_order_direct(req):
        """Tezkor buyurtma — ko'p variant va ko'p miqdor."""
        try:
            body = await req.json()
            pid  = int(body.get("product_id", 0))
            uid  = int(body.get("user_id", 0))

            if not pid or not uid:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"product_id va user_id kerak"}),
                    content_type="application/json")

            # items: [{size_name, qty}] — har bir variant
            items_raw = body.get("items", [])
            if not items_raw:
                qty = float(body.get("qty", 1) or 1)
                items_raw = [{"size_name": None, "qty": qty}]

            # Faqat qty > 0 bo'lganlarni olish
            items_raw = [
                it for it in items_raw
                if float(it.get("qty", 0) or 0) > 0
            ]
            if not items_raw:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"Kamida bitta variant miqdori kiriting"}),
                    content_type="application/json")

            prod = await db_get(
                "SELECT p.*, s.owner_id as seller_id, s.shop_name, s.id as shop_id "
                "FROM products p JOIN shops s ON p.shop_id=s.id WHERE p.id=?", (pid,))
            if not prod:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"Mahsulot topilmadi"}),
                    content_type="application/json")

            u       = await get_user(uid)
            uname   = ""
            if u:
                uname = u.get("clinic_name") or u.get("full_name") or str(uid)
            uname   = uname or str(uid)
            uphone  = (u.get("phone") or "—") if u else "—"
            uregion = (u.get("region") or "—") if u else "—"
            uaddr   = (u.get("address") or "—") if u else "—"

            import json as _pj

            # Stok tekshiruvi
            stock_warnings = []
            for it in items_raw:
                size = (it.get("size_name") or "").strip()
                qty  = float(it.get("qty", 1) or 1)
                if size:
                    vrow = await db_get(
                        "SELECT stock FROM product_variants "
                        "WHERE product_id=? AND size_name=?", (pid, size))
                    if vrow and vrow["stock"] and 0 < vrow["stock"] < qty:
                        stock_warnings.append(
                            f"{size}: {int(qty)} ta so'raldi, {vrow['stock']} ta bor")

            # Buyurtma satrlari
            order_lines = []
            total       = 0
            lines_txt   = ""
            for it in items_raw:
                size = (it.get("size_name") or "").strip()
                qty  = float(it.get("qty", 1) or 1)
                if qty <= 0: continue
                sub  = prod["price"] * qty
                total += sub
                order_lines.append({
                    "name":     prod["name"],
                    "size":     size,
                    "qty":      qty,
                    "price":    prod["price"],
                    "unit":     prod["unit"],
                    "subtotal": sub
                })
                line = f"  • *{size}* — {qty:.0f} {prod['unit']}" if size else                        f"  • {qty:.0f} {prod['unit']}"
                lines_txt += line + f" = {sub:,.0f} so\'m\n"

            if not order_lines:
                return _web.Response(
                    text=_json.dumps({"ok":False,"error":"Buyurtma bo'sh"}),
                    content_type="application/json")

            # DB ga yozish
            order_id = await db_insert(
                "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) "
                "VALUES(?,?,?,?)",
                (uid, prod["seller_id"],
                 _pj.dumps(order_lines, ensure_ascii=False), total)
            )

            # Stok ogohlantirish qo'shimcha
            warn_txt = ""
            if stock_warnings:
                warn_txt = "\n\n⚠️ *Diqqat:*\n" + "\n".join(
                    [f"  ⚡ {w}" for w in stock_warnings])

            # Sotuvchiga xabar
            total_dona = sum(float(it.get("qty",1) or 1) for it in items_raw)
            msg_txt = (
                f"⚡ *Tezkor buyurtma #{order_id}!*\n\n"
                f"📦 *{prod['name']}* "
                f"({len(order_lines)} xil razmer, {total_dona:.0f} dona):\n"
                f"{lines_txt}\n"
                f"💰 *Jami: {total:,.0f} so\'m*"
                f"{warn_txt}\n\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"🏥 *{uname}*\n"
                f"📞 {uphone}\n"
                f"📍 {uregion}\n"
                f"🏠 {uaddr}"
            )

            # Qisman qabul imkoniyati bilan
            confirm_kb = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Qabul qildim",
                        callback_data=f"co_confirm_{order_id}_{uid}"),
                    InlineKeyboardButton(
                        text="⚠️ Qisman",
                        callback_data=f"co_partial_{order_id}_{uid}"),
                ],
                [
                    InlineKeyboardButton(
                        text="❌ Mavjud emas",
                        callback_data=f"co_reject_{order_id}_{uid}"),
                ]
            ])

            try:
                await bot.send_message(
                    prod["seller_id"], msg_txt, reply_markup=confirm_kb)
            except Exception as e:
                log.error(f"Quick order notify xato: {e}")

            # Guruhga ham
            shop_info2 = await db_get(
                "SELECT * FROM shops WHERE owner_id=? AND status='active'",
                (prod["seller_id"],))
            if shop_info2 and shop_info2.get("group_chat_id"):
                clean_txt = lines_txt.replace("*","").replace("\'","'")
                asyncio.create_task(_post_order_to_group(
                    order_id=order_id,
                    shop=dict(shop_info2),
                    products_txt=clean_txt,
                    total=total,
                    buyer_name=uname,
                    buyer_region=uregion
                ))

            # Xaridorga tasdiqlash
            try:
                await bot.send_message(
                    uid,
                    f"✅ *Buyurtma #{order_id} yuborildi!*\n\n"
                    f"📦 *{prod['name']}* — {total_dona:.0f} dona\n"
                    f"💰 {total:,.0f} so\'m\n\n"
                    f"_Sotuvchi tez orada bog\'lanadi._"
                )
            except Exception: pass

            return _web.Response(
                text=_json.dumps({"ok":True,"order_id":order_id,
                                  "warnings": stock_warnings}),
                content_type="application/json",
                headers={"Access-Control-Allow-Origin":"*"})

        except Exception as e:
            log.error(f"quick_order_direct xato: {e}")
            return _web.Response(
                text=_json.dumps({"ok":False,"error":str(e)}),
                content_type="application/json")
    app.router.add_post("/api/catalog/quick_order_direct", _api_quick_order_direct)

    async def _admin_support_list(req):
        uid = int(req.query.get("uid", 0))
        if ADMIN_IDS and uid not in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        rows = await db_all(
            "SELECT s.*, COALESCE(u.clinic_name,u.full_name,CAST(s.user_id AS TEXT)) as uname, "
            "u.phone "
            "FROM support_messages s LEFT JOIN users u ON s.user_id=u.id "
            "ORDER BY s.created_at DESC LIMIT 50"
        )
        data = [{"id":r["id"],"user_id":r["user_id"],"name":r["uname"],"phone":r["phone"],
                 "message":r["message"],"admin_reply":r["admin_reply"],
                 "status":r["status"],"created_at":r["created_at"]} for r in rows]
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/support", _admin_support_list)

    async def _admin_support_reply(req):
        try:
            body   = await req.json()
            mid    = int(body.get("msg_id",0))
            uid    = int(body.get("user_id",0))
            reply  = body.get("reply","")
            admin  = int(body.get("admin_id",0))
            if admin not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            await db_run(
                "UPDATE support_messages SET admin_reply=?, status=\'replied\', "
                "replied_at=to_char(now(),\'YYYY-MM-DD HH24:MI:SS\'), admin_id=? WHERE id=?",
                (reply, admin, mid)
            )
            try:
                await bot.send_message(uid, f"💬 *Yordam xizmatidan javob*\n\n{reply}")
            except Exception: pass
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/support_reply", _admin_support_reply)

    # ── ADMIN API ─────────────────────────────────────────────────
    async def _admin_check(req):
        uid = int(req.query.get("uid", 0))
        return uid in ADMIN_IDS

    async def _admin_stats(req):
        uid    = int(req.query.get("uid", 0))
        period = req.query.get("period", "month")
        if ADMIN_IDS and uid not in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q","uid":uid,"admins":ADMIN_IDS}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        now = datetime.now()
        if period == "day":   since = now.strftime("%Y-%m-%d")
        elif period == "week": since = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        elif period == "month": since = now.strftime("%Y-%m")
        else: since = ""

        def p_filter(col="created_at"):
            if not since: return ""
            if period == "month": return f" AND {col} LIKE '{since}%'"
            return f" AND {col} >= '{since}'"

        total_u  = (await db_get("SELECT COUNT(*) as c FROM users"))["c"]
        new_p    = (await db_get(f"SELECT COUNT(*) as c FROM users WHERE 1=1{p_filter()}"))["c"]
        clinics  = (await db_get("SELECT COUNT(*) as c FROM users WHERE role IN ('clinic','zubtex')"))["c"]
        sellers  = (await db_get("SELECT COUNT(*) as c FROM users WHERE role='seller'"))["c"]
        total_n  = (await db_get("SELECT COUNT(*) as c FROM needs"))["c"]
        total_d  = (await db_get("SELECT COUNT(*) as c FROM offers WHERE status='accepted'"))["c"]
        period_n = (await db_get(f"SELECT COUNT(*) as c FROM needs WHERE 1=1{p_filter()}"))["c"]
        period_d = (await db_get(f"SELECT COUNT(*) as c FROM offers WHERE status='accepted'{p_filter()}"))["c"]
        no_off   = (await db_get(
            "SELECT COUNT(DISTINCT batch_id) as c FROM needs n "
            "WHERE status='active' AND NOT EXISTS "
            "(SELECT 1 FROM offers o WHERE o.need_id=n.id)"))["c"]
        rev_all  = await db_get("SELECT COALESCE(SUM(amount),0) as s FROM transactions WHERE status='confirmed'")
        rev_p    = await db_get(f"SELECT COALESCE(SUM(amount),0) as s FROM transactions WHERE status='confirmed'{p_filter()}")
        regions  = await db_all(
            "SELECT u.region, COUNT(DISTINCT u.id) as users, COUNT(n.id) as needs "
            "FROM users u LEFT JOIN needs n ON n.owner_id=u.id "
            "WHERE u.role IN ('clinic','zubtex') AND u.region IS NOT NULL "
            "GROUP BY u.region ORDER BY needs DESC LIMIT 10")
        return _web.Response(
            text=_json.dumps({"ok":True,"data":{
                "total_users":total_u,"new_period":new_p,
                "clinics":clinics,"sellers":sellers,
                "total_needs":total_n,"total_deals":total_d,
                "no_offers":no_off,
                "period":{"needs":period_n,"deals":period_d,
                          "revenue":float(rev_p["s"] if rev_p else 0)},
                "revenue":float(rev_all["s"] if rev_all else 0),
                "regions":[{"region":r["region"],"users":r["users"],"needs":r["needs"]} for r in regions],
            }}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/stats", _admin_stats)

    async def _admin_checks(req):
        uid = int(req.query.get("uid", 0))
        if ADMIN_IDS and uid not in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        txs = await db_all(
            "SELECT t.*, COALESCE(u.clinic_name,u.full_name) as name "
            "FROM transactions t JOIN users u ON t.user_id=u.id "
            "WHERE t.status='pending' AND t.type='topup' ORDER BY t.created_at DESC LIMIT 20"
        )
        data = []
        for tx in txs:
            data.append({"id":tx["id"],"user_id":tx["user_id"],"name":tx["name"],
                         "amount":tx["amount"],"balls":tx["balls"],
                         "created_at":tx["created_at"],"photo":None})
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/checks", _admin_checks)

    async def _admin_check_action(req):
        try:
            body     = await req.json()
            action   = body.get("action")
            tx_id    = int(body.get("tx_id",0))
            user_id  = int(body.get("user_id",0))
            admin_id = int(body.get("admin_id",0))
            if admin_id not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                     content_type="application/json")
            if action == "approve":
                balls = float(body.get("balls",0))
                await db_run("UPDATE transactions SET status='confirmed',confirmed_by=? WHERE id=?",
                             (admin_id, tx_id))
                await add_balance(user_id, balls)
                try: await bot.send_message(user_id, f"🎉 Hisobingizga *{balls:.1f} ball* qo\'shildi!")
                except Exception: pass
            elif action == "reject":
                await db_run("UPDATE transactions SET status='rejected',confirmed_by=? WHERE id=?",
                             (admin_id, tx_id))
                try: await bot.send_message(user_id, "❌ Chekingiz rad etildi.")
                except Exception: pass
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/check_action", _admin_check_action)

    async def _admin_users(req):
        uid    = int(req.query.get("uid", 0))
        filt   = req.query.get("filter", "all")
        if ADMIN_IDS and uid not in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        q_search = req.query.get("q", "").strip()
        query  = ("SELECT id, COALESCE(clinic_name,full_name,username) as name, "
                  "role, region, phone, balance, is_blocked, created_at "
                  "FROM users WHERE 1=1")
        params = []
        if filt == "clinic":    query += " AND role IN ('clinic','zubtex')"
        elif filt == "seller":  query += " AND role='seller'"
        elif filt == "zubtex":  query += " AND role='zubtex'"
        elif filt == "active":  query += " AND is_blocked=0 AND role!='none'"
        elif filt == "blocked": query += " AND is_blocked=1"
        if q_search:
            # ID bo'yicha yoki ism bo'yicha qidirish
            if q_search.isdigit():
                query += " AND id=?"
                params.append(int(q_search))
            else:
                query += (" AND (LOWER(COALESCE(clinic_name,'')) LIKE ? "
                          "OR LOWER(COALESCE(full_name,'')) LIKE ? "
                          "OR LOWER(COALESCE(username,'')) LIKE ? "
                          "OR LOWER(COALESCE(phone,'')) LIKE ?)")
                like = f"%{q_search.lower()}%"
                params.extend([like, like, like, like])
        query += " ORDER BY created_at DESC LIMIT 100"
        rows = await db_all(query, tuple(params))
        data = [{"id":r["id"],"name":r["name"],"role":r["role"],
                 "region":r["region"],"phone":r["phone"],
                 "balance":float(r["balance"] or 0),
                 "is_blocked":r["is_blocked"],
                 "created_at":r["created_at"]} for r in rows]
        return _web.Response(text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
                             content_type="application/json",
                             headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/users", _admin_users)

    async def _admin_settings(req):
        uid = int(req.query.get("uid", 0))
        if ADMIN_IDS and uid not in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        data = {
            "ball_price":  await get_setting("ball_price") or "1000",
            "card_number": await get_setting("card_number") or "",
            "elon_price":  await get_setting("elon_price") or "0",
        }
        return _web.Response(text=_json.dumps({"ok":True,"data":data}),
                             content_type="application/json",
                             headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/settings", _admin_settings)

    async def _admin_save_settings(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            for key in ["ball_price","card_number","elon_price"]:
                if key in body: await update_setting(key, str(body[key]))
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/save_settings", _admin_save_settings)

    async def _admin_broadcast_api(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            text  = body.get("text","")
            users = await db_all("SELECT id FROM users WHERE is_blocked=0")
            sent  = 0
            for u in users:
                try:
                    await bot.send_message(u["id"], text)
                    sent += 1
                    await asyncio.sleep(0.05)
                except Exception: pass
            return _web.Response(text=_json.dumps({"ok":True,"sent":sent}),
                                 content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/broadcast", _admin_broadcast_api)

    async def _admin_page(req):
        path = os.path.join(BASE_DIR, "webapp", "admin.html")
        if os.path.exists(path):
            return _web.FileResponse(path)
        return _web.Response(text="admin.html topilmadi", status=404)
    app.router.add_get("/admin", _admin_page)

    async def _site_page(req):
        path = os.path.join(BASE_DIR, "webapp", "site.html")
        if os.path.exists(path):
            return _web.FileResponse(path)
        return _web.Response(text="site.html topilmadi", status=404)
    app.router.add_get("/site", _site_page)
    app.router.add_get("/", _site_page)

    async def _admin_excel_dl(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text="ruxsat yo'q", status=403)
        path = await build_admin_excel()
        if not path:
            return _web.Response(text="Excel yaratib bo'lmadi", status=500)
        return _web.FileResponse(path, headers={
            "Content-Disposition": f"attachment; filename=xazdent_admin.xlsx"
        })
    app.router.add_get("/api/admin/excel", _admin_excel_dl)

    # ── MARKET ANALYTICS ──────────────────────────────────────────────────
    async def _admin_market(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}),
                                 content_type="application/json")
        total_p  = (await db_get("SELECT COUNT(*) as c FROM products WHERE is_active=1"))["c"]
        total_s  = (await db_get("SELECT COUNT(*) as c FROM shops WHERE status='active'"))["c"]
        now      = datetime.now()
        month    = now.strftime("%Y-%m")
        views_m  = (await db_get("SELECT COUNT(*) as c FROM product_views WHERE created_at LIKE ?", (f"{month}%",)))["c"]
        views_t  = (await db_get("SELECT COUNT(*) as c FROM product_views"))["c"]
        orders_m = (await db_get("SELECT COUNT(*) as c FROM needs WHERE created_at LIKE ?", (f"{month}%",)))["c"]
        orders_t = (await db_get("SELECT COUNT(*) as c FROM needs"))["c"]
        avg_p    = await db_get("SELECT AVG(price) as a FROM products WHERE is_active=1")
        no_photo = (await db_get("SELECT COUNT(*) as c FROM products WHERE photo_file_id IS NULL AND is_active=1"))["c"]
        no_stock = (await db_get("SELECT COUNT(*) as c FROM products WHERE stock=0 AND is_active=1"))["c"]
        top_viewed = await db_all(
            "SELECT p.name, s.shop_name, COUNT(v.id) as views "
            "FROM product_views v JOIN products p ON v.product_id=p.id "
            "JOIN shops s ON p.shop_id=s.id "
            "GROUP BY p.id ORDER BY views DESC LIMIT 10")
        top_searches = await db_all(
            "SELECT query, COUNT(*) as cnt FROM search_logs "
            "GROUP BY query ORDER BY cnt DESC LIMIT 10")
        top_ordered = await db_all(
            "SELECT p.name, s.shop_name, COUNT(o.id) as orders "
            "FROM offers o JOIN needs n ON o.need_id=n.id "
            "JOIN products p ON p.name=n.product_name "
            "JOIN shops s ON p.shop_id=s.id "
            "WHERE o.status='accepted' "
            "GROUP BY p.id ORDER BY orders DESC LIMIT 10")
        return _web.Response(
            text=_json.dumps({"ok":True,"data":{
                "total_products":total_p,"total_shops":total_s,
                "views_month":views_m,"views_total":views_t,
                "orders_month":orders_m,"orders_total":orders_t,
                "avg_price":float(avg_p["a"] if avg_p and avg_p["a"] else 0),
                "no_photo_count":no_photo,"no_stock_count":no_stock,
                "top_viewed":[{"name":r["name"],"shop_name":r["shop_name"],"views":r["views"]} for r in top_viewed],
                "top_searches":[{"query":r["query"],"cnt":r["cnt"]} for r in top_searches],
                "top_ordered":[{"name":r["name"],"shop_name":r["shop_name"],"orders":r["orders"]} for r in top_ordered],
            }}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/market", _admin_market)

    # ── E'LON YUBORISH ────────────────────────────────────────────────────
    async def _admin_announce(req):
        try:
            body       = await req.json()
            admin_id   = int(body.get("admin_id",0))
            if admin_id not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                     content_type="application/json")
            text       = body.get("text","")
            audiences  = body.get("audiences",[])
            regions    = body.get("regions",[])
            single_uid = int(body.get("single_uid",0))

            if not text:
                return _web.Response(text=_json.dumps({"ok":False,"error":"matn kerak"}),
                                     content_type="application/json")

            recipients = []
            if single_uid:
                recipients = [{"id": single_uid}]
            else:
                query = "SELECT id FROM users WHERE is_blocked=0"
                params = []
                if audiences:
                    ph = ",".join(["?" for _ in audiences])
                    query += f" AND role IN ({ph})"
                    params.extend(audiences)
                if regions:
                    ph = ",".join(["?" for _ in regions])
                    query += f" AND region IN ({ph})"
                    params.extend(regions)
                recipients = await db_all(query, tuple(params))

            sent = 0
            for u in recipients:
                try:
                    await bot.send_message(u["id"], text)
                    sent += 1
                    await asyncio.sleep(0.04)
                except Exception: pass

            # Tarixga saqlash
            await db_insert(
                "INSERT INTO transactions(user_id,amount,balls,type,note,status) VALUES(?,0,0,'announce',?,\'sent\')",
                (admin_id, f"sent:{sent}|text:{text[:100]}")
            )
            return _web.Response(
                text=_json.dumps({"ok":True,"sent":sent}),
                content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/announce", _admin_announce)

    async def _admin_announce_history(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
        rows = await db_all(
            "SELECT note, created_at FROM transactions "
            "WHERE type='announce' ORDER BY created_at DESC LIMIT 20")
        data = []
        for r in rows:
            note = r["note"] or ""
            sent = 0; text = ""
            for part in note.split("|"):
                if part.startswith("sent:"): sent = int(part[5:])
                if part.startswith("text:"): text = part[5:]
            data.append({"text":text,"sent_count":sent,"created_at":r["created_at"]})
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/announce_history", _admin_announce_history)

    # ── OBUNA ─────────────────────────────────────────────────────────────
    async def _admin_subscriptions(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
        now = datetime.now()
        subs = await db_all(
            "SELECT s.*, u.region, sh.shop_name, "
            "COALESCE(u.clinic_name,u.full_name) as name "
            "FROM subscriptions s "
            "JOIN users u ON s.user_id=u.id "
            "LEFT JOIN shops sh ON sh.owner_id=s.user_id AND sh.status='active' "
            "ORDER BY s.created_at DESC")
        data = []
        for s in subs:
            paid_until = s["paid_until"] or s["trial_ends_at"] or ""
            days_left  = 0
            if paid_until:
                try:
                    ends = datetime.fromisoformat(paid_until[:10])
                    days_left = (ends - now).days
                except Exception: pass
            data.append({
                "user_id":s["user_id"],"name":s["name"],"shop_name":s["shop_name"],
                "region":s["region"],"status":s["status"],
                "paid_until":paid_until,"days_left":days_left
            })
        stats = {
            "total":   len(data),
            "active":  sum(1 for d in data if d["status"]=="active"),
            "trial":   sum(1 for d in data if d["status"]=="trial"),
            "expired": sum(1 for d in data if d["status"]=="expired"),
        }
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data,"stats":stats}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/subscriptions", _admin_subscriptions)

    async def _admin_extend_sub(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            uid = int(body.get("target_uid",0))
            sub = await db_get("SELECT * FROM subscriptions WHERE user_id=?", (uid,))
            now = datetime.now()
            if sub and sub["paid_until"]:
                try: base = datetime.fromisoformat(sub["paid_until"][:10])
                except: base = now
            else:
                base = now
            new_until = (base + timedelta(days=30)).strftime("%Y-%m-%d")
            if sub:
                await db_run("UPDATE subscriptions SET paid_until=?,status='active' WHERE user_id=?",
                             (new_until, uid))
            else:
                await db_insert(
                    "INSERT INTO subscriptions(user_id,status,paid_until) VALUES(?,\'active\',?)",
                    (uid, new_until))
            try: await bot.send_message(uid, f"✅ Obunangiz {new_until} gacha uzaytirildi!")
            except Exception: pass
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/extend_sub", _admin_extend_sub)

    # ── SHIKOYATLAR ───────────────────────────────────────────────────────
    async def _admin_complaints(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
        status = req.query.get("status","new")
        rows = await db_all(
            "SELECT c.*, "
            "COALESCE(uf.clinic_name,uf.full_name) as from_name, "
            "COALESCE(ua.clinic_name,ua.full_name) as against_name "
            "FROM complaints c "
            "JOIN users uf ON c.from_user_id=uf.id "
            "JOIN users ua ON c.against_user_id=ua.id "
            "WHERE c.status=? ORDER BY c.created_at DESC", (status,))
        data = [{"id":r["id"],"from_user_id":r["from_user_id"],"against_user_id":r["against_user_id"],
                 "from_name":r["from_name"],"against_name":r["against_name"],
                 "reason":r["reason"],"status":r["status"],"created_at":r["created_at"]} for r in rows]
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/complaints", _admin_complaints)

    async def _admin_close_complaint(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            await db_run("UPDATE complaints SET status='closed' WHERE id=?",
                         (int(body.get("complaint_id",0)),))
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/close_complaint", _admin_close_complaint)

    # ── USER MSG & BLOCK ──────────────────────────────────────────────────
    async def _admin_msg_user(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            uid  = int(body.get("target_uid",0))
            text = body.get("text","")
            await bot.send_message(uid, text)
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/msg_user", _admin_msg_user)

    async def _admin_toggle_block(req):
        try:
            body  = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
            uid   = int(body.get("target_uid",0))
            block = int(body.get("block",0))
            await db_run("UPDATE users SET is_blocked=? WHERE id=?", (block, uid))
            if block:
                await db_run("UPDATE shops SET status='suspended' WHERE owner_id=?", (uid,))
            else:
                await db_run("UPDATE shops SET status='active' WHERE owner_id=?", (uid,))
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/toggle_block", _admin_toggle_block)

    async def _admin_user_stats(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}),
                                 content_type="application/json")
        now = datetime.now()
        d28 = (now - timedelta(days=28)).strftime("%Y-%m-%d")
        total  = (await db_get("SELECT COUNT(*) as c FROM users"))["c"]
        clinic = (await db_get("SELECT COUNT(*) as c FROM users WHERE role IN ('clinic','zubtex')"))["c"]
        seller = (await db_get("SELECT COUNT(*) as c FROM users WHERE role='seller'"))["c"]
        zubtex = (await db_get("SELECT COUNT(*) as c FROM users WHERE role='zubtex'"))["c"]
        new_28 = (await db_get("SELECT COUNT(*) as c FROM users WHERE created_at >= ?", (d28,)))["c"]
        regions = await db_all(
            "SELECT region, "
            "COUNT(*) as total, "
            "SUM(CASE WHEN role IN ('clinic','zubtex') THEN 1 ELSE 0 END) as clinic, "
            "SUM(CASE WHEN role='seller' THEN 1 ELSE 0 END) as seller, "
            "SUM(CASE WHEN role='zubtex' THEN 1 ELSE 0 END) as zubtex, "
            "SUM(CASE WHEN created_at >= ? THEN 1 ELSE 0 END) as new_28 "
            "FROM users WHERE region IS NOT NULL AND role != 'none' "
            "GROUP BY region ORDER BY total DESC LIMIT 15",
            (d28,)
        )
        return _web.Response(
            text=_json.dumps({"ok":True,"data":{
                "total":total,"clinic":clinic,"seller":seller,
                "zubtex":zubtex,"new_28":new_28,
                "regions":[{"region":r["region"],"total":r["total"],
                            "clinic":r["clinic"],"seller":r["seller"],
                            "zubtex":r["zubtex"],"new_28":r["new_28"]} for r in regions]
            }}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/user_stats", _admin_user_stats)

    # ── CATALOG ORDERS (admin) ─────────────────────────────────────────
    async def _admin_catalog_orders(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        filter_uid = req.query.get("uid_filter") or req.query.get("uid")
        # uid_filter — foydalanuvchi buyurtmalari uchun
        uid_filter = int(req.query.get("uid_filter", 0)) if req.query.get("uid_filter") else 0

        query = (
            "SELECT co.*, "
            "COALESCE(ub.clinic_name,ub.full_name,CAST(co.buyer_id AS TEXT)) as buyer_name, "
            "COALESCE(us.clinic_name,us.full_name,CAST(co.seller_id AS TEXT)) as seller_name "
            "FROM catalog_orders co "
            "LEFT JOIN users ub ON co.buyer_id=ub.id "
            "LEFT JOIN users us ON co.seller_id=us.id "
            "WHERE 1=1"
        )
        params = []
        if uid_filter:
            query += " AND (co.buyer_id=? OR co.seller_id=?)"
            params.extend([uid_filter, uid_filter])
        query += " ORDER BY co.created_at DESC LIMIT 50"
        rows = await db_all(query, tuple(params))
        import json as _pj
        data = []
        for r in rows:
            # Mahsulotlar qisqacha
            try:
                items = _pj.loads(r["products_json"] or "[]")
                products_short = ", ".join([
                    f"{it.get('name','?')} x{int(it.get('qty',1))}"
                    for it in items[:2]
                ])
                if len(items) > 2:
                    products_short += f" +{len(items)-2} ta"
            except Exception:
                products_short = "?"
            data.append({
                "id": r["id"],
                "buyer_name": r["buyer_name"],
                "seller_name": r["seller_name"],
                "buyer_id": r["buyer_id"],
                "seller_id": r["seller_id"],
                "products_short": products_short,
                "total_amount": r["total_amount"],
                "status": r["status"],
                "created_at": r["created_at"],
            })
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/catalog_orders", _admin_catalog_orders)

    # ── ADMIN PRODUCTS ────────────────────────────────────────────────
    async def _admin_products(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        rows = await db_all(
            "SELECT p.*, s.shop_name "
            "FROM products p JOIN shops s ON p.shop_id=s.id "
            "ORDER BY p.created_at DESC LIMIT 100"
        )
        data = [{
            "id": r["id"], "name": r["name"], "price": r["price"],
            "unit": r["unit"], "shop_name": r["shop_name"],
            "is_active": r["is_active"], "stock": r["stock"] or 0,
            "article_code": r.get("article_code") or "",
        } for r in rows]
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/products", _admin_products)

    # ── BLOCK/UNBLOCK PRODUCT ─────────────────────────────────────────
    async def _admin_block_product(req):
        try:
            body = await req.json()
            if int(body.get("admin_id",0)) not in ADMIN_IDS:
                return _web.Response(text=_json.dumps({"ok":False,"error":"ruxsat yo'q"}),
                                     content_type="application/json")
            pid   = int(body.get("product_id",0))
            block = int(body.get("block",0))
            await db_run("UPDATE products SET is_active=? WHERE id=?", (0 if block else 1, pid))
            return _web.Response(text=_json.dumps({"ok":True}), content_type="application/json")
        except Exception as e:
            return _web.Response(text=_json.dumps({"ok":False,"error":str(e)}),
                                 content_type="application/json")
    app.router.add_post("/api/admin/block_product", _admin_block_product)

    # ── USER BALANCE ──────────────────────────────────────────────────
    async def _admin_user_balance(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}),
                                 content_type="application/json",
                                 headers={"Access-Control-Allow-Origin":"*"})
        target = int(req.query.get("uid_target", req.query.get("uid",0)))
        u = await get_user(target)
        total_paid = await db_get(
            "SELECT COALESCE(SUM(amount),0) as s FROM transactions "
            "WHERE user_id=? AND status='confirmed' AND type='topup'", (target,))
        return _web.Response(
            text=_json.dumps({
                "ok": True,
                "balance": float(u["balance"] or 0) if u else 0,
                "total_paid": float(total_paid["s"] if total_paid else 0),
            }),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/user_balance", _admin_user_balance)

    # ── ADMIN STATS — PERIOD SUPPORT ──────────────────────────────────────
    async def _admin_orders(req):
        if not int(req.query.get("uid",0)) in ADMIN_IDS:
            return _web.Response(text=_json.dumps({"ok":False}), content_type="application/json")
        filt = req.query.get("filter","all")
        query = (
            "SELECT b.id, b.created_at, "
            "COALESCE(u.clinic_name,u.full_name) as clinic_name, u.region, "
            "COUNT(DISTINCT n.id) as prod_count, "
            "COUNT(DISTINCT o.id) as offer_count, "
            "b.expires_at, b.status "
            "FROM batches b "
            "JOIN users u ON b.owner_id=u.id "
            "LEFT JOIN needs n ON n.batch_id=b.id "
            "LEFT JOIN offers o ON o.need_id=n.id "
            "WHERE 1=1"
        )
        if filt == "active":   query += " AND b.status='active'"
        if filt == "done":     query += " AND b.status='done'"
        if filt == "no_offers": query += " AND b.status='active'"
        query += " GROUP BY b.id ORDER BY b.created_at DESC LIMIT 50"
        rows = await db_all(query)
        now = datetime.now()
        data = []
        for r in rows:
            hours_left = 0
            if r["expires_at"]:
                try:
                    exp = datetime.fromisoformat(r["expires_at"][:19])
                    hours_left = (exp - now).total_seconds() / 3600
                except Exception: pass
            products_str = ""
            prods = await db_all("SELECT product_name,quantity,unit FROM needs WHERE batch_id=? LIMIT 3", (r["id"],))
            products_str = ", ".join([f"{p['product_name']} {p['quantity']}{p['unit']}" for p in prods])
            if filt == "no_offers" and r["offer_count"] > 0: continue
            data.append({
                "id":r["id"],"clinic_name":r["clinic_name"],"region":r["region"],
                "products":products_str,"offer_count":r["offer_count"],
                "hours_left":hours_left,"status":r["status"]
            })
        return _web.Response(
            text=_json.dumps({"ok":True,"data":data}, ensure_ascii=False),
            content_type="application/json",
            headers={"Access-Control-Allow-Origin":"*"})
    app.router.add_get("/api/admin/orders", _admin_orders)
    webapp_dir = os.path.join(BASE_DIR, "webapp")
    if os.path.isdir(webapp_dir):
        app.router.add_static("/static", webapp_dir)
    runner = _web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", 8080))
    site = _web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    log.info(f"🌐 Web server: http://0.0.0.0:{port}")

async def expire_checker():
    """Har 15 daqiqada muddati o'tgan ehtiyojlarni yopadi."""
    while True:
        try:
            now = datetime.now().isoformat()
            expired = await db_all(
                "SELECT * FROM needs WHERE status='active' AND expires_at IS NOT NULL AND expires_at < ?",
                (now,),
            )
            for n in expired:
                await db_run("UPDATE needs SET status='done' WHERE id=?", (n["id"],))
                # Kanal postini o'chirishga harakat
                if n.get("channel_message_id"):
                    try:
                        await bot.delete_message(CHANNEL_ID, n["channel_message_id"])
                    except Exception:
                        pass
            if expired:
                log.info(f"⏰ {len(expired)} ta ehtiyoj muddati tugadi")
        except Exception as e:
            log.error(f"Expire checker xato: {e}")
        await asyncio.sleep(900)  # 15 daqiqa

async def main():
    await init_db()
    dp.include_router(router)
    log.info("🦷 XAZDENT Bot ishga tushdi!")
    await start_webserver()
    asyncio.create_task(expire_checker())
    asyncio.create_task(delivery_checker())
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())
